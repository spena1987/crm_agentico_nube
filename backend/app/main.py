import os
import re
import time
import asyncio
import logging
from concurrent.futures import ThreadPoolExecutor
from typing import List, Dict, Any, Optional
from fastapi import FastAPI, HTTPException, Body, UploadFile, File, Form, BackgroundTasks, Query, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel
from contextlib import asynccontextmanager

from fastapi.responses import Response, StreamingResponse, FileResponse
import io
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from app.db import (
    supabase, 
    actualizar_bot_disabled,
    archivar_conversacion,
    obtener_metricas_conversaciones,
    obtener_conversaciones,
    obtener_mensajes_conversacion,
    marcar_mensajes_conversacion_leidos,
    is_lid_number,
    get_paciente_by_lid,
    get_paciente_by_nombre_aproximado,
    get_ultimo_paciente_activo,
    get_paciente_by_telefono,
    crear_paciente,
    get_or_create_conversacion,
    guardar_mensaje,
    get_paciente_by_dni, 
    get_paciente_by_geclisa_id, 
    crear_o_actualizar_paciente_geclisa,
    vincular_o_fusionar_paciente_con_geclisa,
    registrar_dni_paciente_nuevo_crm,
    asignar_medico_paciente,
    list_nomencladores,
    get_nomenclador_by_id,
    create_nomenclador,
    update_nomenclador,
    delete_nomenclador,
    list_practicas_con_arancel,
    create_or_update_practica,
    delete_practica,
    upsert_arancel_practica,
    buscar_practicas_presupuesto,
    bulk_import_practicas_aranceles,
    guardar_transcripcion_mensaje,
    get_asesorias_by_paciente,
    get_asesorias_confirmadas_pendientes,
    crear_asesoria_quirurgica,
    actualizar_asesoria_quirurgica,
    eliminar_asesoria_quirurgica,
    get_presupuestos_by_paciente,
    cambiar_estado_presupuesto,
    vincular_presupuesto_a_asesoria,
    eliminar_presupuesto,
    crear_presupuesto_rapido,
    get_evoluciones_by_asesoria,
    crear_evolucion_asesoria,
    eliminar_evolucion_asesoria,
    get_paciente_contexto_360,
    get_configuracion_quirurgica,
    actualizar_configuracion_quirurgica,
    get_pipeline_quirurgico,
    enriquecer_practicas_geclisa_con_crm,
    guardar_practica_crm_con_arancel,
    guardar_practica_crm_integral,
    listar_catalogo_completo_crm,
    eliminar_practica_crm,
    get_plantillas_preparaciones,
    get_plantilla_preparacion_by_id,
    create_plantilla_preparacion,
    update_plantilla_preparacion,
    delete_plantilla_preparacion,
    get_plantillas_consentimientos,
    get_plantilla_consentimiento_by_id,
    create_plantilla_consentimiento,
    update_plantilla_consentimiento,
    delete_plantilla_consentimiento,
    get_aranceles_por_practica,
    crear_arancel_practica,
    actualizar_arancel,
    eliminar_arancel,
    get_practica_resumen_operativo,
    render_consent_template,
    generar_mensaje_ameno_presupuesto,
    enviar_presupuesto_por_whatsapp,
    get_configuracion_quirofano,
    actualizar_configuracion_quirofano,
    get_quirofanos,
    crear_quirofano,
    actualizar_quirofano,
    eliminar_quirofano,
    get_quirofano_bloques,
    crear_quirofano_bloque,
    eliminar_quirofano_bloque,
    get_quirofano_bloqueos,
    crear_quirofano_bloqueo,
    eliminar_quirofano_bloqueo,
    get_turnos_quirofano,
    get_turno_quirofano_by_id,
    crear_turno_quirofano,
    actualizar_turno_quirofano,
    eliminar_turno_quirofano,
    get_consentimiento_by_token,
    registrar_firma_consentimiento,
    get_prestadores,
    crear_prestador,
    actualizar_prestador,
    eliminar_prestador,
    get_modelos_lio,
    get_modelo_lio_por_id,
    crear_modelo_lio,
    actualizar_modelo_lio,
    eliminar_modelo_lio,
    get_modelos_lio_items,
    get_all_modelos_lio_items,
    validar_gtin_unico,
    sincronizar_lentes_masivos,
    crear_modelo_lio_item,
    eliminar_modelo_lio_item,
    resolver_sku_lio,
    get_catalogo_maestro,
    crear_item_catalogo_maestro,
    actualizar_item_catalogo_maestro,
    eliminar_item_catalogo_maestro,
    cambiar_estado_turno_quirofano,
    subir_consentimiento_turno_a_geclisa,
    subir_parte_quirurgico_turno_a_geclisa,
    desvincular_documento_geclisa_turno,
    obtener_datos_pulsera_turno,
    marcar_pulsera_impresa,
    procesar_escaneo_qr_turno
)
from app.agent import procesar_mensaje_agente, transcribir_audio_con_gemini
from app.services.copilot_service import (
    sugerir_respuesta_copilot,
    mejorar_redaccion_copilot,
    resumir_conversacion_copilot
)
from app.services.phone_normalizer import normalize_phone_number
from app.whatsapp import (
    iniciar_daemon_whatsapp, 
    whatsapp_manager
)
from app.services.pdf_service import PDF_DIR
from app.services.media_service import media_service, STATIC_MEDIA_DIR
from app.services.media_cleaner import purgar_archivos_antiguos, obtener_estadisticas_storage
from app.services.tools import crear_borrador_presupuesto
from app.services.config_service import load_settings, save_settings
from app.services.agent_orchestrator import orchestrator, AVAILABLE_TOOLS_MAP
from app.services.alcon_catalog_service import alcon_catalog_service
from app.services.geclisa_client import GeclisaClient
from app.services.logger_service import log_event, get_logs, get_logs_stats

geclisa_client = GeclisaClient()

# Configurar logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s"
)
logger = logging.getLogger(__name__)

async def cron_limpieza_diaria_media():
    """
    Tarea en segundo plano que ejecuta la purga de archivos multimedia > 30 días una vez cada 24 horas.
    """
    while True:
        try:
            await asyncio.sleep(60) # Esperar 1 minuto tras arranque del backend
            logger.info("Ejecutando rutina periódica de retención y limpieza de multimedia (> 30 días)...")
            purgar_archivos_antiguos(dias_retencion=30, dry_run=False)
        except Exception as e:
            logger.error(f"Error en rutina periódica de limpieza de multimedia: {e}")
        
        # Esperar 24 horas para la siguiente ejecución
        await asyncio.sleep(24 * 3600)

# Modelos de validación Pydantic
class SimuladorMensaje(BaseModel):
    telefono: str
    mensaje: str

class ToggleBotRequest(BaseModel):
    bot_disabled: bool

class ItemPresupuestoInput(BaseModel):
    codigo_servicio: str
    nombre_prestacion: Optional[str] = None
    cantidad: int = 1
    precio_unitario: Optional[float] = None
    moneda: Optional[str] = "ARS"

class PresupuestoInput(BaseModel):
    paciente_id: str
    items: List[ItemPresupuestoInput]

class SendMessageRequest(BaseModel):
    telefono: Optional[str] = None
    mensaje: Optional[str] = None
    phone: Optional[str] = None
    message: Optional[str] = None
    paciente_id: Optional[str] = None
    conversacion_id: Optional[str] = None
    is_internal_note: Optional[bool] = False
    quoted_message_id: Optional[str] = None
    quoted_message_data: Optional[Dict[str, Any]] = None

class ReactMessageRequest(BaseModel):
    emoji: str

class CopilotSugerirRequest(BaseModel):
    conversacion_id: Optional[str] = None
    paciente_id: Optional[str] = None
    historial: Optional[List[Dict[str, Any]]] = None
    contexto_paciente: Optional[Dict[str, Any]] = None

class CopilotMejorarRequest(BaseModel):
    texto: str

class CopilotResumirRequest(BaseModel):
    conversacion_id: Optional[str] = None
    historial: Optional[List[Dict[str, Any]]] = None

class TestMessageRequest(BaseModel):
    telefono: str
    mensaje: Optional[str] = "¡Hola desde MedCRM! Prueba de vinculación exitosa. 🩺"

# Ciclo de vida de la aplicación FastAPI
@asynccontextmanager
async def lifespan(app: FastAPI):
    logger.info("Iniciando aplicación CRM Médico + WhatsApp Baileys Gateway...")
    iniciar_daemon_whatsapp()
    asyncio.create_task(cron_limpieza_diaria_media())
    yield
    logger.info("Deteniendo aplicación CRM Médico...")

app = FastAPI(
    title="CRM Médico API + Gestor de WhatsApp Baileys",
    description="Backend en FastAPI para gestión clínica médica, sincronización WhatsApp Baileys y agente Gemini.",
    version="3.0.0",
    lifespan=lifespan
)

# Configuración de CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ====================================================================
# SERVIDOR INTELIGENTE DE ESTÁTICOS Y PDFs ON-DEMAND (CON AUTO-REGENERACIÓN)
# ====================================================================

@app.get("/static/{filename}")
def servir_archivo_estatico(filename: str):
    """
    Sirve archivos estáticos y PDFs de presupuestos.
    Si el archivo no existe en el disco local (ej: reinicio o nuevo despliegue de contenedor en Railway),
    lo reconstruye y regenera dinámicamente desde la base de datos de Supabase on-the-fly.
    """
    # 1. Protección contra Path Traversal
    safe_filename = os.path.basename(filename)
    if ".." in filename or "/" in filename or "\\" in filename or safe_filename != filename:
        raise HTTPException(status_code=400, detail="Nombre de archivo no válido.")
        
    file_path = os.path.join(PDF_DIR, safe_filename)
    
    # 2. Si no existe en disco, verificar si es un presupuesto y regenerar
    if not os.path.exists(file_path):
        if safe_filename.startswith("presupuesto_") and safe_filename.endswith(".pdf"):
            presupuesto_id = safe_filename.replace("presupuesto_", "").replace(".pdf", "")
            try:
                if supabase:
                    pres_resp = supabase.table("presupuestos") \
                        .select("*, pacientes(*), items_presupuesto(*, servicios_precios(*)), asesorias_quirurgicas!presupuestos_asesoria_id_fkey(*)") \
                        .eq("id", presupuesto_id) \
                        .execute()
                        
                    if pres_resp.data:
                        p_data = pres_resp.data[0]
                        paciente = p_data.get("pacientes") or {}
                        items_db = p_data.get("items_presupuesto") or []
                        
                        from app.services.pdf_service import generar_pdf_presupuesto
                        generar_pdf_presupuesto(p_data, paciente, items_db)
            except Exception as e:
                logger.error(f"Error regenerando PDF de presupuesto on-demand ({safe_filename}): {e}")
        elif safe_filename.startswith("consentimiento_") and safe_filename.endswith(".pdf"):
            turno_id = safe_filename.replace("consentimiento_", "").replace(".pdf", "")
            try:
                if supabase:
                    t_resp = supabase.table("turnos_quirofano").select("*, pacientes(*), quirofanos(nombre, codigo)").eq("id", turno_id).limit(1).execute()
                    if t_resp.data:
                        t_data = t_resp.data[0]
                        # Resolver texto desde el Nomenclador
                        practica_cod = t_data.get("practica_codigo") or ""
                        practica_id = t_data.get("practica_id") or ""
                        practica_nombre = t_data.get("practica_nombre") or ""
                        
                        cuerpo_template = None
                        resumen_practica = get_practica_resumen_operativo(practica_id or practica_cod or practica_nombre)
                        if resumen_practica and resumen_practica.get("habilitar_consentimiento") and resumen_practica.get("texto_consentimiento"):
                            cuerpo_template = resumen_practica["texto_consentimiento"]
                            
                        if not cuerpo_template:
                            config = get_configuracion_quirofano()
                            plantillas = config.get("plantillas_consentimiento") or []
                            cuerpo_template = plantillas[0]["cuerpo"] if plantillas else "Consentimiento informado para procedimiento quirúrgico."
                            
                        ojo = t_data.get("ojo") or "OD"
                        ojo_desc = "OJO DERECHO (OD)" if ojo == "OD" else "OJO IZQUIERDO (OI)" if ojo == "OI" else "AMBOS OJOS (AO)"
                        
                        cuerpo_final = render_consent_template(
                            cuerpo_template,
                            {
                                "paciente": pac.get("nombre") or "Paciente",
                                "dni": pac.get("dni") or "-",
                                "cirujano": t_data.get("cirujano_nombre") or "Médico Cirujano",
                                "medico": t_data.get("cirujano_nombre") or "Médico Cirujano",
                                "practica": t_data.get("practica_nombre") or "Cirugía Oftalmológica",
                                "cirugia": t_data.get("practica_nombre") or "Cirugía Oftalmológica",
                                "ojo_intervenido": ojo_desc,
                                "ojo": ojo_desc,
                                "quirofano": (t_data.get("quirofanos") or {}).get("nombre") or "Quirófano Central",
                                "fecha": str(t_data.get("fecha_cirugia") or ""),
                                "fecha_cirugia": str(t_data.get("fecha_cirugia") or ""),
                                "hora_cirugia": str(t_data.get("hora_inicio") or "")[:5],
                                "hora_inicio": str(t_data.get("hora_inicio") or "")[:5]
                            }
                        )
                        
                        from app.services.pdf_service import generar_pdf_consentimiento_informado
                        generar_pdf_consentimiento_informado(
                            turno=t_data,
                            paciente=pac,
                            texto_consentimiento=cuerpo_final,
                            firma_img_base64=t_data.get("consentimiento_firma_img"),
                            firma_metadata={
                                "fecha_hora": str(t_data.get("consentimiento_firmado_at") or ""),
                                "ip_origen": str(t_data.get("consentimiento_firma_ip") or "Web-Client")
                            }
                        )
            except Exception as e:
                logger.error(f"Error regenerando PDF de consentimiento on-demand ({safe_filename}): {e}")

    # 3. Si aún no existe, devolver 404
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="El archivo solicitado no fue encontrado en el servidor.")
        
    media_type = "application/pdf" if safe_filename.endswith(".pdf") else "application/octet-stream"
    return FileResponse(
        path=file_path,
        media_type=media_type,
        filename=safe_filename,
        headers={
            "Content-Disposition": f"inline; filename={safe_filename}",
            "X-Content-Type-Options": "nosniff"
        }
    )

@app.get("/static/media/{subfolder}/{filename}")
def servir_archivo_media(subfolder: str, filename: str):
    """
    Sirve archivos multimedia recibidos por WhatsApp (audios, imágenes, stickers, documentos, videos).
    """
    safe_subfolder = os.path.basename(subfolder)
    safe_filename = os.path.basename(filename)
    if ".." in subfolder or ".." in filename or "/" in filename or "\\" in filename:
        raise HTTPException(status_code=400, detail="Ruta de archivo no válida.")
        
    file_path = os.path.join(STATIC_MEDIA_DIR, safe_subfolder, safe_filename)
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="Archivo multimedia no encontrado.")
        
    import mimetypes
    guessed_mime, _ = mimetypes.guess_type(file_path)
    if safe_filename.endswith(".ogg"):
        guessed_mime = "audio/ogg"
    elif safe_filename.endswith(".webp"):
        guessed_mime = "image/webp"
    elif safe_filename.endswith(".mp3"):
        guessed_mime = "audio/mpeg"
        
    return FileResponse(
        path=file_path,
        media_type=guessed_mime or "application/octet-stream",
        filename=safe_filename,
        headers={
            "Content-Disposition": f"inline; filename={safe_filename}",
            "Cache-Control": "public, max-age=86400"
        }
    )

# ====================================================================
# ENDPOINTS GENERALES Y HEALTH
# ====================================================================

@app.get("/")
def read_root():
    st = whatsapp_manager.get_status()
    return {
        "status": "online",
        "servicio": "MedCRM - Gestor de Mensajería & Clínica",
        "whatsapp_engine": "Baileys (Node.js)",
        "whatsapp_status": st.get("status", "UNKNOWN"),
        "pdf_storage_dir": PDF_DIR
    }

@app.get("/api/health")
def health_check():
    supabase_ok = False
    if supabase:
        try:
            supabase.table("servicios_precios").select("id").limit(1).execute()
            supabase_ok = True
        except Exception:
            pass
            
    st = whatsapp_manager.get_status()
    return {
        "api": "ok",
        "supabase": "conectado" if supabase_ok else "desconectado",
        "whatsapp_engine": "Baileys",
        "whatsapp_status": st.get("status", "UNKNOWN")
    }

@app.get("/api/version")
def version_check():
    st = whatsapp_manager.get_status()
    return {
        "version": "3.0.0",
        "engine": "Baileys",
        "status": st.get("status", "UNKNOWN"),
        "is_logged_in": st.get("is_logged_in", False)
    }

# ====================================================================
# ENDPOINTS DE GESTIÓN DE WHATSAPP Y VINCULACIÓN QR
# ====================================================================

@app.get("/api/whatsapp/status")
def get_whatsapp_status():
    """
    Retorna el estado de la conexión, si está autenticado, QR disponible y datos del móvil.
    """
    return whatsapp_manager.get_status()

@app.get("/api/whatsapp/qr")
def get_whatsapp_qr():
    """
    Retorna el código QR activo en formato base64 Data-URI para renderizado directo en UI.
    """
    return whatsapp_manager.get_qr_data()

@app.get("/api/whatsapp/logs")
def get_whatsapp_logs(limit: int = 40):
    """
    Retorna los logs recientes del daemon de WhatsApp para el visor de consola.
    """
    return {"logs": whatsapp_manager.get_logs(limit=limit)}

@app.post("/api/whatsapp/connect")
def connect_whatsapp(force: bool = False):
    """
    Inicia o fuerza el reinicio de la conexión con WhatsApp y generación de QR.
    """
    logger.info(f"Petición de conexión WhatsApp recibida (force={force})")
    iniciar_daemon_whatsapp(force_restart=force)
    return {"success": True, "message": "Proceso de conexión iniciado.", "status": whatsapp_manager.get_status()}

@app.post("/api/whatsapp/logout")
def logout_whatsapp():
    """
    Cierra la sesión de WhatsApp, desvincula el dispositivo y reinicia el estado.
    """
    logger.info("Petición de Logout de WhatsApp recibida")
    ok = whatsapp_manager.desconectar_y_logout()
    if not ok:
        raise HTTPException(status_code=500, detail="Error al cerrar sesión de WhatsApp.")
    return {"success": True, "message": "Sesión cerrada correctamente. Dispositivo desvinculado."}

class IncomingWebhookMessage(BaseModel):
    message_id: Optional[str] = None
    from_me: bool = False
    phone: str
    jid: Optional[str] = None
    remote_jid_alt: Optional[str] = None
    addressing_mode: Optional[str] = None
    name: Optional[str] = "Paciente"
    text: str = ""
    message_type: str = "text"
    media: Optional[Dict[str, Any]] = None
    timestamp: Optional[int] = None
    raw_message: Optional[Dict[str, Any]] = None

def procesar_agente_ia_background(conversacion_id: str, clean_phone: str, texto: str, remote_jid: Optional[str] = None):
    """
    Ejecuta el Agente IA de Gemini en segundo plano con filtro rápido de escalamiento
    y auto-lectura al responder con éxito, enrutando por el canal activo dinámico.
    """
    try:
        texto_lower = texto.lower()
        # Filtro rápido de palabras clave para derivación inmediata a operador humano (0ms latencia)
        patron_humano = r'\b(humano|persona|secretaria|operador|asesor|asesora|hablar con alguien|atencion humana|atención personalizada)\b'
        if re.search(patron_humano, texto_lower):
            logger.info(f"Petición explícita de atención humana detectada en '{texto[:40]}'. Escalando...")
            from app.services.tools import escalar_a_operador_humano
            escalar_a_operador_humano(conversacion_id=conversacion_id, motivo="Solicitud expresa del paciente por mensaje", nivel_urgencia="alta")
            whatsapp_manager.enviar_mensaje(
                clean_phone,
                "Te comunico de inmediato con nuestro equipo de atención humana. Un asesor se contactará a la brevedad.",
                conversacion_id=conversacion_id,
                emisor="bot",
                remote_jid=remote_jid
            )
            return

        respuesta_agente = procesar_mensaje_agente(conversacion_id=conversacion_id, mensaje_texto_o_paciente_id=texto)
        if respuesta_agente:
            whatsapp_manager.enviar_mensaje(clean_phone, respuesta_agente, conversacion_id=conversacion_id, emisor="bot", remote_jid=remote_jid)
            # Auto-lectura de la consulta gestionada 100% por IA (no generar badges falsos a operadores humanos)
            try:
                marcar_mensajes_conversacion_leidos(conversacion_id)
            except Exception as read_err:
                logger.warning(f"Error en auto-lectura por IA para {conversacion_id}: {read_err}")
    except Exception as agent_err:
        logger.error(f"Error procesando respuesta de agente IA en background: {agent_err}")

def unwrap_whatsapp_message(msg_dict: Any) -> Dict[str, Any]:
    """
    Desempaqueta recursivamente mensajes encapsulados de WhatsApp emitidos por dispositivos iOS (iPhone)
    o configuraciones de privacidad (Mensajes Efímeros / Ephemeral, ViewOnce, Documentos con Caption, etc.).
    """
    if not isinstance(msg_dict, dict):
        return {}
    
    current = msg_dict
    for _ in range(10):  # Limitar profundidad para evitar recursión infinita
        if not isinstance(current, dict):
            break
        if "ephemeralMessage" in current and isinstance(current["ephemeralMessage"], dict):
            current = current["ephemeralMessage"].get("message", {})
        elif "viewOnceMessage" in current and isinstance(current["viewOnceMessage"], dict):
            current = current["viewOnceMessage"].get("message", {})
        elif "viewOnceMessageV2" in current and isinstance(current["viewOnceMessageV2"], dict):
            current = current["viewOnceMessageV2"].get("message", {})
        elif "viewOnceMessageV2Extension" in current and isinstance(current["viewOnceMessageV2Extension"], dict):
            current = current["viewOnceMessageV2Extension"].get("message", {})
        elif "documentWithCaptionMessage" in current and isinstance(current["documentWithCaptionMessage"], dict):
            current = current["documentWithCaptionMessage"].get("message", {})
        elif "botInvokeMessage" in current and isinstance(current["botInvokeMessage"], dict):
            current = current["botInvokeMessage"].get("message", {})
        elif "templateMessage" in current and isinstance(current["templateMessage"], dict):
            hydrated = current["templateMessage"].get("hydratedTemplate", {})
            current = hydrated or current["templateMessage"]
        elif "interactiveMessage" in current and isinstance(current["interactiveMessage"], dict):
            body = current["interactiveMessage"].get("body", {})
            if isinstance(body, dict) and "text" in body:
                return {"conversation": body["text"]}
            break
        else:
            break

    return current if isinstance(current, dict) else {}

@app.post("/api/whatsapp/webhook/incoming")
async def receive_incoming_whatsapp_message(request: Request, background_tasks: BackgroundTasks):
    """
    Webhook unificado que recibe eventos y mensajes entrantes de WhatsApp desde Evolution API v2,
    persiste el mensaje en Supabase y despacha el Agente IA en segundo plano.
    """
    try:
        body = await request.json()
    except Exception:
        return {"status": "error", "detail": "Invalid JSON"}

    try:
        # Caso 1: Eventos de Evolution API v2
        if isinstance(body, dict) and "event" in body:
            event_name = str(body.get("event", "")).upper()
            data = body.get("data", {})

            # 1.1 Actualización de estados de entrega (SERVER_ACK, DELIVERY_ACK, READ)
            if event_name in ["MESSAGES_UPDATE", "MESSAGES.UPDATE"]:
                updates = data if isinstance(data, list) else [data]
                for upd in updates:
                    if not isinstance(upd, dict):
                        continue
                    key = upd.get("key", {}) if isinstance(upd.get("key"), dict) else {}
                    msg_id = key.get("id") or upd.get("id")
                    status_raw = upd.get("update", {}).get("status") if isinstance(upd.get("update"), dict) else upd.get("status")
                    if msg_id and status_raw is not None and supabase:
                        status_label = "enviado"
                        s_str = str(status_raw).upper()
                        if s_str in ["3", "DELIVERY_ACK", "DELIVERED"]:
                            status_label = "entregado"
                        elif s_str in ["4", "5", "READ", "PLAYED"]:
                            status_label = "leido"
                        try:
                            m_res = supabase.table("mensajes").select("id, metadata_json").filter("metadata_json->>whatsapp_message_id", "eq", msg_id).execute()
                            if m_res.data:
                                row = m_res.data[0]
                                meta = row.get("metadata_json") or {}
                                if isinstance(meta, dict):
                                    meta["delivery_status"] = status_label
                                    supabase.table("mensajes").update({"metadata_json": meta}).eq("id", row["id"]).execute()
                        except Exception as upd_err:
                            logger.warning(f"Error actualizando delivery_status para {msg_id}: {upd_err}")
                return {"status": "processed", "event": event_name}

            # 1.2 Mensaje entrante de WhatsApp (MESSAGES_UPSERT)
            if event_name in ["MESSAGES_UPSERT", "MESSAGES.UPSERT"]:
                key = data.get("key", {}) if isinstance(data.get("key"), dict) else {}
                from_me = key.get("fromMe", False)
                if from_me:
                    return {"status": "ignored", "reason": "outgoing_message"}

                remote_jid = key.get("remoteJid", "")
                remote_jid_alt = (
                    key.get("remoteJidAlt") 
                    or data.get("remoteJidAlt") 
                    or key.get("participant") 
                    or data.get("participant") 
                    or ""
                )
                addressing_mode = key.get("addressingMode") or data.get("addressingMode") or ""
                message_id = key.get("id", "")
                push_name = data.get("pushName", "Paciente")
                
                # Desempaquetar capas envolventes de iOS / Ephemeral / ViewOnce
                raw_msg = data.get("message", {}) or {}
                msg_content = unwrap_whatsapp_message(raw_msg)
                
                texto = ""
                message_type = "texto"
                media_info = None
                base64_data = data.get("base64")

                # REACCION A UN MENSAJE
                if isinstance(msg_content, dict) and "reactionMessage" in msg_content:
                    react = msg_content.get("reactionMessage", {})
                    target_id = react.get("key", {}).get("id")
                    emoji = react.get("text", "")
                    if target_id and emoji and supabase:
                        try:
                            t_res = supabase.table("mensajes").select("id, metadata_json").filter("metadata_json->>whatsapp_message_id", "eq", target_id).execute()
                            if t_res.data:
                                t_row = t_res.data[0]
                                t_meta = t_row.get("metadata_json") or {}
                                if isinstance(t_meta, dict):
                                    reactions_list = t_meta.get("reactions") or []
                                    reactions_list.append({
                                        "emisor": "paciente",
                                        "emoji": emoji,
                                        "timestamp": datetime.datetime.now(datetime.timezone.utc).isoformat()
                                    })
                                    t_meta["reactions"] = reactions_list
                                    supabase.table("mensajes").update({"metadata_json": t_meta}).eq("id", t_row["id"]).execute()
                        except Exception as re_err:
                            logger.warning(f"Error procesando reacción en webhook: {re_err}")
                    return {"status": "processed", "event": "reaction"}

                if isinstance(msg_content, dict):
                    # TEXTO SIMPLE
                    if "conversation" in msg_content:
                        texto = msg_content.get("conversation", "")
                        message_type = "texto"
                    elif "extendedTextMessage" in msg_content:
                        texto = msg_content.get("extendedTextMessage", {}).get("text", "")
                        message_type = "texto"
                    elif "buttonsResponseMessage" in msg_content:
                        texto = msg_content.get("buttonsResponseMessage", {}).get("selectedDisplayText", "") or msg_content.get("buttonsResponseMessage", {}).get("selectedButtonId", "")
                        message_type = "texto"
                    elif "templateButtonReplyMessage" in msg_content:
                        texto = msg_content.get("templateButtonReplyMessage", {}).get("selectedDisplayText", "") or msg_content.get("templateButtonReplyMessage", {}).get("selectedId", "")
                        message_type = "texto"
                    elif "listResponseMessage" in msg_content:
                        texto = msg_content.get("listResponseMessage", {}).get("title", "") or msg_content.get("listResponseMessage", {}).get("description", "")
                        message_type = "texto"
                    elif "interactiveResponseMessage" in msg_content:
                        body_txt = msg_content.get("interactiveResponseMessage", {}).get("body", {}).get("text", "")
                        native_p = msg_content.get("interactiveResponseMessage", {}).get("nativeFlowResponseMessage", {}).get("paramsJson", "")
                        texto = body_txt or native_p
                        message_type = "texto"

                    # IMAGEN
                    elif "imageMessage" in msg_content:
                        message_type = "imagen"
                        img_data = msg_content.get("imageMessage", {})
                        caption = img_data.get("caption", "")
                        texto = caption
                        mime = img_data.get("mimetype", "image/jpeg")
                        b64 = base64_data or img_data.get("base64")
                        if not b64 and message_id:
                            m_resp = whatsapp_manager.get_media_base64(message_id)
                            if m_resp:
                                b64 = m_resp.get("base64")
                                mime = m_resp.get("mimetype", mime)
                        media_info = {"tipo": "imagen", "mime_type": mime, "caption": caption}
                        if b64:
                            try:
                                saved = media_service.save_base64_media(b64, "images", mime, "imagen.jpg")
                                media_info.update(saved)
                            except Exception as e:
                                logger.warning(f"Error guardando imagen base64: {e}")

                    # AUDIO / NOTA DE VOZ
                    elif "audioMessage" in msg_content:
                        message_type = "audio"
                        aud_data = msg_content.get("audioMessage", {})
                        mime = aud_data.get("mimetype", "audio/ogg; codecs=opus")
                        is_ptt = aud_data.get("ptt", True)
                        b64 = base64_data or aud_data.get("base64")
                        if not b64 and message_id:
                            m_resp = whatsapp_manager.get_media_base64(message_id)
                            if m_resp:
                                b64 = m_resp.get("base64")
                                mime = m_resp.get("mimetype", mime)
                        
                        media_info = {"tipo": "audio", "mime_type": mime, "is_voice_note": bool(is_ptt)}
                        if b64:
                            try:
                                saved = media_service.save_base64_media(b64, "audio", mime, "audio.ogg")
                                media_info.update(saved)
                                try:
                                    import base64 as b64_mod
                                    raw_audio = b64_mod.b64decode(b64.split("base64,")[-1])
                                    transcript = transcribir_audio_con_gemini(audio_bytes=raw_audio, mime_type=mime)
                                    if transcript:
                                        media_info["transcripcion"] = transcript
                                        texto = transcript
                                        logger.info(f"✔ Audio de WhatsApp transcripto automáticamente: '{transcript[:60]}'")
                                except Exception as t_err:
                                    logger.warning(f"Error en auto-transcripción de audio: {t_err}")
                            except Exception as e:
                                logger.warning(f"Error guardando audio base64: {e}")

                    # STICKER
                    elif "stickerMessage" in msg_content:
                        message_type = "sticker"
                        stk_data = msg_content.get("stickerMessage", {})
                        mime = stk_data.get("mimetype", "image/webp")
                        b64 = base64_data or stk_data.get("base64")
                        if not b64 and message_id:
                            m_resp = whatsapp_manager.get_media_base64(message_id)
                            if m_resp:
                                b64 = m_resp.get("base64")
                                mime = m_resp.get("mimetype", mime)
                        media_info = {"tipo": "sticker", "mime_type": mime}
                        if b64:
                            try:
                                saved = media_service.save_base64_media(b64, "stickers", mime, "sticker.webp")
                                media_info.update(saved)
                            except Exception as e:
                                logger.warning(f"Error guardando sticker base64: {e}")

                    # DOCUMENTO
                    elif "documentMessage" in msg_content:
                        message_type = "documento"
                        doc_data = msg_content.get("documentMessage", {})
                        filename = doc_data.get("fileName", "Documento.pdf")
                        caption = doc_data.get("caption", "")
                        mime = doc_data.get("mimetype", "application/pdf")
                        b64 = base64_data or doc_data.get("base64")
                        if not b64 and message_id:
                            m_resp = whatsapp_manager.get_media_base64(message_id)
                            if m_resp:
                                b64 = m_resp.get("base64")
                                mime = m_resp.get("mimetype", mime)
                        media_info = {"tipo": "documento", "mime_type": mime, "file_name": filename, "caption": caption}
                        if b64:
                            try:
                                saved = media_service.save_base64_media(b64, "documents", mime, filename)
                                media_info.update(saved)
                            except Exception as e:
                                logger.warning(f"Error guardando documento base64: {e}")

                    # VIDEO
                    elif "videoMessage" in msg_content:
                        message_type = "video"
                        vid_data = msg_content.get("videoMessage", {})
                        caption = vid_data.get("caption", "")
                        mime = vid_data.get("mimetype", "video/mp4")
                        b64 = base64_data or vid_data.get("base64")
                        if not b64 and message_id:
                            m_resp = whatsapp_manager.get_media_base64(message_id)
                            if m_resp:
                                b64 = m_resp.get("base64")
                                mime = m_resp.get("mimetype", mime)
                        media_info = {"tipo": "video", "mime_type": mime, "caption": caption}
                        if b64:
                            try:
                                saved = media_service.save_base64_media(b64, "videos", mime, "video.mp4")
                                media_info.update(saved)
                            except Exception as e:
                                logger.warning(f"Error guardando video base64: {e}")

                    # UBICACION
                    elif "locationMessage" in msg_content:
                        message_type = "ubicacion"
                        loc_data = msg_content.get("locationMessage", {})
                        lat = loc_data.get("degreesLatitude")
                        lng = loc_data.get("degreesLongitude")
                        name = loc_data.get("name", "Ubicación")
                        media_info = {"tipo": "ubicacion", "latitud": lat, "longitud": lng, "nombre": name, "maps_url": f"https://www.google.com/maps?q={lat},{lng}"}
                        texto = f"📍 {name}"

                    # CONTACTO
                    elif "contactMessage" in msg_content or "contactsArrayMessage" in msg_content:
                        message_type = "contacto"
                        c_data = msg_content.get("contactMessage", {})
                        c_name = c_data.get("displayName", "Contacto")
                        c_vcard = c_data.get("vcard", "")
                        media_info = {"tipo": "contacto", "nombre": c_name, "vcard": c_vcard}
                        texto = f"👤 {c_name}"

                # Resolución inteligente del teléfono real (Soporta LID y remoteJidAlt)
                real_phone_str = ""
                if remote_jid_alt and not str(remote_jid_alt).endswith("@lid"):
                    real_phone_str = str(remote_jid_alt).split("@")[0].split(":")[0]
                elif remote_jid and not str(remote_jid).endswith("@lid"):
                    real_phone_str = str(remote_jid).split("@")[0].split(":")[0]
                else:
                    pac_by_lid = get_paciente_by_lid(remote_jid)
                    if pac_by_lid and pac_by_lid.get("telefono"):
                        real_phone_str = pac_by_lid.get("telefono")
                    elif push_name and push_name != "Paciente":
                        pac_by_name = get_paciente_by_nombre_aproximado(push_name)
                        if pac_by_name and pac_by_name.get("telefono"):
                            real_phone_str = pac_by_name.get("telefono")
                        else:
                            lid_clean = "".join(filter(str.isdigit, str(remote_jid)))
                            real_phone_str = f"lid_{lid_clean}"
                    else:
                        lid_clean = "".join(filter(str.isdigit, str(remote_jid)))
                        real_phone_str = f"lid_{lid_clean}"

                clean_phone = normalize_phone_number(real_phone_str) if not str(real_phone_str).startswith("lid_") else str(real_phone_str)
                timestamp = data.get("messageTimestamp", int(time.time()))
                
                payload = IncomingWebhookMessage(
                    message_id=message_id,
                    from_me=from_me,
                    phone=clean_phone,
                    jid=remote_jid,
                    remote_jid_alt=remote_jid_alt,
                    addressing_mode=addressing_mode,
                    name=push_name,
                    text=texto or "",
                    message_type=message_type,
                    media=media_info,
                    timestamp=timestamp
                )
            else:
                return {"status": "ignored", "reason": f"unhandled_event_{event_name}"}
        else:
            payload = IncomingWebhookMessage(**body)

        if payload.from_me:
            return {"status": "ignored", "reason": "outgoing_message"}

        clean_phone = normalize_phone_number(payload.phone) if payload.phone and not payload.phone.startswith("lid_") else (payload.phone or "")
        texto = payload.text.strip() if payload.text else ""
        incoming_jid = payload.jid or payload.phone

        logger.info(f"Mensaje entrante WhatsApp desde {clean_phone} (JID: {incoming_jid}, Alt: {payload.remote_jid_alt}) [{payload.message_type}]: {texto[:50]}")

        if not clean_phone:
            return {"status": "ignored", "reason": "empty_phone"}

        # 1. Obtener o crear paciente (con resolución multicriterio por teléfono, LID o nombre)
        paciente = None
        if clean_phone and not clean_phone.startswith("lid_"):
            paciente = get_paciente_by_telefono(clean_phone)
        if not paciente and incoming_jid:
            paciente = get_paciente_by_lid(incoming_jid)
        if not paciente and payload.name and payload.name != "Paciente":
            paciente = get_paciente_by_nombre_aproximado(payload.name)

        if not paciente:
            nombre = payload.name if payload.name and payload.name != "Paciente" else f"Paciente {clean_phone[-4:] if len(clean_phone) >= 4 else clean_phone}"
            paciente = crear_paciente(telefono=clean_phone, nombre=nombre)
        
        if not paciente:
            logger.error(f"No se pudo crear ni obtener paciente para {clean_phone}")
            return {"status": "error", "detail": "No se pudo obtener paciente"}

        paciente_id = paciente["id"] if isinstance(paciente, dict) else paciente.get("id")

        # 1.1 Interceptor Inteligente de DNI para Pacientes
        if paciente and not paciente.get("dni") and texto:
            dni_match = re.search(r'\b(\d{7,8})\b', texto)
            if dni_match:
                potential_dni = dni_match.group(1)
                try:
                    res_geclisa = geclisa_client.buscar_paciente_por_dni(potential_dni)
                    if res_geclisa and res_geclisa.get("encontrado"):
                        paciente_vinculado = vincular_o_fusionar_paciente_con_geclisa(
                            paciente_temporal_id=paciente_id,
                            datos_geclisa=res_geclisa,
                            telefono_whatsapp=clean_phone
                        )
                        if paciente_vinculado:
                            paciente = paciente_vinculado
                            paciente_id = paciente.get("id")
                    else:
                        registrar_dni_paciente_nuevo_crm(paciente_id, potential_dni)
                        paciente["dni"] = potential_dni
                except Exception as g_err:
                    logger.warning(f"Error en auto-vinculación de Geclisa desde webhook: {g_err}")

        # 2. Conversación
        conversacion = get_or_create_conversacion(paciente_id)
        if not conversacion:
            return {"status": "error", "detail": "No se pudo obtener conversación"}

        conversacion_id = conversacion["id"] if isinstance(conversacion, dict) else conversacion.get("id")

        # 3. Guardar mensaje entrante del paciente (con deduplicación)
        if payload.message_id and supabase:
            try:
                existente = supabase.table("mensajes").select("id").filter("metadata_json->>whatsapp_message_id", "eq", payload.message_id).execute()
                if existente.data and len(existente.data) > 0:
                    logger.info(f"Mensaje {payload.message_id} ya registrado previamente. Omitiendo duplicado.")
                    return {"status": "ignored", "reason": "duplicate_message_id"}
            except Exception as dedup_err:
                logger.warning(f"Error verificando duplicado: {dedup_err}")

        meta = {
            "whatsapp_message_id": payload.message_id,
            "remote_jid": incoming_jid,
            "remote_jid_alt": payload.remote_jid_alt or None,
            "addressing_mode": payload.addressing_mode or None,
            "push_name": payload.name,
            "gateway": "evolution_api_v2",
            "tipo": payload.message_type
        }
        if payload.media and isinstance(payload.media, dict):
            meta.update(payload.media)

        caption_texto = payload.media.get("caption") if (payload.media and isinstance(payload.media, dict)) else None
        
        # Etiqueta amigable de contenido para la vista previa en la bandeja de chats
        if payload.message_type == "imagen":
            contenido_final = caption_texto or "📷 [Imagen]"
        elif payload.message_type == "audio":
            contenido_final = "🎵 [Nota de Voz]"
        elif payload.message_type == "sticker":
            contenido_final = "💟 [Sticker]"
        elif payload.message_type == "documento":
            doc_name = payload.media.get("file_name", "Documento") if payload.media else "Documento"
            contenido_final = caption_texto or f"📄 [{doc_name}]"
        elif payload.message_type == "video":
            contenido_final = caption_texto or "🎥 [Video]"
        elif payload.message_type == "ubicacion":
            contenido_final = texto or "📍 [Ubicación]"
        elif payload.message_type == "contacto":
            contenido_final = texto or "👤 [Contacto]"
        else:
            contenido_final = texto or "[MENSAJE]"

        created_at_iso = None
        if payload.timestamp and payload.timestamp > 0:
            try:
                import datetime
                created_at_iso = datetime.datetime.fromtimestamp(payload.timestamp, tz=datetime.timezone.utc).isoformat()
            except Exception:
                pass

        guardar_mensaje(
            conversacion_id=conversacion_id,
            emisor="paciente",
            contenido=contenido_final,
            metadata_json=meta,
            created_at=created_at_iso
        )

        # 4. Despachar agente IA en background
        bot_disabled = conversacion.get("bot_disabled", False) if isinstance(conversacion, dict) else False
        if not bot_disabled and texto:
            background_tasks.add_task(procesar_agente_ia_background, conversacion_id, clean_phone, texto, incoming_jid)

        return {"status": "processed", "conversacion_id": conversacion_id, "telefono": clean_phone, "remote_jid": incoming_jid}
    except Exception as e:
        logger.error(f"Error procesando mensaje entrante WhatsApp: {e}", exc_info=True)
        log_event(nivel="ERROR", modulo="WHATSAPP", accion="ERROR_WEBHOOK_FASTAPI", mensaje=f"Error procesando mensaje: {e}", detalles={"error": str(e)})
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/conversaciones")
def get_conversaciones_api(incluir_archivadas: bool = True):
    """
    Retorna la lista de todas las conversaciones activas con sus pacientes asociados.
    """
    return obtener_conversaciones(incluir_archivadas=incluir_archivadas)

@app.get("/api/conversaciones/metricas")
def get_conversaciones_metricas_api():
    """
    Retorna los contadores en tiempo real para las pestañas de la bandeja de entrada:
    - Derivados a Humano
    - En Gestión por IA (Bot Activo)
    - Total Activos
    - Resueltos / Archivados
    """
    return obtener_metricas_conversaciones()

@app.post("/api/conversaciones/{conversacion_id}/archivar")
def archivar_conversacion_api(conversacion_id: str, payload: Dict[str, Any] = Body(...)):
    """
    Marca una conversación como archivada / resuelta o la restaura a activa.
    """
    archivada = payload.get("archivada", True)
    logger.info(f"Cambiando estado archivada en conversación {conversacion_id} a {archivada}")
    res = archivar_conversacion(conversacion_id, archivada)
    if not res:
        raise HTTPException(status_code=404, detail="No se pudo actualizar el estado de la conversación.")
    return {"success": True, "conversacion": res}

@app.get("/api/conversaciones/{conversacion_id}/mensajes")
def get_mensajes_conversacion_api(conversacion_id: str):
    """
    Retorna todo el historial de mensajes de una conversación específica.
    """
    return obtener_mensajes_conversacion(conversacion_id)

@app.post("/api/conversaciones/{conversacion_id}/leer")
def marcar_conversacion_leida_api(conversacion_id: str):
    """
    Marca los mensajes entrantes del paciente como leídos por el operador
    y emite el recibo de lectura (doble tilde azul) a WhatsApp.
    """
    try:
        res = marcar_mensajes_conversacion_leidos(conversacion_id)
        w_ids = res.get("whatsapp_message_ids", [])
        telefono = res.get("telefono")

        if telefono and w_ids:
            whatsapp_manager.marcar_como_leido(telefono_o_jid=telefono, message_ids=w_ids)

        return {"success": True, "mensajes_marcados": len(w_ids)}
    except Exception as e:
        logger.error(f"Error marcando conversación {conversacion_id} como leída: {e}")
        return {"success": False, "error": str(e)}

@app.post("/api/conversaciones/{conversacion_id}/marcar-no-leido")
def marcar_conversacion_no_leida_api(conversacion_id: str):
    """
    Marca una conversación como no leída por el operador para dejarla pendiente de atención.
    """
    if not supabase:
        raise HTTPException(status_code=500, detail="Base de datos no disponible")
    try:
        res = supabase.table("conversaciones").select("id, metadata_json").eq("id", conversacion_id).execute()
        if not res.data:
            raise HTTPException(status_code=404, detail="Conversación no encontrada")
        
        meta = res.data[0].get("metadata_json") or {}
        meta["manual_unread"] = True
        
        supabase.table("conversaciones").update({
            "unread_count": 1,
            "metadata_json": meta
        }).eq("id", conversacion_id).execute()

        return {"success": True, "unread_count": 1}
    except Exception as e:
        logger.error(f"Error marcando conversación {conversacion_id} como no leída: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/conversaciones/{conversacion_id}/fijar")
def fijar_conversacion_api(conversacion_id: str, payload: Dict[str, Any] = Body(...)):
    """
    Fija o desfija una conversación para que aparezca prioritariamente al tope de la lista.
    """
    if not supabase:
        raise HTTPException(status_code=500, detail="Base de datos no disponible")
    try:
        fijada = payload.get("fijada", True)
        res = supabase.table("conversaciones").select("id, metadata_json").eq("id", conversacion_id).execute()
        if not res.data:
            raise HTTPException(status_code=404, detail="Conversación no encontrada")
        
        meta = res.data[0].get("metadata_json") or {}
        meta["is_pinned"] = fijada

        supabase.table("conversaciones").update({
            "metadata_json": meta
        }).eq("id", conversacion_id).execute()

        return {"success": True, "is_pinned": fijada}
    except Exception as e:
        logger.error(f"Error fijando conversación {conversacion_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/api/conversaciones/{conversacion_id}")
def eliminar_conversacion_api(conversacion_id: str):
    """
    Elimina una conversación y sus mensajes asociados del CRM.
    """
    if not supabase:
        raise HTTPException(status_code=500, detail="Base de datos no disponible")
    try:
        supabase.table("mensajes").delete().eq("conversacion_id", conversacion_id).execute()
        supabase.table("conversaciones").delete().eq("id", conversacion_id).execute()
        return {"success": True, "conversacion_id": conversacion_id}
    except Exception as e:
        logger.error(f"Error eliminando conversación {conversacion_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/mensajes/{mensaje_id}/transcribir")
def transcribir_mensaje_api(mensaje_id: str):
    """
    Transcribe un mensaje de audio bajo demanda utilizando Google Gemini Multimodal.
    Retorna el texto transcripto y lo almacena en metadata_json.transcripcion del mensaje.
    """
    if not supabase:
        raise HTTPException(status_code=500, detail="Base de datos Supabase no configurada.")
    
    try:
        res = supabase.table("mensajes").select("*").eq("id", mensaje_id).execute()
        if not res.data:
            raise HTTPException(status_code=404, detail="Mensaje no encontrado.")
        
        msg = res.data[0]
        meta = msg.get("metadata_json") or {}
        if isinstance(meta, str):
            import json
            meta = json.loads(meta)

        # Si ya está transcripto, retornar directamente
        if meta.get("transcripcion") and str(meta.get("transcripcion")).strip():
            return {
                "success": True,
                "transcripcion": meta.get("transcripcion"),
                "mensaje_id": mensaje_id,
                "cached": True
            }

        audio_url = meta.get("media_url") or meta.get("relative_url")
        if not audio_url:
            raise HTTPException(status_code=400, detail="El mensaje no contiene una URL de audio válida.")

        # Si la URL es relativa y apunta a static
        if not audio_url.startswith("http"):
            audio_url = f"http://localhost:8000/{audio_url.lstrip('/')}"

        logger.info(f"Iniciando transcripción de audio para mensaje {mensaje_id}...")
        texto_transcrito = transcribir_audio_con_gemini(audio_url)
        
        # Persistir en la base de datos
        guardar_transcripcion_mensaje(mensaje_id, texto_transcrito)

        return {
            "success": True,
            "transcripcion": texto_transcrito,
            "mensaje_id": mensaje_id,
            "cached": False
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error procesando transcripción de audio: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Error al transcribir el audio con Gemini: {str(e)}")

@app.post("/api/mantenimiento/purgar-media")
def purgar_media_api(dias: int = 30, dry_run: bool = False):
    """
    Ejecuta la purga de archivos multimedia antiguos (> 30 días) en Supabase Storage
    auto-transcribiendo los audios previos con Gemini para preservar su contenido textual.
    """
    logger.info(f"Endpoint de mantenimiento invocado: purga con retención de {dias} días (dry_run={dry_run})")
    return purgar_archivos_antiguos(dias_retencion=dias, dry_run=dry_run)

@app.get("/api/mantenimiento/estadisticas-storage")
def estadisticas_storage_api():
    """
    Retorna métricas de archivos almacenados, espacio utilizado y conteo de audios transcriptos.
    """
    return obtener_estadisticas_storage()

@app.post("/api/whatsapp/send-message")
def send_message_api(payload: SendMessageRequest):
    """
    Envía un mensaje de texto directamente al WhatsApp real del paciente o lo guarda como NOTA INTERNA privada del equipo.
    """
    texto_final = (payload.mensaje or payload.message or "").strip()
    if not texto_final:
        raise HTTPException(status_code=400, detail="El contenido del mensaje no puede estar vacío.")

    # 1. NOTA INTERNA (Solo visible en el CRM, NUNCA se envía al paciente por WhatsApp)
    if payload.is_internal_note:
        logger.info(f"Registrando nota interna en conversación {payload.conversacion_id}")
        meta = {
            "is_internal_note": True,
            "tipo": "nota_interna",
            "autor": "Operador Humano"
        }
        msg = guardar_mensaje(
            conversacion_id=payload.conversacion_id,
            emisor="operador",
            contenido=texto_final,
            metadata_json=meta
        )
        return {
            "success": True,
            "is_internal_note": True,
            "guardado_db": True,
            "mensaje": msg
        }

    # 2. MENSAJE SALIENTE A WHATSAPP
    telefono_raw = (payload.telefono or payload.phone or "").strip()
    telefono_final = normalize_phone_number(telefono_raw) if telefono_raw else ""
    conversacion_id = payload.conversacion_id
    paciente_id = payload.paciente_id

    # A. Resolver conversacion_id si se pasó paciente_id
    if not conversacion_id and paciente_id:
        try:
            conv = get_or_create_conversacion(paciente_id)
            if conv:
                conversacion_id = conv.get("id")
        except Exception as e:
            logger.warning(f"No se pudo obtener/crear conversación para paciente {paciente_id}: {e}")

    # B. Resolver conversacion_id si se pasó teléfono
    if not conversacion_id and telefono_final:
        try:
            pac = get_paciente_by_telefono(telefono_final)
            if not pac and paciente_id and supabase:
                p_resp = supabase.table("pacientes").select("*").eq("id", paciente_id).execute()
                if p_resp.data:
                    pac = p_resp.data[0]
            if pac:
                conv = get_or_create_conversacion(pac["id"])
                if conv:
                    conversacion_id = conv.get("id")
                    if not paciente_id:
                        paciente_id = pac["id"]
        except Exception as e:
            logger.warning(f"No se pudo resolver conversación por teléfono {telefono_final}: {e}")

    # C. Recuperar y normalizar teléfono si faltaba
    if not telefono_final and conversacion_id:
        try:
            if supabase:
                conv_data = supabase.table("conversaciones").select("paciente_id, pacientes(telefono)").eq("id", conversacion_id).execute()
                if conv_data.data and len(conv_data.data) > 0:
                    p_data = conv_data.data[0].get("pacientes")
                    raw_tel = None
                    if isinstance(p_data, list) and len(p_data) > 0:
                        raw_tel = p_data[0].get("telefono")
                    elif isinstance(p_data, dict):
                        raw_tel = p_data.get("telefono")
                    if raw_tel:
                        telefono_final = normalize_phone_number(raw_tel)
        except Exception as e:
            logger.warning(f"No se pudo recuperar teléfono por conversación {conversacion_id}: {e}")

    if not telefono_final and paciente_id:
        try:
            if supabase:
                pac_res = supabase.table("pacientes").select("telefono").eq("id", paciente_id).execute()
                if pac_res.data and len(pac_res.data) > 0:
                    raw_tel = pac_res.data[0].get("telefono")
                    if raw_tel:
                        telefono_final = normalize_phone_number(raw_tel)
        except Exception as e:
            logger.warning(f"No se pudo recuperar teléfono por paciente_id {payload.paciente_id}: {e}")

    if not telefono_final:
        raise HTTPException(status_code=400, detail="Debe especificar un número de teléfono de destino válido.")

    logger.info(f"Enviando mensaje saliente a {telefono_final} [operador] (conversacion_id: {conversacion_id})")
    result = whatsapp_manager.enviar_mensaje(
        telefono_o_jid=telefono_final,
        texto=texto_final,
        conversacion_id=conversacion_id,
        emisor="operador",
        quoted_message_id=payload.quoted_message_id,
        quoted_message_data=payload.quoted_message_data
    )
    if "error" in result and not result.get("guardado_db"):
        raise HTTPException(status_code=400, detail=result["error"])
    return result

@app.post("/api/mensajes/{mensaje_id}/reaccionar")
def reaccionar_mensaje_api(mensaje_id: str, payload: ReactMessageRequest):
    """
    Envía una reacción de emoji a un mensaje por WhatsApp y actualiza Supabase.
    """
    if not supabase:
        raise HTTPException(status_code=500, detail="Base de datos no disponible")

    res = supabase.table("mensajes").select("id, emisor, whatsapp_message_id, conversacion_id, metadata_json").eq("id", mensaje_id).execute()
    if not res.data or len(res.data) == 0:
        raise HTTPException(status_code=404, detail="Mensaje no encontrado")

    msg = res.data[0]
    wa_msg_id = msg.get("whatsapp_message_id")
    conv_id = msg.get("conversacion_id")
    meta = msg.get("metadata_json") or {}
    
    if wa_msg_id and not meta.get("is_internal_note"):
        active_jid = meta.get("dispatched_jid")
        if not active_jid:
            active_jid = get_active_jid_for_paciente_o_conversacion(conversacion_id=conv_id)
        
        if active_jid:
            whatsapp_manager.enviar_reaccion(
                message_id=wa_msg_id,
                remote_jid=str(active_jid),
                emoji=payload.emoji,
                from_me=(msg.get("emisor") == "operador")
            )

    reactions = meta.get("reactions") or []
    reactions = [r for r in reactions if r.get("emisor") != "operador"]
    if payload.emoji:
        reactions.append({
            "emisor": "operador",
            "emoji": payload.emoji,
            "created_at": datetime.now().isoformat()
        })
    meta["reactions"] = reactions

    supabase.table("mensajes").update({"metadata_json": meta}).eq("id", mensaje_id).execute()
    return {"success": True, "reactions": reactions, "mensaje_id": mensaje_id}

@app.delete("/api/mensajes/{mensaje_id}")
def eliminar_mensaje_api(mensaje_id: str):
    """
    Elimina un mensaje del chat en el CRM.
    """
    if not supabase:
        raise HTTPException(status_code=500, detail="Base de datos no disponible")

    supabase.table("mensajes").delete().eq("id", mensaje_id).execute()
    return {"success": True, "mensaje_id": mensaje_id}

@app.get("/api/whatsapp/chequeo-canal")
def chequeo_canal_whatsapp_api(telefono: str):
    """
    Inspecciona en tiempo real el canal activo, el paciente, la conversación y el estado onWhatsApp para un número telefónico.
    """
    clean_tel = normalize_phone_number(telefono) if telefono else ""
    if not clean_tel:
        raise HTTPException(status_code=400, detail="Debe especificar un número de teléfono.")

    paciente = get_paciente_by_telefono(clean_tel)
    conversacion = None
    active_remote_jid = None
    if paciente and supabase:
        try:
            c_resp = supabase.table("conversaciones").select("*").eq("paciente_id", paciente["id"]).execute()
            if c_resp.data and len(c_resp.data) > 0:
                conversacion = c_resp.data[0]
                active_remote_jid = (conversacion.get("metadata_json") or {}).get("active_remote_jid")
        except Exception:
            pass

    # Consultar estado en pasarela Node.js
    node_status = {}
    try:
        r = httpx.get(f"{whatsapp_manager.service_url}/channel-check?phone={clean_tel}", timeout=4.0)
        if r.status_code == 200:
            node_status = r.json()
    except Exception as err:
        node_status = {"error": str(err)}

    return {
        "telefono_solicitado": telefono,
        "telefono_normalizado": clean_tel,
        "paciente": paciente,
        "conversacion_id": conversacion.get("id") if conversacion else None,
        "active_remote_jid_en_db": active_remote_jid,
        "gateway_node": node_status
    }

# ====================================================================
# ENDPOINTS COPILOTO DE IA (GEMINI)
# ====================================================================

@app.post("/api/chat/copilot/sugerir")
def copilot_sugerir_api(payload: CopilotSugerirRequest):
    """
    Genera una sugerencia de respuesta redactada con Gemini basada en el historial del chat y la ficha clínica.
    """
    msgs = payload.historial or []
    contexto = payload.contexto_paciente or {}
    
    if payload.conversacion_id and (not msgs or not contexto):
        try:
            if not msgs:
                msgs = obtener_mensajes_conversacion(payload.conversacion_id)
            if payload.paciente_id and not contexto:
                contexto = get_paciente_contexto_360(payload.paciente_id)
        except Exception as err:
            logger.warning(f"Error cargando contexto para copilot sugerir: {err}")

    sugerencia = sugerir_respuesta_copilot(msgs, contexto)
    return {"success": True, "sugerencia": sugerencia}

@app.post("/api/chat/copilot/mejorar")
def copilot_mejorar_api(payload: CopilotMejorarRequest):
    """
    Mejora la redacción, tono, calidez y ortografía de un borrador de mensaje.
    """
    mejorado = mejorar_redaccion_copilot(payload.texto)
    return {"success": True, "texto_mejorado": mejorado}

@app.post("/api/chat/copilot/resumir")
def copilot_resumir_api(payload: CopilotResumirRequest):
    """
    Genera un resumen ejecutivo en 3 puntos clave de la conversación médica/administrativa.
    """
    msgs = payload.historial or []
    if payload.conversacion_id and not msgs:
        try:
            msgs = obtener_mensajes_conversacion(payload.conversacion_id)
        except Exception as err:
            logger.warning(f"Error cargando mensajes para copilot resumir: {err}")

    resumen = resumir_conversacion_copilot(msgs)
    return {"success": True, "resumen": resumen}

@app.post("/api/whatsapp/send-test")
def send_test_message(payload: TestMessageRequest):
    """
    Envía un mensaje de prueba al teléfono indicado para verificar la conexión activa.
    """
    logger.info(f"Enviando mensaje de prueba a {payload.telefono}")
    result = whatsapp_manager.enviar_mensaje(
        telefono_o_jid=payload.telefono,
        texto=payload.mensaje or "Mensaje de prueba desde MedCRM"
    )
    if "error" in result and not result.get("guardado_db"):
        raise HTTPException(status_code=400, detail=result["error"])
    return result

@app.post("/api/whatsapp/send-media")
async def send_media_api(
    file: UploadFile = File(...),
    telefono: str = Form(...),
    conversacion_id: Optional[str] = Form(None),
    caption: Optional[str] = Form("")
):
    """
    Envía un archivo multimedia (imagen, PDF, estudio médico) al WhatsApp del paciente
    y lo almacena con metadatos estructurados en Supabase.
    """
    logger.info(f"API: Enviando archivo {file.filename} a {telefono} (conversacion_id: {conversacion_id})")
    try:
        content = await file.read()
        mime_type = file.content_type or "application/octet-stream"
        original_name = file.filename or "archivo"
        
        subfolder = "images" if "image" in mime_type else "documents"
        tipo = "imagen" if "image" in mime_type else "audio" if "audio" in mime_type else "documento"
        media_type_baileys = "image" if "image" in mime_type else "audio" if "audio" in mime_type else "document"

        # Guardar en almacenamiento estático
        saved = media_service.save_media_bytes(
            data=content,
            subfolder=subfolder,
            mime_type=mime_type,
            original_filename=original_name,
            prefix="crm_out"
        )
        
        media_url_final = saved["media_url"]

        # Intentar subir a Supabase Storage Bucket whatsapp-media
        if supabase:
            try:
                storage_path = f"media/{int(time.time())}_{original_name}"
                res = supabase.storage.from_("whatsapp-media").upload(
                    file=content,
                    path=storage_path,
                    file_options={"content-type": mime_type, "upsert": "true"}
                )
                public_res = supabase.storage.from_("whatsapp-media").get_public_url(storage_path)
                if public_res:
                    media_url_final = public_res
            except Exception as sup_err:
                logger.warning(f"No se pudo subir a Supabase Storage, usando URL local: {sup_err}")

        # Enviar vía WhatsApp Baileys
        result = whatsapp_manager.enviar_multimedia(
            telefono=telefono,
            media_url=media_url_final,
            media_type=media_type_baileys,
            caption=caption or "",
            filename=original_name
        )
        
        # Guardar mensaje saliente del operador en Supabase
        if conversacion_id:
            try:
                guardar_mensaje(
                    conversacion_id=conversacion_id,
                    emisor="operador",
                    contenido=caption or original_name,
                    metadata_json={
                        "tipo": tipo,
                        "media_url": media_url_final,
                        "relative_url": saved.get("relative_url"),
                        "file_name": original_name,
                        "mime_type": mime_type,
                        "file_size_bytes": len(content),
                        "caption": caption or "",
                        "delivery_status": "enviado"
                    }
                )
            except Exception as db_save_err:
                logger.error(f"Error guardando mensaje multimedia en BD: {db_save_err}")

        return {
            "success": True,
            "media": {
                **saved,
                "media_url": media_url_final,
                "tipo": tipo
            },
            "whatsapp_result": result
        }
    except Exception as e:
        logger.error(f"Error enviando archivo multimedia: {e}")
        raise HTTPException(status_code=500, detail=f"Error al procesar archivo: {str(e)}")

# ====================================================================
# ENDPOINTS DE CONFIGURACIÓN DEL SISTEMA Y BOT
# ====================================================================

@app.get("/api/settings")
def get_system_settings():
    """
    Obtiene la configuración actual del consultorio, bot de IA y reglas de escalamiento.
    """
    return load_settings()

@app.post("/api/settings")
def update_system_settings(payload: Dict[str, Any] = Body(...)):
    """
    Actualiza y persiste la configuración del consultorio y bot de IA.
    """
    try:
        updated = save_settings(payload)
        return {"success": True, "settings": updated}
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error actualizando configuraciones: {str(e)}")

# ====================================================================
# ENDPOINTS DEL SISTEMA MULTI-AGENTE (PROMPT LAYERING & PERSONAS)
# ====================================================================

class GlobalDirectivesUpdate(BaseModel):
    nombre_clinica: Optional[str] = None
    tono_general: Optional[str] = None
    guardrails_medicos: Optional[str] = None
    politica_escalamiento: Optional[str] = None
    politica_turnos: Optional[str] = None
    politica_presupuestos: Optional[str] = None
    agente_defecto_codigo: Optional[str] = None

class SituationalAgentCreateUpdate(BaseModel):
    codigo: str
    nombre: str
    descripcion: Optional[str] = ""
    activo: Optional[bool] = True
    temperatura: Optional[float] = 0.2
    directiva_particular: str
    herramientas_habilitadas: Optional[List[str]] = ["buscar_disponibilidad_turnos", "crear_borrador_presupuesto", "escalar_a_operador_humano"]
    criterios_activacion: Optional[Any] = []
    orden: Optional[int] = 0

class AgentSimulatorRequest(BaseModel):
    mensaje: str
    agente_codigo: Optional[str] = "AUTO"
    paciente_nombre: Optional[str] = None
    paciente_etapa: Optional[str] = "CONSULTA_GENERAL"
    medico_asignado: Optional[str] = None

class AssignAgentConversationRequest(BaseModel):
    agente_codigo: str

@app.get("/api/agentes/config")
def get_multiagent_config():
    """
    Retorna la configuración completa del sistema multi-agente:
    directivas globales de la clínica, lista de agentes situacionales y herramientas disponibles.
    """
    try:
        globales = orchestrator.get_global_directives()
        agentes = list(orchestrator.get_all_agents().values())
        tools_list = list(AVAILABLE_TOOLS_MAP.keys())
        return {
            "success": True,
            "globales": globales,
            "agentes": agentes,
            "available_tools": tools_list
        }
    except Exception as e:
        logger.error(f"Error obteniendo configuración multi-agente: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))

@app.put("/api/agentes/globales")
def update_global_directives(payload: GlobalDirectivesUpdate):
    """
    Actualiza las directivas globales y guardrails médicos de la clínica en Supabase.
    """
    if not supabase:
        raise HTTPException(status_code=503, detail="Supabase no configurado.")
    try:
        data_to_update = {k: v for k, v in payload.dict().items() if v is not None}
        data_to_update["updated_at"] = "now()"
        
        # Verificar si existe el registro singleton
        check = supabase.table("agentes_directivas_globales").select("id").limit(1).execute()
        if check.data and len(check.data) > 0:
            rec_id = check.data[0]["id"]
            resp = supabase.table("agentes_directivas_globales").update(data_to_update).eq("id", rec_id).execute()
        else:
            resp = supabase.table("agentes_directivas_globales").insert(data_to_update).execute()
        
        orchestrator.invalidate_cache()
        log_event(
            nivel="INFO",
            modulo="SISTEMA",
            accion="CONFIGURACION_AGENTES",
            mensaje="Directivas globales de la clínica actualizadas.",
            detalles=data_to_update
        )
        return {"success": True, "globales": orchestrator.get_global_directives()}
    except Exception as e:
        logger.error(f"Error actualizando directivas globales: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/agentes/situacionales")
def create_situational_agent(payload: SituationalAgentCreateUpdate):
    """
    Crea un nuevo agente situacional en Supabase.
    """
    if not supabase:
        raise HTTPException(status_code=503, detail="Supabase no configurado.")
    try:
        clean_code = payload.codigo.strip().upper().replace(" ", "_")
        agent_data = payload.dict()
        agent_data["codigo"] = clean_code
        agent_data["updated_at"] = "now()"
        
        resp = supabase.table("agentes_situacionales").insert(agent_data).execute()
        orchestrator.invalidate_cache()
        
        log_event(
            nivel="INFO",
            modulo="SISTEMA",
            accion="CREAR_AGENTE_SITUACIONAL",
            mensaje=f"Agente situacional creado: {payload.nombre} ({clean_code})",
            detalles=agent_data
        )
        return {"success": True, "agente": resp.data[0] if resp.data else agent_data}
    except Exception as e:
        logger.error(f"Error creando agente situacional: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))

@app.put("/api/agentes/situacionales/{agente_id}")
def update_situational_agent(agente_id: str, payload: SituationalAgentCreateUpdate):
    """
    Actualiza las pautas, herramientas o estado de un agente situacional existente.
    """
    if not supabase:
        raise HTTPException(status_code=503, detail="Supabase no configurado.")
    try:
        agent_data = {k: v for k, v in payload.dict().items() if v is not None}
        if "codigo" in agent_data:
            agent_data["codigo"] = agent_data["codigo"].strip().upper().replace(" ", "_")
        agent_data["updated_at"] = "now()"

        resp = supabase.table("agentes_situacionales").update(agent_data).eq("id", agente_id).execute()
        orchestrator.invalidate_cache()
        
        log_event(
            nivel="INFO",
            modulo="SISTEMA",
            accion="ACTUALIZAR_AGENTE_SITUACIONAL",
            mensaje=f"Agente situacional actualizado: {payload.nombre}",
            detalles=agent_data
        )
        return {"success": True, "agente": resp.data[0] if resp.data else agent_data}
    except Exception as e:
        logger.error(f"Error actualizando agente situacional: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/api/agentes/situacionales/{agente_id}")
def delete_situational_agent(agente_id: str):
    """
    Elimina un agente situacional de Supabase.
    """
    if not supabase:
        raise HTTPException(status_code=503, detail="Supabase no configurado.")
    try:
        resp = supabase.table("agentes_situacionales").delete().eq("id", agente_id).execute()
        orchestrator.invalidate_cache()
        return {"success": True, "deleted_id": agente_id}
    except Exception as e:
        logger.error(f"Error eliminando agente situacional: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/agentes/simulador")
def simulate_agent_prompt(payload: AgentSimulatorRequest):
    """
    Ejecuta una prueba en vivo con un perfil y mensaje dado sin persistir mensajes ni modificar WhatsApp.
    Permite verificar cómo responde cada agente situacional en tiempo real.
    """
    t_start = time.time()
    try:
        import uuid
        temp_conv_id = str(uuid.uuid4())
        override_code = None if payload.agente_codigo in ["AUTO", "", None] else payload.agente_codigo
        
        # Inyectar mock de paciente si fue provisto
        mock_paciente_id = None
        
        respuesta = procesar_mensaje_agente(
            conversacion_id=temp_conv_id,
            mensaje_texto_o_paciente_id=payload.mensaje,
            guardar_en_db=False,
            agente_override_codigo=override_code
        )
        
        duracion = int((time.time() - t_start) * 1000)
        
        # Determinar qué agente resolvió la consulta
        if override_code:
            ag_usado = orchestrator.get_agent_by_code(override_code)
        else:
            ag_usado = orchestrator.determine_active_agent(mensaje_texto=payload.mensaje)

        return {
            "success": True,
            "respuesta": respuesta,
            "agente_utilizado": {
                "codigo": ag_usado.get("codigo"),
                "nombre": ag_usado.get("nombre"),
                "temperatura": ag_usado.get("temperatura")
            },
            "duracion_ms": duracion
        }
    except Exception as e:
        logger.error(f"Error en simulador agéntico: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))

@app.patch("/api/conversaciones/{conversacion_id}/agente")
def assign_agent_to_conversation(conversacion_id: str, payload: AssignAgentConversationRequest):
    """
    Asigna manualmente un agente situacional a una conversación específica (o 'AUTO' para enrutamiento inteligente).
    """
    if not supabase:
        raise HTTPException(status_code=503, detail="Supabase no configurado.")
    try:
        resp = supabase.table("conversaciones").update({"agente_asignado_codigo": payload.agente_codigo}).eq("id", conversacion_id).execute()
        return {"success": True, "agente_asignado_codigo": payload.agente_codigo}
    except Exception as e:
        logger.error(f"Error asignando agente a conversación {conversacion_id}: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))

# ====================================================================
# ENDPOINTS DE CONVERSACIONES, PRESUPUESTOS Y SIMULADOR
# ====================================================================

@app.post("/api/conversaciones/{conversacion_id}/toggle-bot")
def toggle_bot(conversacion_id: str, payload: ToggleBotRequest):
    """
    Habilita o deshabilita la atención del bot automático en una conversación específica.
    """
    logger.info(f"Alternar bot en conversación {conversacion_id} a {payload.bot_disabled}")
    res = actualizar_bot_disabled(conversacion_id, payload.bot_disabled)
    if not res:
        raise HTTPException(status_code=404, detail="Conversación no encontrada.")
    return {"success": True, "conversacion": res}

@app.post("/api/simulate-message")
def simulate_message(payload: SimuladorMensaje):
    """
    Simula un mensaje de WhatsApp entrante para desarrollo/pruebas.
    """
    logger.info(f"Simulando mensaje de {payload.telefono}: {payload.mensaje}")
    
    class MockSender:
        def __init__(self, sender):
            self.User = sender
            self.Sender = sender

    class MockInfo:
        def __init__(self, sender):
            self.Sender = sender
            self.PushName = f"Paciente Test ({sender[-4:]})"

    class MockConversationMessage:
        def __init__(self, text):
            self.conversation = text
            self.extendedTextMessage = None

    class MockEvent:
        def __init__(self, sender, text):
            self.Info = MockInfo(sender)
            self.Message = MockConversationMessage(text)

    class MockClient:
        def send_message(self, jid, text):
            logger.info(f"[WHATSAPP SIMULADO] ENVIANDO A {jid}: {text}")
            return {"status": "sent", "to": str(jid), "content": text}

    mock_client = MockClient()
    mock_event = MockEvent(payload.telefono, payload.mensaje)
    
    # Delegar al procesador de mensajes
    whatsapp_manager._handle_incoming_message(mock_client, mock_event)
    
    return {
        "success": True,
        "mensaje": f"Mensaje procesado en simulación para {payload.telefono}."
    }

@app.post("/api/presupuestos")
def create_presupuesto_api(payload: Dict[str, Any] = Body(...)):
    """
    Crea un presupuesto para un paciente y genera el PDF correspondiente con soporte multi-moneda centralizado.
    """
    logger.info(f"API: Crear presupuesto para paciente {payload.get('paciente_id')}")
    try:
        paciente_id = payload.get("paciente_id")
        if not paciente_id:
            raise HTTPException(status_code=400, detail="El ID del paciente es obligatorio.")
        
        items_raw = payload.get("items", [])
        if not items_raw:
            raise HTTPException(status_code=400, detail="Debe incluir al menos una prestación médica.")
            
        items_parsed = []
        for it in items_raw:
            items_parsed.append({
                "codigo": it.get("codigo") or it.get("codigo_servicio") or "PRACT",
                "nombre": it.get("nombre") or it.get("nombre_prestacion") or "Prestación Médica",
                "cantidad": int(it.get("cantidad", 1)),
                "precio_unitario": float(it.get("precio_unitario", 0.0)),
                "moneda": str(it.get("moneda", "ARS")).upper()
            })
            
        presupuesto = crear_presupuesto_rapido({
            "paciente_id": paciente_id,
            "asesoria_id": payload.get("asesoria_id"),
            "estado": payload.get("estado", "enviado"),
            "items": items_parsed
        })
        
        return {
            "success": True,
            "mensaje": "Presupuesto médico y PDF generados correctamente.",
            "presupuesto": presupuesto
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error al crear presupuesto: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/presupuestos/plantilla-preview")
def preview_plantilla_presupuesto():
    """
    Genera un PDF de muestra al vuelo para previsualizar el diseño de la plantilla institucional.
    """
    try:
        from app.services.pdf_service import generar_pdf_presupuesto
        from datetime import date
        
        sample_presupuesto = {
            "id": "A1B2C3D4",
            "created_at": date.today().isoformat(),
            "estado": "borrador",
            "total": 30500.00
        }
        sample_paciente = {
            "nombre": "GARCÍA, MARÍA JOSÉ",
            "telefono": "+54 9 11 4444-5555",
            "email": "maria.garcia@ejemplo.com",
            "dni": "35.890.123",
            "obra_social": "OSDE 310"
        }
        sample_items = [
            {
                "codigo": "420101",
                "nombre_prestacion": "Consulta Médica Especializada en Fertilidad",
                "cantidad": 1,
                "precio_unitario": 8500.00,
                "moneda": "ARS",
                "subtotal": 8500.00
            },
            {
                "codigo": "180104",
                "nombre_prestacion": "Ecografía Ginecológica Transvaginal de Alta Resolución",
                "cantidad": 1,
                "precio_unitario": 22000.00,
                "moneda": "ARS",
                "subtotal": 22000.00
            },
            {
                "codigo": "FIV-ICSI-01",
                "nombre_prestacion": "Tratamiento FIV + ICSI Completo con Criopreservación",
                "cantidad": 1,
                "precio_unitario": 1500.00,
                "moneda": "USD",
                "subtotal": 1500.00
            }
        ]
        
        filename = generar_pdf_presupuesto(sample_presupuesto, sample_paciente, sample_items)
        return {
            "success": True,
            "pdf_url": f"/static/{filename}"
        }
    except Exception as e:
        logger.error(f"Error al generar preview de plantilla: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# ====================================================================
# ENDPOINTS DE INTEGRACIÓN CON GECLISA (HOSPITALARIO)
# ====================================================================

@app.get("/api/geclisa/pacientes/buscar-por-dni")
def buscar_geclisa_dni(dni: str):
    """
    Busca al paciente en Geclisa por número de DNI y verifica si ya está registrado en el CRM.
    """
    if not dni:
        raise HTTPException(status_code=400, detail="Debe ingresar un número de DNI.")
        
    resultado = geclisa_client.buscar_paciente_por_dni(dni)
    if not resultado.get("encontrado"):
        return resultado

    # Verificar si ya existe en Supabase
    ficha_id = resultado.get("ficha_id")
    dni_resultado = resultado.get("dni") or dni
    
    paciente_crm = None
    if ficha_id:
        paciente_crm = get_paciente_by_geclisa_id(ficha_id)
    if not paciente_crm and dni_resultado:
        paciente_crm = get_paciente_by_dni(dni_resultado)

    resultado["ya_en_crm"] = bool(paciente_crm)
    resultado["crm_paciente_id"] = paciente_crm.get("id") if paciente_crm else None
    return resultado

@app.get("/api/geclisa/pacientes/buscar-por-ficha")
def buscar_geclisa_ficha(ficha_id: int):
    """
    Busca la ficha en Geclisa directamente por fichaId y verifica si ya está registrada en el CRM.
    """
    if not ficha_id:
        raise HTTPException(status_code=400, detail="Debe ingresar un ID de ficha válido.")
        
    resultado = geclisa_client.buscar_paciente_por_ficha(ficha_id)
    if not resultado.get("encontrado"):
        return resultado

    # Verificar si ya existe en Supabase
    paciente_crm = get_paciente_by_geclisa_id(ficha_id)
    if not paciente_crm and resultado.get("dni"):
        paciente_crm = get_paciente_by_dni(resultado.get("dni"))

    resultado["ya_en_crm"] = bool(paciente_crm)
    resultado["crm_paciente_id"] = paciente_crm.get("id") if paciente_crm else None
    return resultado

@app.get("/api/geclisa/turnos/pendientes/{ficha_id}")
def obtener_turnos_pendientes_geclisa(ficha_id: int):
    """
    Consulta en tiempo real a Geclisa los turnos pendientes/agendados de una ficha de paciente (GET /api/Turnos/pendientes/{fichaId}).
    No almacena en la base de datos local.
    """
    try:
        resultado = geclisa_client.obtener_turnos_pendientes_ficha(ficha_id)
        return resultado
    except Exception as e:
        logger.error(f"Error al obtener turnos pendientes de Geclisa para ficha {ficha_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/geclisa/pacientes/importar")
def importar_paciente_geclisa(payload: Dict[str, Any] = Body(...)):
    """
    Importa o actualiza un paciente desde Geclisa hacia la base de datos del CRM (Supabase).
    """
    if not payload:
        raise HTTPException(status_code=400, detail="El payload con los datos del paciente está vacío.")
        
    try:
        paciente = crear_o_actualizar_paciente_geclisa(payload)
        if not paciente or not isinstance(paciente, dict):
            raise HTTPException(status_code=500, detail="No se pudo registrar el paciente en la base de datos Supabase.")
            
        nombre = paciente.get("nombre", "importado")
        return {
            "success": True,
            "mensaje": f"Paciente {nombre} importado correctamente.",
            "paciente": paciente
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error al importar paciente desde Geclisa: {e}")
        raise HTTPException(status_code=500, detail=f"Error al importar paciente: {str(e)}")

@app.post("/api/geclisa/pacientes/sincronizar/{paciente_id}")
def sincronizar_paciente_geclisa(paciente_id: str):
    """
    Consulta en vivo a Geclisa y actualiza los datos del paciente en Supabase.
    """
    if not supabase:
        raise HTTPException(status_code=500, detail="Base de datos Supabase no conectada.")
        
    try:
        # 1. Obtener paciente de Supabase
        res = supabase.table("pacientes").select("*").eq("id", paciente_id).execute()
        if not res.data or len(res.data) == 0:
            raise HTTPException(status_code=404, detail="Paciente no encontrado en el CRM.")
            
        paciente_actual = res.data[0]
        ficha_id = paciente_actual.get("geclisa_ficha_id")
        dni = paciente_actual.get("dni")
        
        datos_geclisa = None
        if ficha_id:
            logger.info(f"Sincronizando paciente {paciente_id} por fichaId {ficha_id}...")
            datos_geclisa = geclisa_client.buscar_paciente_por_ficha(int(ficha_id))
        elif dni:
            logger.info(f"Sincronizando paciente {paciente_id} por DNI {dni}...")
            datos_geclisa = geclisa_client.buscar_paciente_por_dni(str(dni))
            
        if not datos_geclisa or not datos_geclisa.get("encontrado"):
            msg = datos_geclisa.get("mensaje") if datos_geclisa else "El paciente no posee Ficha ID ni DNI para consultar en Geclisa."
            raise HTTPException(status_code=404, detail=msg or "No se encontraron datos en Geclisa para este paciente.")

        # 2. Mezclar datos preservando campos propios del CRM si no vienen de Geclisa
        payload_actualizado = {
            "id": paciente_id,
            "geclisa_ficha_id": datos_geclisa.get("ficha_id") or ficha_id,
            "dni": datos_geclisa.get("dni") or dni,
            "nombre_completo": datos_geclisa.get("nombre_completo") or paciente_actual.get("nombre"),
            "nombre": datos_geclisa.get("nombre") or paciente_actual.get("nombre"),
            "telefono": datos_geclisa.get("telefono") or paciente_actual.get("telefono"),
            "telefono_fijo": datos_geclisa.get("telefono_fijo") or paciente_actual.get("telefono_fijo"),
            "email": datos_geclisa.get("email") or paciente_actual.get("email"),
            "nro_hc": datos_geclisa.get("nro_hc") or paciente_actual.get("nro_hc"),
            "obra_social": datos_geclisa.get("obra_social") or paciente_actual.get("obra_social"),
            "plan_cobertura": datos_geclisa.get("plan_cobertura") or paciente_actual.get("plan_cobertura"),
            "direccion": datos_geclisa.get("direccion") or paciente_actual.get("direccion"),
            "fecha_nacimiento": datos_geclisa.get("fecha_nacimiento") or paciente_actual.get("fecha_nacimiento"),
            "sexo": datos_geclisa.get("sexo") or paciente_actual.get("sexo"),
            "medico_cabecera": paciente_actual.get("medico_cabecera"),
            "medico_cabecera_id": paciente_actual.get("medico_cabecera_id"),
            "medico_cabecera_nombre": paciente_actual.get("medico_cabecera_nombre"),
            "medico_cabecera_matricula": paciente_actual.get("medico_cabecera_matricula"),
            "medico_cabecera_especialidad": paciente_actual.get("medico_cabecera_especialidad"),
        }

        paciente_actualizado = crear_o_actualizar_paciente_geclisa(payload_actualizado)
        return {
            "success": True,
            "mensaje": f"Expediente de {paciente_actualizado.get('nombre')} sincronizado con Geclisa.",
            "paciente": paciente_actualizado
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error al sincronizar paciente con Geclisa: {e}")
        raise HTTPException(status_code=500, detail=f"Error al sincronizar paciente: {str(e)}")

@app.get("/api/geclisa/pacientes/{paciente_id}/historia-clinica")
def obtener_historia_clinica_paciente(paciente_id: str):
    """
    Consulta en vivo las evoluciones de la historia clínica en Geclisa.
    Soporta paciente_id como UUID del CRM, DNI o Ficha ID directamente.
    Si el paciente no tiene 'geclisa_ficha_id' pero tiene 'dni', busca en Geclisa por DNI,
    asocia la ficha encontrada en Supabase y consulta la historia clínica.
    No persiste la historia clínica en Supabase (operación 100% de lectura on-demand).
    """
    try:
        ficha_id = None
        dni = None
        paciente_nombre = None
        paciente_crm_id = None

        # 1. Intentar buscar en Supabase (por UUID, DNI o ficha_id)
        if supabase:
            try:
                res_paciente = None
                # Búsqueda por UUID solo si tiene formato UUID
                if len(str(paciente_id)) == 36 and '-' in str(paciente_id):
                    res_paciente = supabase.table("pacientes").select("*").eq("id", paciente_id).execute()

                if not res_paciente or not res_paciente.data:
                    # Búsqueda por DNI
                    res_paciente = supabase.table("pacientes").select("*").ilike("dni", f"%{paciente_id}%").execute()

                if not res_paciente or not res_paciente.data:
                    # Búsqueda por geclisa_ficha_id si es numérico
                    if str(paciente_id).isdigit():
                        res_paciente = supabase.table("pacientes").select("*").eq("geclisa_ficha_id", int(paciente_id)).execute()

                if res_paciente and res_paciente.data and len(res_paciente.data) > 0:
                    paciente = res_paciente.data[0]
                    paciente_crm_id = paciente.get("id")
                    ficha_id = paciente.get("geclisa_ficha_id")
                    dni = paciente.get("dni")
                    paciente_nombre = paciente.get("nombre")
            except Exception as db_err:
                logger.warning(f"Aviso al consultar paciente en Supabase: {db_err}")

        # Si no se encontró en Supabase pero paciente_id parece ser un DNI o Ficha directa
        if not dni and not ficha_id:
            if str(paciente_id).isdigit():
                if len(str(paciente_id)) >= 7:
                    dni = str(paciente_id)
                else:
                    ficha_id = int(paciente_id)

        # 2. Si no tiene ficha_id pero tiene DNI, resolver por DNI contra Geclisa
        if not ficha_id and dni:
            dni_limpio = "".join(filter(str.isdigit, str(dni)))
            if dni_limpio:
                datos_dni = geclisa_client.buscar_paciente_por_dni(dni_limpio)
                if datos_dni.get("encontrado") and datos_dni.get("ficha_id"):
                    ficha_id = int(datos_dni["ficha_id"])
                    if not paciente_nombre:
                        paciente_nombre = datos_dni.get("nombre_completo")
                    # Guardar ficha_id en Supabase si tenemos el ID del paciente en CRM
                    if paciente_crm_id and supabase:
                        try:
                            supabase.table("pacientes").update({"geclisa_ficha_id": ficha_id}).eq("id", paciente_crm_id).execute()
                            logger.info(f"Ficha #{ficha_id} auto-vinculada por DNI para paciente {paciente_crm_id}")
                        except Exception as err_upd:
                            logger.warning(f"No se pudo auto-vincular ficha #{ficha_id} en Supabase: {err_upd}")

        # 3. Si aún no tenemos ficha_id, retornar aviso amigable (encontrado: False con HTTP 200)
        if not ficha_id:
            if not dni:
                return {
                    "encontrado": False,
                    "motivo": "sin_dni",
                    "mensaje": "El paciente no posee número de DNI ni Ficha Geclisa registrada en el CRM."
                }
            else:
                return {
                    "encontrado": False,
                    "motivo": "sin_ficha_geclisa",
                    "mensaje": f"No se encontró ninguna historia clínica ni ficha activa en Geclisa para el DNI {dni}."
                }

        # 4. Consultar Historia Clínica Resumen en Geclisa
        resultado_hc = geclisa_client.obtener_historia_clinica_resumen(int(ficha_id))
        if not resultado_hc.get("encontrado"):
            msg = resultado_hc.get("mensaje") or "No se encontraron evoluciones en Geclisa."
            return {
                "success": False,
                "encontrado": False,
                "motivo": "no_encontrado_geclisa",
                "ficha_id": ficha_id,
                "mensaje": msg,
                "detail": msg
            }

        evoluciones_lista = resultado_hc.get("evoluciones_recientes", [])
        return {
            "success": True,
            "encontrado": True,
            "paciente_id": paciente_crm_id or paciente_id,
            "ficha_id": ficha_id,
            "paciente_nombre": paciente_nombre,
            "paciente_dni": dni,
            "fecha_generacion": resultado_hc.get("fecha_generacion"),
            "evoluciones": evoluciones_lista,
            "evoluciones_recientes": evoluciones_lista,
            "total_evoluciones": len(evoluciones_lista),
            "data": {
                "evoluciones": evoluciones_lista,
                "evoluciones_recientes": evoluciones_lista,
                "total_evoluciones": len(evoluciones_lista),
                "paciente_nombre": paciente_nombre,
                "ficha_id": ficha_id
            }
        }

    except Exception as e:
        logger.error(f"Error al obtener historia clínica para paciente {paciente_id}: {e}")
        return {
            "success": False,
            "encontrado": False,
            "motivo": "error_servidor",
            "mensaje": f"Error al consultar historia clínica en Geclisa: {str(e)}",
            "detail": str(e)
        }

@app.get("/api/geclisa/pacientes/{paciente_id}/indicaciones")
def obtener_indicaciones_paciente(paciente_id: str):
    """
    Consulta en vivo las indicaciones médicas, protocolos de medicación y recetas
    en Geclisa para el paciente indicado. Operación 100% de lectura on-demand.
    """
    try:
        ficha_id = None
        dni = None
        paciente_nombre = None
        paciente_crm_id = None

        # 1. Intentar buscar en Supabase (por UUID, DNI o ficha_id)
        if supabase:
            try:
                res_paciente = None
                if len(str(paciente_id)) == 36 and '-' in str(paciente_id):
                    res_paciente = supabase.table("pacientes").select("*").eq("id", paciente_id).execute()

                if not res_paciente or not res_paciente.data:
                    res_paciente = supabase.table("pacientes").select("*").ilike("dni", f"%{paciente_id}%").execute()

                if not res_paciente or not res_paciente.data:
                    if str(paciente_id).isdigit():
                        res_paciente = supabase.table("pacientes").select("*").eq("geclisa_ficha_id", int(paciente_id)).execute()

                if res_paciente and res_paciente.data and len(res_paciente.data) > 0:
                    paciente = res_paciente.data[0]
                    paciente_crm_id = paciente.get("id")
                    ficha_id = paciente.get("geclisa_ficha_id")
                    dni = paciente.get("dni")
                    paciente_nombre = paciente.get("nombre")
            except Exception as db_err:
                logger.warning(f"Aviso al consultar paciente en Supabase: {db_err}")

        # Fallback si paciente_id es directamente DNI o Ficha
        if not dni and not ficha_id:
            if str(paciente_id).isdigit():
                if len(str(paciente_id)) >= 7:
                    dni = str(paciente_id)
                else:
                    ficha_id = int(paciente_id)

        # 2. Si no tiene ficha_id pero tiene DNI, resolver por DNI contra Geclisa
        if not ficha_id and dni:
            dni_limpio = "".join(filter(str.isdigit, str(dni)))
            if dni_limpio:
                datos_dni = geclisa_client.buscar_paciente_por_dni(dni_limpio)
                if datos_dni.get("encontrado") and datos_dni.get("ficha_id"):
                    ficha_id = int(datos_dni["ficha_id"])
                    if not paciente_nombre:
                        paciente_nombre = datos_dni.get("nombre_completo")
                    if paciente_crm_id and supabase:
                        try:
                            supabase.table("pacientes").update({"geclisa_ficha_id": ficha_id}).eq("id", paciente_crm_id).execute()
                        except Exception as err_upd:
                            logger.warning(f"No se pudo auto-vincular ficha #{ficha_id} en Supabase: {err_upd}")

        if not ficha_id:
            if not dni:
                msg = "El paciente no posee número de DNI ni Ficha Geclisa registrada en el CRM."
                return {
                    "success": False,
                    "encontrado": False,
                    "motivo": "sin_dni",
                    "mensaje": msg,
                    "detail": msg
                }
            else:
                msg = f"No se encontró ninguna ficha activa en Geclisa para el DNI {dni}."
                return {
                    "success": False,
                    "encontrado": False,
                    "motivo": "sin_ficha_geclisa",
                    "mensaje": msg,
                    "detail": msg
                }

        # 3. Consultar Indicaciones Médicas en Geclisa
        resultado_ind = geclisa_client.obtener_indicaciones_medicas(int(ficha_id))
        indicaciones_lista = resultado_ind.get("indicaciones", [])
        return {
            "success": True,
            "encontrado": True,
            "paciente_id": paciente_crm_id or paciente_id,
            "ficha_id": ficha_id,
            "paciente_nombre": paciente_nombre,
            "paciente_dni": dni,
            "indicaciones": indicaciones_lista,
            "recetas": indicaciones_lista,
            "total_indicaciones": len(indicaciones_lista),
            "data": {
                "indicaciones": indicaciones_lista,
                "recetas": indicaciones_lista,
                "total_indicaciones": len(indicaciones_lista),
                "paciente_nombre": paciente_nombre,
                "ficha_id": ficha_id
            }
        }

    except Exception as e:
        logger.error(f"Error al obtener indicaciones para paciente {paciente_id}: {e}")
        return {
            "success": False,
            "encontrado": False,
            "motivo": "error_servidor",
            "mensaje": f"Error al consultar indicaciones en Geclisa: {str(e)}",
            "detail": str(e)
        }

@app.put("/api/pacientes/{paciente_id}")
def actualizar_paciente_crm(paciente_id: str, payload: Dict[str, Any] = Body(...)):
    """
    Actualiza directamente los datos de un paciente en el CRM (Supabase).
    """
    if not supabase:
        raise HTTPException(status_code=500, detail="Base de datos Supabase no conectada.")
        
    try:
        from app.services.phone_normalizer import normalize_phone_number
        
        datos_actualizar = {}
        if "nombre" in payload and payload["nombre"]:
            datos_actualizar["nombre"] = str(payload["nombre"]).strip()
        if "dni" in payload:
            datos_actualizar["dni"] = str(payload["dni"]).strip() if payload["dni"] else None
        if "geclisa_ficha_id" in payload:
            datos_actualizar["geclisa_ficha_id"] = int(payload["geclisa_ficha_id"]) if payload["geclisa_ficha_id"] else None
        if "nro_hc" in payload:
            datos_actualizar["nro_hc"] = str(payload["nro_hc"]).strip() if payload["nro_hc"] else None
        if "telefono" in payload and payload["telefono"]:
            raw_tel = str(payload["telefono"]).strip()
            datos_actualizar["telefono"] = normalize_phone_number(raw_tel) if not raw_tel.startswith("temp_") else raw_tel
        if "telefono_fijo" in payload:
            datos_actualizar["telefono_fijo"] = str(payload["telefono_fijo"]).strip() if payload["telefono_fijo"] else None
        if "email" in payload:
            datos_actualizar["email"] = str(payload["email"]).strip() if payload["email"] else None
        if "obra_social" in payload:
            datos_actualizar["obra_social"] = str(payload["obra_social"]).strip() if payload["obra_social"] else None
        if "plan_cobertura" in payload:
            datos_actualizar["plan_cobertura"] = str(payload["plan_cobertura"]).strip() if payload["plan_cobertura"] else None
        if "medico_cabecera" in payload:
            datos_actualizar["medico_cabecera"] = str(payload["medico_cabecera"]).strip() if payload["medico_cabecera"] else None
        if "medico_cabecera_id" in payload:
            datos_actualizar["medico_cabecera_id"] = int(payload["medico_cabecera_id"]) if payload["medico_cabecera_id"] else None
        if "medico_cabecera_nombre" in payload:
            datos_actualizar["medico_cabecera_nombre"] = str(payload["medico_cabecera_nombre"]).strip() if payload["medico_cabecera_nombre"] else None
        if "medico_cabecera_matricula" in payload:
            datos_actualizar["medico_cabecera_matricula"] = str(payload["medico_cabecera_matricula"]).strip() if payload["medico_cabecera_matricula"] else None
        if "medico_cabecera_especialidad" in payload:
            datos_actualizar["medico_cabecera_especialidad"] = str(payload["medico_cabecera_especialidad"]).strip() if payload["medico_cabecera_especialidad"] else None
        if "direccion" in payload:
            datos_actualizar["direccion"] = str(payload["direccion"]).strip() if payload["direccion"] else None
        if "fecha_nacimiento" in payload:
            datos_actualizar["fecha_nacimiento"] = str(payload["fecha_nacimiento"]).strip() if payload["fecha_nacimiento"] else None
        if "sexo" in payload:
            datos_actualizar["sexo"] = str(payload["sexo"]).strip() if payload["sexo"] else None
        if "historial_notas" in payload:
            datos_actualizar["historial_notas"] = str(payload["historial_notas"]) if payload["historial_notas"] is not None else None

        res = supabase.table("pacientes").update(datos_actualizar).eq("id", paciente_id).execute()
        if not res.data or len(res.data) == 0:
            raise HTTPException(status_code=404, detail="Paciente no encontrado para actualizar.")

        paciente_actualizado = res.data[0]
        return {
            "success": True,
            "mensaje": f"Datos de {paciente_actualizado.get('nombre')} modificados correctamente.",
            "paciente": paciente_actualizado
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error al actualizar paciente {paciente_id}: {e}")
        raise HTTPException(status_code=500, detail=f"Error al actualizar paciente: {str(e)}")

@app.delete("/api/pacientes/{paciente_id}")
def eliminar_paciente_crm(paciente_id: str):
    """
    Elimina permanentemente a un paciente y todo su historial relacionado en cascada (conversaciones, mensajes, presupuestos).
    """
    if not supabase:
        raise HTTPException(status_code=500, detail="Base de datos Supabase no conectada.")
        
    try:
        # Verificar existencia
        check_p = supabase.table("pacientes").select("id, nombre").eq("id", paciente_id).execute()
        if not check_p.data or len(check_p.data) == 0:
            raise HTTPException(status_code=404, detail="Paciente no encontrado.")
            
        nombre_paciente = check_p.data[0].get("nombre", "Paciente")
        
        # Eliminar en Supabase (las claves foráneas en CASCADE borrarán conversaciones, mensajes y presupuestos)
        supabase.table("pacientes").delete().eq("id", paciente_id).execute()
        
        logger.info(f"Paciente {paciente_id} ({nombre_paciente}) eliminado con éxito en cascada.")
        return {
            "success": True,
            "mensaje": f"Expediente de {nombre_paciente} y todos sus registros asociados han sido eliminados correctamente."
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error al eliminar paciente {paciente_id}: {e}")
        raise HTTPException(status_code=500, detail=f"Error al eliminar paciente: {str(e)}")


@app.get("/api/geclisa/prestadores/buscar")
def buscar_geclisa_prestadores(query: Optional[str] = None, q: Optional[str] = None):
    """
    Busca prestadores médicos en Geclisa por nombre, apellido o número de matrícula.
    Operación a demanda activada exclusivamente por la búsqueda del usuario.
    Acepta tanto 'query' como 'q'.
    """
    termino = (query or q or "").strip()
    try:
        prestadores = geclisa_client.buscar_prestadores(termino)
        return {
            "success": True,
            "query": termino,
            "total": len(prestadores),
            "prestadores": prestadores
        }
    except Exception as e:
        logger.error(f"Error al consultar prestadores en Geclisa: {e}")
        raise HTTPException(status_code=500, detail=f"Error al consultar prestadores: {str(e)}")

@app.get("/api/geclisa/prestadores/{pre_id}")
def detalle_geclisa_prestador(pre_id: int):
    """
    Obtiene los datos detallados de un prestador específico por su preId de Geclisa.
    """
    try:
        prestador = geclisa_client.obtener_prestador_por_id(pre_id)
        return prestador
    except Exception as e:
        logger.error(f"Error al obtener detalle de prestador {pre_id}: {e}")
        raise HTTPException(status_code=500, detail=f"Error al consultar prestador: {str(e)}")

@app.patch("/api/pacientes/{paciente_id}/medico")
def asignar_medico_a_paciente(paciente_id: str, payload: Dict[str, Any] = Body(...)):
    """
    Asigna o actualiza el médico de cabecera en la ficha del paciente en el CRM (Supabase).
    Payload esperado: { "pre_id": int | null, "nombre": str | null, "matricula": str | null, "especialidad": str | null }
    """
    try:
        paciente_actualizado = asignar_medico_paciente(paciente_id, payload)
        if not paciente_actualizado:
            raise HTTPException(status_code=404, detail="Paciente no encontrado o no se pudo actualizar.")
        return {
            "success": True,
            "mensaje": "Médico de cabecera asignado correctamente.",
            "paciente": paciente_actualizado
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error al asignar médico a paciente {paciente_id}: {e}")
        raise HTTPException(status_code=500, detail=f"Error al asignar médico: {str(e)}")


# ====================================================================
# ENDPOINTS REST: NOMENCLADORES Y ARANCELES (GECLISA LIVE + CRM LOCAL)
# ====================================================================

@app.get("/api/geclisa/nomenclador/tipos")
def get_geclisa_nomenclador_tipos():
    """
    Obtiene los tipos de nomencladores configurados en Geclisa (GET /api/Nomenclador/tipos).
    """
    try:
        tipos = geclisa_client.obtener_tipos_nomenclador()
        return {"success": True, "tipos": tipos}
    except Exception as e:
        logger.error(f"Error al obtener tipos de nomenclador de Geclisa: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/geclisa/nomenclador/buscar")
def buscar_geclisa_nomenclador_practicas(nom_id: Optional[int] = None, q: Optional[str] = ""):
    """
    Busca prácticas en tiempo real en Geclisa por tipo de nomenclador y texto de búsqueda,
    cruzándolas con la base de datos del CRM para indicar si ya tienen arancel configurado.
    """
    try:
        raw_practicas = geclisa_client.buscar_practicas_nomenclador(nom_id=nom_id, search_string=q or "")
        enriquecidas = enriquecer_practicas_geclisa_con_crm(raw_practicas)
        return {
            "success": True,
            "total": len(enriquecidas),
            "practicas": enriquecidas
        }
    except Exception as e:
        logger.error(f"Error al buscar prácticas en Geclisa: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/nomenclador/guardar-practica-arancel")
def guardar_practica_arancel_api(payload: Dict[str, Any] = Body(...)):
    """
    Guarda o actualiza una práctica (desde Geclisa o Manual) y su arancel/vigencia en el CRM.
    """
    try:
        if not payload.get("codigo") or not payload.get("nombre"):
            raise HTTPException(status_code=400, detail="El código y nombre de la práctica son obligatorios.")
        res = guardar_practica_crm_con_arancel(payload)
        return {
            "success": True,
            "mensaje": f"Práctica {payload.get('codigo')} configurada correctamente en el CRM.",
            "resultado": res
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error al guardar práctica y arancel: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/nomenclador/practicas-configuradas")
def get_practicas_configuradas_crm(
    moneda: Optional[str] = None,
    origen: Optional[str] = None,
    q: Optional[str] = None
):
    """
    Lista el catálogo consolidado de prácticas configuradas en el CRM (Geclisa y Manuales).
    """
    try:
        practicas = listar_catalogo_completo_crm(filtro_moneda=moneda, filtro_origen=origen, q=q)
        return {
            "success": True,
            "total": len(practicas),
            "practicas": practicas
        }
    except Exception as e:
        logger.error(f"Error al listar prácticas configuradas: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/api/nomenclador/practicas-configuradas/{practica_id}")
def eliminar_practica_crm_api(practica_id: str):
    """
    Elimina una práctica del catálogo del CRM.
    """
    try:
        ok = eliminar_practica_crm(practica_id)
        return {"success": ok, "mensaje": "Práctica eliminada del catálogo del CRM."}
    except Exception as e:
        logger.error(f"Error al eliminar práctica {practica_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# ====================================================================
# GUARDADO INTEGRAL Y RESUMEN OPERATIVO MULTIDIMENSIONAL
# ====================================================================

@app.post("/api/nomenclador/guardar-practica-integral")
def guardar_practica_integral_api(payload: Dict[str, Any] = Body(...)):
    """
    Guarda o actualiza una práctica de manera integral con sus reglas multidimensionales:
    Aranceles temporales, Preparación (Plantilla o Custom), Consentimiento (Plantilla o Custom).
    """
    try:
        if not payload.get("codigo") or not payload.get("nombre"):
            raise HTTPException(status_code=400, detail="El código y nombre de la práctica son obligatorios.")
        res = guardar_practica_crm_integral(payload)
        return {
            "success": True,
            "mensaje": f"Práctica {payload.get('codigo')} guardada y configurada exitosamente.",
            "resultado": res
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error al guardar práctica integral: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/nomenclador/practica/{practica_id_or_codigo}/resumen-operativo")
def get_resumen_operativo_api(practica_id_or_codigo: str, fecha: Optional[str] = None):
    """
    Resuelve todos los aspectos operativos de una práctica para la fecha dada
    (arancel vigente, preparación resuelta y consentimiento resuelto).
    """
    try:
        res = get_practica_resumen_operativo(practica_id_or_codigo, fecha_consulta=fecha)
        if not res:
            raise HTTPException(status_code=404, detail="Práctica no encontrada en el CRM.")
        return {"success": True, "practica": res}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error al obtener resumen operativo: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# ====================================================================
# ENDPOINTS: PLANTILLAS MAESTRAS DE PREPARACIONES
# ====================================================================

@app.get("/api/nomenclador/plantillas/preparaciones")
def list_plantillas_preparaciones_api():
    try:
        data = get_plantillas_preparaciones()
        return {"success": True, "total": len(data), "plantillas": data}
    except Exception as e:
        logger.error(f"Error listando plantillas de preparaciones: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/nomenclador/plantillas/preparaciones")
def create_plantilla_preparacion_api(payload: Dict[str, Any] = Body(...)):
    try:
        if not payload.get("titulo") or not payload.get("texto_indicaciones"):
            raise HTTPException(status_code=400, detail="El título y las indicaciones son obligatorios.")
        item = create_plantilla_preparacion(payload)
        return {"success": True, "mensaje": "Plantilla de preparación creada con éxito.", "plantilla": item}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error creando plantilla de preparación: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.put("/api/nomenclador/plantillas/preparaciones/{plantilla_id}")
def update_plantilla_preparacion_api(plantilla_id: str, payload: Dict[str, Any] = Body(...)):
    try:
        item = update_plantilla_preparacion(plantilla_id, payload)
        return {"success": True, "mensaje": "Plantilla de preparación actualizada.", "plantilla": item}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error actualizando plantilla de preparación: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/api/nomenclador/plantillas/preparaciones/{plantilla_id}")
def delete_plantilla_preparacion_api(plantilla_id: str):
    try:
        ok = delete_plantilla_preparacion(plantilla_id)
        return {"success": ok, "mensaje": "Plantilla de preparación eliminada."}
    except Exception as e:
        logger.error(f"Error eliminando plantilla de preparación: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# ====================================================================
# ENDPOINTS: PLANTILLAS MAESTRAS DE CONSENTIMIENTOS INFORMADOS
# ====================================================================

@app.get("/api/nomenclador/plantillas/consentimientos")
def list_plantillas_consentimientos_api():
    try:
        data = get_plantillas_consentimientos()
        return {"success": True, "total": len(data), "plantillas": data}
    except Exception as e:
        logger.error(f"Error listando plantillas de consentimientos: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/nomenclador/plantillas/consentimientos")
def create_plantilla_consentimiento_api(payload: Dict[str, Any] = Body(...)):
    try:
        if not payload.get("titulo") or not payload.get("cuerpo_legal"):
            raise HTTPException(status_code=400, detail="El título y el cuerpo legal son obligatorios.")
        item = create_plantilla_consentimiento(payload)
        return {"success": True, "mensaje": "Plantilla de consentimiento creada con éxito.", "plantilla": item}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error creando plantilla de consentimiento: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.put("/api/nomenclador/plantillas/consentimientos/{plantilla_id}")
def update_plantilla_consentimiento_api(plantilla_id: str, payload: Dict[str, Any] = Body(...)):
    try:
        item = update_plantilla_consentimiento(plantilla_id, payload)
        return {"success": True, "mensaje": "Plantilla de consentimiento actualizada.", "plantilla": item}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error actualizando plantilla de consentimiento: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/api/nomenclador/plantillas/consentimientos/{plantilla_id}")
def delete_plantilla_consentimiento_api(plantilla_id: str):
    try:
        ok = delete_plantilla_consentimiento(plantilla_id)
        return {"success": ok, "mensaje": "Plantilla de consentimiento eliminada."}
    except Exception as e:
        logger.error(f"Error eliminando plantilla de consentimiento: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# ====================================================================
# ENDPOINTS: HISTORIAL Y CRUD DE ARANCELES (1:N) POR PRÁCTICA
# ====================================================================

@app.get("/api/nomenclador/practicas/{practica_id}/aranceles")
def get_aranceles_practica_api(practica_id: str):
    try:
        data = get_aranceles_por_practica(practica_id)
        return {"success": True, "total": len(data), "aranceles": data}
    except Exception as e:
        logger.error(f"Error al obtener aranceles de la práctica {practica_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/nomenclador/practicas/{practica_id}/aranceles")
def crear_arancel_practica_api(practica_id: str, payload: Dict[str, Any] = Body(...)):
    try:
        item = crear_arancel_practica(practica_id, payload)
        return {"success": True, "mensaje": "Tarifa agregada al historial con éxito.", "arancel": item}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error al crear arancel para práctica {practica_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.put("/api/nomenclador/aranceles/{arancel_id}")
def actualizar_arancel_api(arancel_id: str, payload: Dict[str, Any] = Body(...)):
    try:
        item = actualizar_arancel(arancel_id, payload)
        return {"success": True, "mensaje": "Tarifa actualizada con éxito.", "arancel": item}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error al actualizar arancel {arancel_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/api/nomenclador/aranceles/{arancel_id}")
def eliminar_arancel_api(arancel_id: str):
    try:
        ok = eliminar_arancel(arancel_id)
        return {"success": ok, "mensaje": "Tarifa eliminada del historial."}
    except Exception as e:
        logger.error(f"Error al eliminar arancel {arancel_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/nomencladores")
def get_all_nomencladores():
    """
    Lista todos los nomencladores configurados en el CRM.
    """
    try:
        data = list_nomencladores()
        return {"success": True, "nomencladores": data}
    except Exception as e:
        logger.error(f"Error al obtener nomencladores: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/nomencladores")
def crear_nuevo_nomenclador(payload: Dict[str, Any] = Body(...)):
    """
    Crea un nuevo nomenclador en el CRM (con moneda ARS o USD).
    """
    try:
        if not payload.get("nombre"):
            raise HTTPException(status_code=400, detail="El nombre del nomenclador es obligatorio.")
        creado = create_nomenclador(payload)
        return {"success": True, "mensaje": "Nomenclador creado exitosamente.", "nomenclador": creado}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error al crear nomenclador: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.put("/api/nomencladores/{nomenclador_id}")
def editar_nomenclador(nomenclador_id: str, payload: Dict[str, Any] = Body(...)):
    """
    Actualiza datos de un nomenclador.
    """
    try:
        act = update_nomenclador(nomenclador_id, payload)
        return {"success": True, "mensaje": "Nomenclador actualizado.", "nomenclador": act}
    except Exception as e:
        logger.error(f"Error al actualizar nomenclador {nomenclador_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/api/nomencladores/{nomenclador_id}")
def borrar_nomenclador(nomenclador_id: str):
    """
    Elimina un nomenclador y todas sus prácticas asociadas.
    """
    try:
        ok = delete_nomenclador(nomenclador_id)
        return {"success": ok, "mensaje": "Nomenclador eliminado exitosamente."}
    except Exception as e:
        logger.error(f"Error al eliminar nomenclador {nomenclador_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/nomenclador/practicas")
def get_practicas_nomenclador(
    nomenclador_id: Optional[str] = None,
    fecha: Optional[str] = None,
    q: Optional[str] = None,
    limit: int = 100,
    offset: int = 0
):
    """
    Lista las prácticas resolviendo el arancel y moneda vigentes para la fecha.
    """
    try:
        res = list_practicas_con_arancel(
            nomenclador_id=nomenclador_id,
            fecha_consulta=fecha,
            q=q,
            limit=limit,
            offset=offset
        )
        return {"success": True, "total": res["total"], "practicas": res["practicas"]}
    except Exception as e:
        logger.error(f"Error al listar prácticas: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/nomenclador/practicas")
def guardar_practica(payload: Dict[str, Any] = Body(...)):
    """
    Crea o actualiza una práctica individual del catálogo.
    """
    try:
        if not payload.get("nomenclador_id") or not payload.get("codigo") or not payload.get("nombre"):
            raise HTTPException(status_code=400, detail="Nomenclador, código y nombre son obligatorios.")
        guardada = create_or_update_practica(payload)
        return {"success": True, "mensaje": "Práctica guardada correctamente.", "practica": guardada}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error al guardar práctica: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/api/nomenclador/practicas/{practica_id}")
def borrar_practica(practica_id: str):
    """
    Elimina una práctica del catálogo.
    """
    try:
        ok = delete_practica(practica_id)
        return {"success": ok, "mensaje": "Práctica eliminada."}
    except Exception as e:
        logger.error(f"Error al eliminar práctica {practica_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/nomenclador/practicas/{practica_id}/arancel")
def agregar_arancel_vigencia(practica_id: str, payload: Dict[str, Any] = Body(...)):
    """
    Registra un nuevo arancel con fecha de vigencia para una práctica.
    """
    try:
        if "precio" not in payload:
            raise HTTPException(status_code=400, detail="El precio es obligatorio.")
        nuevo_arancel = upsert_arancel_practica(practica_id, payload)
        return {"success": True, "mensaje": "Arancel con vigencia guardado.", "arancel": nuevo_arancel}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error al registrar arancel: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# ====================================================================
# BÚSQUEDA RÁPIDA MULTI-MONEDA PARA PRESUPUESTOS (NATIVO CRM)
# ====================================================================

@app.get("/api/nomenclador/buscar-presupuesto")
def buscar_para_presupuesto(q: Optional[str] = "", fecha: Optional[str] = None):
    """
    Búsqueda optimizada por texto para el modal de presupuestos con arancel vigente.
    """
    try:
        resultados = buscar_practicas_presupuesto(q=q or "", fecha_consulta=fecha)
        return {
            "success": True,
            "query": q,
            "total": len(resultados),
            "resultados": resultados
        }
    except Exception as e:
        logger.error(f"Error al buscar prácticas para presupuesto: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# ====================================================================
# IMPORTADOR Y EXPORTADOR MASIVO EXCEL (.XLSX / .CSV)
# ====================================================================

@app.post("/api/nomenclador/importar-excel")
async def importar_practicas_excel(
    file: UploadFile = File(...),
    nomenclador_id: str = Form(...),
    modo: str = Form("upsert"),
    default_vigencia_desde: Optional[str] = Form(None),
    default_vigencia_hasta: Optional[str] = Form(None),
    default_moneda: str = Form("ARS")
):
    """
    Procesa la importación masiva de prácticas y aranceles desde un archivo Excel o CSV.
    """
    try:
        nom = get_nomenclador_by_id(nomenclador_id)
        if not nom:
            raise HTTPException(status_code=404, detail="El Nomenclador de destino no existe.")
            
        moneda_nom = default_moneda or nom.get("moneda_default", "ARS")
        contents = await file.read()
        filename = (file.filename or "").lower()
        
        parsed_rows = []
        
        if filename.endswith(".xlsx") or filename.endswith(".xls"):
            wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
            sheet = wb.active
            
            headers = []
            for cell in sheet[1]:
                val = str(cell.value or "").strip().lower()
                headers.append(val)
                
            def get_col_idx(names):
                for name in names:
                    for i, h in enumerate(headers):
                        if name in h:
                            return i
                return -1
                
            col_cod = get_col_idx(["cod", "codigo", "código", "id"])
            col_nom = get_col_idx(["nom", "nombre", "practica", "práctica", "descripcion", "descripción"])
            col_cat = get_col_idx(["cat", "categoria", "categoría", "tipo", "rubro"])
            col_pre = get_col_idx(["pre", "precio", "arancel", "valor", "importe", "monto"])
            col_mon = get_col_idx(["mon", "moneda", "currency"])
            col_vdes = get_col_idx(["desde", "vigencia_desde", "inicio"])
            col_vhas = get_col_idx(["hasta", "vigencia_hasta", "fin"])
            
            if col_cod == -1 or col_nom == -1:
                raise HTTPException(status_code=400, detail="El archivo Excel debe contener al menos las columnas 'Código' y 'Nombre/Práctica'.")
                
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row or not any(row):
                    continue
                codigo = str(row[col_cod] or "").strip()
                nombre = str(row[col_nom] or "").strip()
                if not codigo or not nombre:
                    continue
                    
                categoria = str(row[col_cat] or "General").strip() if col_cat != -1 and row[col_cat] else "General"
                precio = row[col_pre] if col_pre != -1 and row[col_pre] is not None else 0.0
                moneda = str(row[col_mon] or moneda_nom).strip().upper() if col_mon != -1 and row[col_mon] else moneda_nom
                
                v_desde = None
                if col_vdes != -1 and row[col_vdes]:
                    v_raw = row[col_vdes]
                    v_desde = v_raw.strftime("%Y-%m-%d") if hasattr(v_raw, "strftime") else str(v_raw).strip()
                    
                v_hasta = None
                if col_vhas != -1 and row[col_vhas]:
                    v_raw = row[col_vhas]
                    v_hasta = v_raw.strftime("%Y-%m-%d") if hasattr(v_raw, "strftime") else str(v_raw).strip()
                    
                parsed_rows.append({
                    "codigo": codigo,
                    "nombre": nombre,
                    "categoria": categoria,
                    "precio": precio,
                    "moneda": moneda,
                    "vigencia_desde": v_desde,
                    "vigencia_hasta": v_hasta
                })
        elif filename.endswith(".csv"):
            text_content = contents.decode("utf-8-sig", errors="ignore")
            delimiter = ";" if ";" in text_content[:200] else ","
            reader = csv.DictReader(text_content.splitlines(), delimiter=delimiter)
            
            for row in reader:
                norm_row = {k.strip().lower(): v for k, v in row.items() if k}
                codigo = norm_row.get("codigo") or norm_row.get("código") or norm_row.get("cod") or ""
                nombre = norm_row.get("nombre") or norm_row.get("practica") or norm_row.get("práctica") or norm_row.get("descripcion") or ""
                if not codigo or not nombre:
                    continue
                categoria = norm_row.get("categoria") or norm_row.get("categoría") or "General"
                precio = norm_row.get("precio") or norm_row.get("arancel") or norm_row.get("valor") or 0.0
                moneda = norm_row.get("moneda") or moneda_nom
                v_desde = norm_row.get("vigencia_desde") or norm_row.get("desde")
                v_hasta = norm_row.get("vigencia_hasta") or norm_row.get("hasta")
                
                parsed_rows.append({
                    "codigo": str(codigo).strip(),
                    "nombre": str(nombre).strip(),
                    "categoria": str(categoria).strip(),
                    "precio": precio,
                    "moneda": str(moneda).strip().upper(),
                    "vigencia_desde": v_desde,
                    "vigencia_hasta": v_hasta
                })
        else:
            raise HTTPException(status_code=400, detail="Formato no soportado. Por favor sube un archivo .xlsx o .csv")
            
        if not parsed_rows:
            raise HTTPException(status_code=400, detail="No se encontraron filas con datos válidos en el archivo.")
            
        res_import = bulk_import_practicas_aranceles(
            nomenclador_id=nomenclador_id,
            rows=parsed_rows,
            modo=modo,
            default_vigencia_desde=default_vigencia_desde,
            default_vigencia_hasta=default_vigencia_hasta,
            default_moneda=moneda_nom
        )
        
        return {
            "success": True,
            "mensaje": f"Se procesaron {res_import['total_procesadas']} filas exitosamente ({res_import['insertadas']} aranceles actualizados).",
            "detalle": res_import
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error al importar Excel: {e}")
        raise HTTPException(status_code=500, detail=f"Error durante la importación: {str(e)}")

@app.get("/api/nomenclador/descargar-plantilla")
def descargar_plantilla_excel():
    """
    Genera y descarga un archivo Excel .xlsx oficial con encabezados, instrucciones y filas de muestra en ARS y USD.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Nomenclador_Plantilla"
    
    headers = ["Codigo", "Nombre", "Categoria", "Precio", "Moneda", "Vigencia_Desde", "Vigencia_Hasta", "Descripcion"]
    ws.append(headers)
    
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    ejemplos = [
        ["420101", "Consulta Médica en Consultorio", "Consultas", 8500.00, "ARS", "2026-09-01", "2026-12-31", "Consulta ambulatoria general"],
        ["180104", "Ecografía Ginecológica / Tocoginecológica", "Diagnóstico", 22000.00, "ARS", "2026-09-01", "", "Ecografía pelviana transvaginal"],
        ["FIV-ICSI-01", "Tratamiento FIV + ICSI Alta Complejidad", "Fertilidad", 1500.00, "USD", "2026-09-01", "", "Incluye estimulación y laboratorio"],
        ["KIT-MED-02", "Kit de Medicación y Criopreservación", "Laboratorio", 450.00, "USD", "2026-09-01", "2027-03-01", "Mantenimiento anual"]
    ]
    
    for row in ejemplos:
        ws.append(row)
        
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 14)
        
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    headers_resp = {
        "Content-Disposition": "attachment; filename=plantilla_nomenclador_crm.xlsx"
    }
    return Response(content=output.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers_resp)

@app.get("/api/nomenclador/exportar-excel")
def exportar_nomenclador_excel(nomenclador_id: Optional[str] = None):
    """
    Exporta el catálogo completo o de un nomenclador específico a un archivo Excel .xlsx descargable.
    """
    try:
        res = list_practicas_con_arancel(nomenclador_id=nomenclador_id, limit=5000)
        practicas = res.get("practicas", [])
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Catalogo_Practicas"
        
        headers = ["Nomenclador", "Codigo", "Nombre", "Categoria", "Precio", "Moneda", "Vigencia_Desde", "Vigencia_Hasta", "Estado"]
        ws.append(headers)
        
        header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
        for p in practicas:
            estado = "Activa" if p.get("activo") else "Inactiva"
            ws.append([
                p.get("nomenclador_nombre", ""),
                p.get("codigo", ""),
                p.get("nombre", ""),
                p.get("categoria", "General"),
                p.get("precio", 0.0),
                p.get("moneda", "ARS"),
                p.get("vigencia_desde", "") or "",
                p.get("vigencia_hasta", "") or "",
                estado
            ])
            
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 14)
            
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        headers_resp = {
            "Content-Disposition": "attachment; filename=catalogo_nomencladores_crm.xlsx"
        }
        return Response(content=output.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers_resp)
    except Exception as e:
        logger.error(f"Error al exportar Excel: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ====================================================================
# ENDPOINTS REST: SISTEMA DE LOGS Y AUDITORÍA EN TIEMPO REAL
# ====================================================================

@app.get("/api/logs")
def consultar_logs_sistema(
    limit: int = 50,
    offset: int = 0,
    nivel: Optional[str] = None,
    modulo: Optional[str] = None,
    search: Optional[str] = None,
    paciente_id: Optional[str] = None,
    desde: Optional[str] = None,
    hasta: Optional[str] = None
):
    """
    Retorna la lista paginada de logs y auditoría del sistema con filtros avanzados.
    """
    try:
        resultado = get_logs(
            limit=limit,
            offset=offset,
            nivel=nivel,
            modulo=modulo,
            search=search,
            paciente_id=paciente_id,
            desde=desde,
            hasta=hasta
        )
        return {
            "success": True,
            "logs": resultado.get("logs", []),
            "total": resultado.get("total", 0)
        }
    except Exception as e:
        logger.error(f"Error al consultar logs: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/logs/stats")
def obtener_estadisticas_logs():
    """
    Retorna métricas consolidadas de salud y volumen de eventos de las últimas 24 horas.
    """
    try:
        stats = get_logs_stats()
        return {"success": True, "stats": stats}
    except Exception as e:
        logger.error(f"Error al calcular estadísticas de logs: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/logs/client")
def registrar_log_desde_cliente(payload: Dict[str, Any] = Body(...)):
    """
    Permite al frontend registrar eventos o errores ocurridos en el navegador del usuario.
    """
    try:
        nivel = payload.get("nivel", "ERROR")
        modulo = payload.get("modulo", "FRONTEND")
        accion = payload.get("accion", "ERROR_CLIENTE")
        mensaje = payload.get("mensaje", "Error reportado desde la interfaz web")
        detalles = payload.get("detalles", {})
        duracion_ms = payload.get("duracion_ms")
        http_status = payload.get("http_status")
        trace = payload.get("trace")

        log_entry = log_event(
            nivel=nivel,
            modulo=modulo,
            accion=accion,
            mensaje=mensaje,
            detalles=detalles,
            duracion_ms=duracion_ms,
            http_status=http_status,
            trace=trace
        )
        return {"success": True, "log": log_entry}
    except Exception as e:
        logger.error(f"Error registrando log de cliente: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# ====================================================================
# ENDPOINTS: MÓDULO DE ASESORÍAS QUIRÚRGICAS (PIPELINE)
# ====================================================================

@app.get("/api/asesorias-quirurgicas/paciente/{paciente_id}")
def obtener_asesorias_paciente(paciente_id: str):
    """
    Retorna el listado de asesorías quirúrgicas de un paciente.
    """
    try:
        asesorias = get_asesorias_by_paciente(paciente_id)
        return {"success": True, "asesorias": asesorias}
    except Exception as e:
        logger.error(f"Error al obtener asesorías del paciente {paciente_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/asesorias-quirurgicas")
def crear_nueva_asesoria(payload: Dict[str, Any] = Body(...)):
    """
    Crea un nuevo caso de asesoramiento quirúrgico para un paciente.
    """
    if not payload.get("paciente_id"):
        raise HTTPException(status_code=400, detail="El ID del paciente es obligatorio.")
    try:
        asesoria = crear_asesoria_quirurgica(payload)
        return {"success": True, "mensaje": "Caso de asesoría quirúrgica registrado.", "asesoria": asesoria}
    except Exception as e:
        logger.error(f"Error al crear asesoría quirúrgica: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.put("/api/asesorias-quirurgicas/{asesoria_id}")
def actualizar_asesoria(asesoria_id: str, payload: Dict[str, Any] = Body(...)):
    """
    Actualiza estado, fechas, condiciones o notas de un caso quirúrgico.
    """
    try:
        asesoria = actualizar_asesoria_quirurgica(asesoria_id, payload)
        return {"success": True, "mensaje": "Caso quirúrgico actualizado.", "asesoria": asesoria}
    except Exception as e:
        logger.error(f"Error al actualizar asesoría {asesoria_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/api/asesorias-quirurgicas/{asesoria_id}")
def eliminar_asesoria(asesoria_id: str):
    """
    Elimina un caso de asesoría quirúrgica.
    """
    try:
        ok = eliminar_asesoria_quirurgica(asesoria_id)
        return {"success": ok, "mensaje": "Caso quirúrgico eliminado."}
    except Exception as e:
        logger.error(f"Error al eliminar asesoría {asesoria_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# ====================================================================
# ENDPOINTS: PRESUPUESTOS INTEGRADOS CON EXPEDIENTE DE PACIENTE
# ====================================================================

@app.get("/api/presupuestos/paciente/{paciente_id}")
@app.get("/api/pacientes/{paciente_id}/presupuestos")
def obtener_presupuestos_paciente(paciente_id: str):
    """
    Retorna el listado histórico de presupuestos emitidos a un paciente con sus ítems detallados.
    """
    try:
        presupuestos = get_presupuestos_by_paciente(paciente_id)
        return {"success": True, "presupuestos": presupuestos}
    except Exception as e:
        logger.error(f"Error al obtener presupuestos del paciente {paciente_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.put("/api/presupuestos/{presupuesto_id}/estado")
def actualizar_estado_presupuesto(presupuesto_id: str, payload: Dict[str, Any] = Body(...)):
    """
    Actualiza el estado de un presupuesto ('borrador', 'enviado', 'aprobado', 'rechazado')
    y sincroniza automáticamente el caso quirúrgico vinculado.
    """
    nuevo_estado = payload.get("estado")
    asesoria_id = payload.get("asesoria_id")
    if not nuevo_estado:
        raise HTTPException(status_code=400, detail="El nuevo estado es obligatorio.")
    try:
        presupuesto = cambiar_estado_presupuesto(presupuesto_id, nuevo_estado, asesoria_id)
        return {
            "success": True, 
            "mensaje": f"Presupuesto marcado como {nuevo_estado}.",
            "presupuesto": presupuesto
        }
    except Exception as e:
        logger.error(f"Error al actualizar estado del presupuesto {presupuesto_id}: {e}")
@app.put("/api/presupuestos/{presupuesto_id}/vincular-asesoria")
def vincular_presupuesto_asesoria_api(presupuesto_id: str, payload: Dict[str, Any] = Body(...)):
    """
    Asocia bidireccionalmente un presupuesto existente con un caso quirúrgico.
    """
    asesoria_id = payload.get("asesoria_id")
    if not asesoria_id:
        raise HTTPException(status_code=400, detail="El ID de la asesoría quirúrgica es obligatorio.")
    try:
        presupuesto = vincular_presupuesto_a_asesoria(presupuesto_id, asesoria_id)
        return {"success": True, "mensaje": "Presupuesto vinculado exitosamente al caso quirúrgico.", "presupuesto": presupuesto}
    except Exception as e:
        logger.error(f"Error al vincular presupuesto {presupuesto_id} a asesoría {asesoria_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/api/presupuestos/{presupuesto_id}")
def eliminar_presupuesto_endpoint(presupuesto_id: str):
    """
    Elimina un presupuesto y sus ítems, desvinculándolo si estaba asociado a una asesoría.
    """
    try:
        ok = eliminar_presupuesto(presupuesto_id)
        return {"success": ok, "mensaje": "Presupuesto eliminado correctamente."}
    except Exception as e:
        logger.error(f"Error al eliminar presupuesto {presupuesto_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/presupuestos/crear-rapido")
def emitir_presupuesto_rapido(payload: Dict[str, Any] = Body(...)):
    """
    Crea un presupuesto con ítems, calcula el total, genera el PDF membretado oficial del CRM
    y vincula el resultado al paciente y su caso quirúrgico.
    """
    if not payload.get("paciente_id"):
        raise HTTPException(status_code=400, detail="El ID del paciente es obligatorio.")
    if not payload.get("items") or len(payload["items"]) == 0:
        raise HTTPException(status_code=400, detail="Debe incluir al menos un ítem o prestación.")
    try:
        presupuesto = crear_presupuesto_rapido(payload)
        return {
            "success": True,
            "mensaje": "Presupuesto médico emitido y PDF generado correctamente.",
            "presupuesto": presupuesto
        }
    except Exception as e:
        logger.error(f"Error al emitir presupuesto rápido: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/presupuestos")
def listar_todos_presupuestos_api():
    """
    Retorna la lista completa de presupuestos con totales discriminados en ARS y USD,
    datos del paciente e ítems asociados.
    """
    try:
        if not supabase:
            return {"success": False, "presupuestos": []}
        resp = supabase.table("presupuestos")\
            .select("id, paciente_id, asesoria_id, estado, total, total_ars, total_usd, pdf_url, created_at, pacientes(id, nombre, telefono, dni), items_presupuesto(id, cantidad, precio_unitario, subtotal, servicios_precios(id, codigo, nombre_prestacion))")\
            .order("created_at", desc=True)\
            .execute()
        return {"success": True, "presupuestos": resp.data or []}
    except Exception as e:
        logger.error(f"Error al listar presupuestos: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/presupuestos/{presupuesto_id}/mensaje-sugerido")
def obtener_mensaje_sugerido_presupuesto_api(presupuesto_id: str):
    """
    Genera el borrador de texto ameno de WhatsApp para un presupuesto específico.
    """
    try:
        p_resp = supabase.table("presupuestos")\
            .select("*, pacientes(*), items_presupuesto(*, servicios_precios(*))")\
            .eq("id", presupuesto_id)\
            .execute()
            
        if not p_resp.data:
            raise HTTPException(status_code=404, detail="Presupuesto no encontrado.")
            
        presupuesto = p_resp.data[0]
        paciente = presupuesto.get("pacientes") or {}
        items_raw = presupuesto.get("items_presupuesto") or []
        
        items = []
        for it in items_raw:
            srv = it.get("servicios_precios") or {}
            items.append({
                "codigo": srv.get("codigo") or "",
                "nombre": srv.get("nombre_prestacion") or "Prestación Médica",
                "precio_unitario": float(it.get("precio_unitario") or 0.0),
                "cantidad": int(it.get("cantidad") or 1),
                "subtotal": float(it.get("subtotal") or 0.0),
                "moneda": "USD" if float(presupuesto.get("total_usd") or 0) > 0 and float(presupuesto.get("total_ars") or 0) == 0 else "ARS"
            })
            
        mensaje = generar_mensaje_ameno_presupuesto(presupuesto, paciente, items)
        
        return {
            "success": True,
            "presupuesto_id": presupuesto_id,
            "paciente_nombre": paciente.get("nombre"),
            "telefono": paciente.get("telefono"),
            "total_ars": float(presupuesto.get("total_ars") or 0.0),
            "total_usd": float(presupuesto.get("total_usd") or 0.0),
            "pdf_url": presupuesto.get("pdf_url"),
            "mensaje_sugerido": mensaje
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error al generar mensaje sugerido para presupuesto {presupuesto_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/presupuestos/{presupuesto_id}/mensaje-seguimiento")
def obtener_mensaje_seguimiento_presupuesto_api(presupuesto_id: str, tipo: str = "seguimiento"):
    """
    Genera un mensaje de re-contacto comercial cordial para seguimiento o aviso por vencimiento.
    """
    try:
        p_resp = supabase.table("presupuestos")\
            .select("*, pacientes(*), items_presupuesto(*, servicios_precios(*))")\
            .eq("id", presupuesto_id)\
            .execute()
            
        if not p_resp.data:
            raise HTTPException(status_code=404, detail="Presupuesto no encontrado.")
            
        presupuesto = p_resp.data[0]
        paciente = presupuesto.get("pacientes") or {}
        items_raw = presupuesto.get("items_presupuesto") or []
        
        items = []
        for it in items_raw:
            srv = it.get("servicios_precios") or {}
            items.append({
                "codigo": srv.get("codigo") or "",
                "nombre": srv.get("nombre_prestacion") or "Prestación Médica",
                "precio_unitario": float(it.get("precio_unitario") or 0.0),
                "cantidad": int(it.get("cantidad") or 1),
                "subtotal": float(it.get("subtotal") or 0.0),
                "moneda": str(it.get("moneda") or srv.get("moneda") or "ARS").upper()
            })
            
        mensaje = generar_mensaje_seguimiento_presupuesto(presupuesto, paciente, items, tipo=tipo)
        
        return {
            "success": True,
            "presupuesto_id": presupuesto_id,
            "paciente_nombre": paciente.get("nombre"),
            "telefono": paciente.get("telefono"),
            "tipo": tipo,
            "mensaje_seguimiento": mensaje
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error al generar mensaje de seguimiento para presupuesto {presupuesto_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/presupuestos/{presupuesto_id}/duplicar")
def duplicar_presupuesto_api(presupuesto_id: str):
    """
    Retorna los datos preformateados de un presupuesto existente para clonarlo en el cotizador.
    """
    try:
        data = obtener_datos_duplicar_presupuesto(presupuesto_id)
        return {"success": True, **data}
    except Exception as e:
        logger.error(f"Error al duplicar presupuesto {presupuesto_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/presupuestos/{presupuesto_id}/enviar-whatsapp")
def enviar_presupuesto_whatsapp_api(presupuesto_id: str, payload: Dict[str, Any] = Body(...)):
    """
    Envía el PDF del presupuesto por WhatsApp junto con el mensaje personalizado o ameno al paciente.
    """
    try:
        telefono_override = payload.get("telefono")
        mensaje_custom = payload.get("mensaje")
        
        res = enviar_presupuesto_por_whatsapp(
            presupuesto_id=presupuesto_id,
            telefono_override=telefono_override,
            mensaje_custom=mensaje_custom
        )
        return res
    except Exception as e:
        logger.error(f"Error al enviar presupuesto {presupuesto_id} por WhatsApp: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/presupuestos/{presupuesto_id}/pdf")
def obtener_pdf_presupuesto(presupuesto_id: str):
    """
    Sirve el archivo PDF membretado oficial de un presupuesto directamente como stream / descarga.
    """
    return servir_archivo_estatico(f"presupuesto_{presupuesto_id}.pdf")

# ====================================================================
# ENDPOINTS: BITÁCORA Y EVOLUCIONES DE ASESORAMIENTO QUIRÚRGICO
# ====================================================================

@app.get("/api/asesorias-quirurgicas/{asesoria_id}/evoluciones")
def listar_evoluciones_asesoria(asesoria_id: str):
    """
    Retorna el historial cronológico de notas de evolución de un caso quirúrgico.
    """
    try:
        evoluciones = get_evoluciones_by_asesoria(asesoria_id)
        return {"success": True, "evoluciones": evoluciones}
    except Exception as e:
        logger.error(f"Error al listar evoluciones de asesoría {asesoria_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/asesorias-quirurgicas/{asesoria_id}/evoluciones")
def registrar_evolucion_asesoria(asesoria_id: str, payload: Dict[str, Any] = Body(...)):
    """
    Registra una nueva evolución en la bitácora del asesoramiento quirúrgico.
    """
    if not payload.get("paciente_id"):
        raise HTTPException(status_code=400, detail="El ID del paciente es obligatorio.")
    if not payload.get("contenido") or not str(payload["contenido"]).strip():
        raise HTTPException(status_code=400, detail="El contenido de la evolución no puede estar vacío.")
    try:
        payload["asesoria_id"] = asesoria_id
        evolucion = crear_evolucion_asesoria(payload)
        return {
            "success": True,
            "mensaje": "Nota de evolución registrada correctamente.",
            "evolucion": evolucion
        }
    except Exception as e:
        logger.error(f"Error al registrar evolución de asesoría {asesoria_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/api/asesorias-quirurgicas/evoluciones/{evolucion_id}")
def borrar_evolucion_asesoria(evolucion_id: str):
    """
    Elimina una nota de evolución por su ID.
    """
    try:
        ok = eliminar_evolucion_asesoria(evolucion_id)
        return {"success": ok, "mensaje": "Evolución eliminada."}
    except Exception as e:
        logger.error(f"Error al eliminar evolución {evolucion_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# ====================================================================
# ENDPOINTS: CONFIGURACIÓN QUIRÚRGICA & PIPELINE LEAD-TO-SURGERY
# ====================================================================

@app.get("/api/configuracion-quirurgica")
def obtener_config_quirurgica():
    """
    Retorna la configuración de SLA, plantillas WhatsApp y checklist quirúrgico.
    """
    try:
        config = get_configuracion_quirurgica()
        return {"success": True, "configuracion": config}
    except Exception as e:
        logger.error(f"Error al obtener configuración quirúrgica: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.put("/api/configuracion-quirurgica")
def guardar_config_quirurgica(payload: Dict[str, Any] = Body(...)):
    """
    Actualiza la configuración de SLA, plantillas de WhatsApp o checklist.
    """
    try:
        config = actualizar_configuracion_quirurgica(payload)
        return {"success": True, "mensaje": "Configuración guardada correctamente.", "configuracion": config}
    except Exception as e:
        logger.error(f"Error al guardar configuración quirúrgica: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/pipeline-quirurgico")
def obtener_pipeline():
    """
    Retorna el tablero global Kanban de cirugías por etapa con KPIs de ingresos.
    """
    try:
        data = get_pipeline_quirurgico()
        return {"success": True, **data}
    except Exception as e:
        logger.error(f"Error al obtener pipeline quirúrgico: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# ====================================================================
# ENDPOINTS: CONFIGURACIÓN DE QUIRÓFANOS Y CONSENTIMIENTOS
# ====================================================================

@app.get("/api/configuracion-quirofano")
def obtener_config_quirofano():
    try:
        config = get_configuracion_quirofano()
        return {"success": True, "configuracion": config}
    except Exception as e:
        logger.error(f"Error al obtener configuración de quirófano: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.put("/api/configuracion-quirofano")
def guardar_config_quirofano(payload: Dict[str, Any] = Body(...)):
    try:
        config = actualizar_configuracion_quirofano(payload)
        return {"success": True, "mensaje": "Configuración guardada correctamente.", "configuracion": config}
    except Exception as e:
        logger.error(f"Error al actualizar configuración de quirófano: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# Quirófanos / Salas
@app.get("/api/quirofanos")
def listar_quirofanos(solo_activos: bool = False):
    try:
        salas = get_quirofanos(solo_activos=solo_activos)
        return {"success": True, "quirofanos": salas}
    except Exception as e:
        logger.error(f"Error al listar salas de quirófano: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/quirofanos")
def crear_sala_quirofano(payload: Dict[str, Any] = Body(...)):
    try:
        sala = crear_quirofano(payload)
        return {"success": True, "quirofano": sala}
    except Exception as e:
        logger.error(f"Error al crear sala de quirófano: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.put("/api/quirofanos/{quirofano_id}")
def actualizar_sala_quirofano(quirofano_id: str, payload: Dict[str, Any] = Body(...)):
    try:
        sala = actualizar_quirofano(quirofano_id, payload)
        return {"success": True, "quirofano": sala}
    except Exception as e:
        logger.error(f"Error al actualizar sala de quirófano {quirofano_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/api/quirofanos/{quirofano_id}")
def eliminar_sala_quirofano(quirofano_id: str):
    try:
        ok = eliminar_quirofano(quirofano_id)
        return {"success": ok}
    except Exception as e:
        logger.error(f"Error al eliminar sala {quirofano_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# Bloques Médicos
@app.get("/api/quirofanos/bloques-medicos")
def listar_bloques_medicos(quirofano_id: Optional[str] = None):
    try:
        bloques = get_quirofano_bloques(quirofano_id=quirofano_id)
        return {"success": True, "bloques": bloques}
    except Exception as e:
        logger.error(f"Error al listar bloques de quirófano: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/quirofanos/bloques-medicos")
def crear_bloque_medico(payload: Dict[str, Any] = Body(...)):
    try:
        bloque = crear_quirofano_bloque(payload)
        return {"success": True, "bloque": bloque}
    except Exception as e:
        logger.error(f"Error al crear bloque médico: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/api/quirofanos/bloques-medicos/{bloque_id}")
def eliminar_bloque_medico(bloque_id: str):
    try:
        ok = eliminar_quirofano_bloque(bloque_id)
        return {"success": ok}
    except Exception as e:
        logger.error(f"Error al eliminar bloque médico {bloque_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# Bloqueos de Horario ("No dar turno")
@app.get("/api/quirofano-bloqueos")
def listar_bloqueos(fecha_desde: Optional[str] = None, fecha_hasta: Optional[str] = None):
    try:
        bloqueos = get_quirofano_bloqueos(fecha_desde=fecha_desde, fecha_hasta=fecha_hasta)
        return {"success": True, "bloqueos": bloqueos}
    except Exception as e:
        logger.error(f"Error al listar bloqueos: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/quirofano-bloqueos")
def crear_bloqueo(payload: Dict[str, Any] = Body(...)):
    try:
        bloqueo = crear_quirofano_bloqueo(payload)
        return {"success": True, "bloqueo": bloqueo}
    except Exception as e:
        logger.error(f"Error al crear bloqueo: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/api/quirofano-bloqueos/{bloqueo_id}")
def eliminar_bloqueo(bloqueo_id: str):
    try:
        ok = eliminar_quirofano_bloqueo(bloqueo_id)
        return {"success": ok}
    except Exception as e:
        logger.error(f"Error al eliminar bloqueo {bloqueo_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# Turnos Quirúrgicos
# Casos de Asesoramiento Quirúrgico Confirmados Pendientes de Quirófano
@app.get("/api/asesorias-quirurgicas/pendientes-quirofano")
def listar_asesorias_confirmadas():
    try:
        casos = get_asesorias_confirmadas_pendientes()
        return {"success": True, "casos": casos}
    except Exception as e:
        logger.error(f"Error al listar asesorías confirmadas para quirófano: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/turnos-quirofano")
def listar_turnos_quirofano(
    fecha_desde: Optional[str] = None,
    fecha_hasta: Optional[str] = None,
    quirofano_id: Optional[str] = None,
    cirujano_id: Optional[int] = None,
    estado: Optional[str] = None
):
    try:
        turnos = get_turnos_quirofano(
            fecha_desde=fecha_desde,
            fecha_hasta=fecha_hasta,
            quirofano_id=quirofano_id,
            cirujano_id=cirujano_id,
            estado=estado
        )
        return {"success": True, "turnos": turnos}
    except Exception as e:
        logger.error(f"Error al listar turnos de quirófano: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/turnos-quirofano")
def crear_turno(payload: Dict[str, Any] = Body(...)):
    try:
        turno = crear_turno_quirofano(payload)
        return {"success": True, "turno": turno}
    except Exception as e:
        logger.error(f"Error al crear turno de quirófano: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.put("/api/turnos-quirofano/{turno_id}")
def actualizar_turno(turno_id: str, payload: Dict[str, Any] = Body(...)):
    try:
        turno = actualizar_turno_quirofano(turno_id, payload)
        return {"success": True, "turno": turno}
    except Exception as e:
        logger.error(f"Error al actualizar turno de quirófano {turno_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/api/turnos-quirofano/{turno_id}")
def eliminar_turno(turno_id: str):
    try:
        ok = eliminar_turno_quirofano(turno_id)
        return {"success": ok}
    except Exception as e:
        logger.error(f"Error al eliminar turno {turno_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# ====================================================================
# TRAZABILIDAD QUIRÚRGICA POR CÓDIGO QR & PULSERAS TÉRMICAS
# ====================================================================

class EscanearQRRequest(BaseModel):
    codigo_qr: str
    estacion: Optional[str] = "General"
    usuario_crm: Optional[str] = None
    accion_deseada: Optional[str] = None

@app.post("/api/turnos-quirofano/escanear-qr")
def escanear_qr_turno_api(payload: EscanearQRRequest):
    """
    Procesa la lectura de un código QR de pulsera médica para identificación
    inequívoca del paciente y transición en tiempo real de estadios quirúrgicos.
    """
    res = procesar_escaneo_qr_turno(
        codigo_qr=payload.codigo_qr,
        estacion=payload.estacion,
        usuario_crm=payload.usuario_crm,
        accion_deseada=payload.accion_deseada
    )
    if not res.get("success"):
        raise HTTPException(status_code=400, detail=res.get("error") or "Error al procesar el código QR.")
    return res

@app.get("/api/turnos-quirofano/{turno_id}/datos-pulsera")
def obtener_datos_pulsera_api(turno_id: str):
    """
    Retorna los datos clínicos consolidados para la impresión térmica de la pulsera
    identificatoria en la impresora TSC TDP-225.
    """
    res = obtener_datos_pulsera_turno(turno_id)
    if not res.get("success"):
        raise HTTPException(status_code=404, detail=res.get("error") or "Turno no encontrado.")
    return res

@app.post("/api/turnos-quirofano/{turno_id}/marcar-pulsera-impresa")
def marcar_pulsera_impresa_api(turno_id: str, payload: Optional[Dict[str, Any]] = None):
    """
    Marca que la pulsera identificatoria ha sido impresa y colocada al paciente.
    """
    user = (payload or {}).get("usuario_crm") if isinstance(payload, dict) else None
    res = marcar_pulsera_impresa(turno_id, usuario_crm=user)
    if not res.get("success"):
        raise HTTPException(status_code=400, detail=res.get("error") or "Error al registrar impresión de pulsera.")
    return res

# Disparo de Consentimiento por WhatsApp
@app.post("/api/turnos-quirofano/{turno_id}/enviar-consentimiento-wa")
async def enviar_consentimiento_whatsapp(turno_id: str):
    try:
        turno = get_turno_quirofano_by_id(turno_id)
        if not turno:
            raise HTTPException(status_code=404, detail="Turno no encontrado")
            
        paciente = turno.get("pacientes") or {}
        telefono = paciente.get("telefono")
        if not telefono:
            raise HTTPException(status_code=400, detail="El paciente no tiene teléfono registrado")
            
        config = get_configuracion_quirofano()
        plantilla_msg = config.get("whatsapp_mensaje_envio") or (
            "Hola {paciente}, confirmamos tu turno de cirugía de {cirugia} ({ojo_intervenido}) para el día {fecha_cirugia} a las {hora_cirugia} hs con el Dr. {cirujano}. "
            "Por favor, revisá y firmá tu Consentimiento Informado en tu celular desde el siguiente enlace seguro: {enlace_firma}"
        )
        
        token = turno.get("consentimiento_token")
        if not token:
            import secrets
            token = secrets.token_urlsafe(24)
            actualizar_turno_quirofano(turno_id, {"consentimiento_token": token})
            
        # Determinar base URL pública o del frontend (Vercel en producción)
        base_app_url = os.getenv("NEXT_PUBLIC_APP_URL") or os.getenv("APP_URL") or os.getenv("FRONTEND_URL") or "https://crm-agentico-nube-tn4d.vercel.app"
        enlace_firma = f"{base_app_url}/consentimiento/{token}"
        
        ojo = turno.get("ojo") or "OD"
        ojo_desc = "Ojo Derecho" if ojo == "OD" else "Ojo Izquierdo" if ojo == "OI" else "Ambos Ojos"
        
        mensaje_final = plantilla_msg.format(
            paciente=paciente.get("nombre") or "Paciente",
            cirugia=turno.get("practica_nombre") or "Cirugía",
            ojo_intervenido=ojo_desc,
            fecha_cirugia=str(turno.get("fecha_cirugia") or ""),
            hora_cirugia=str(turno.get("hora_inicio") or "")[:5],
            cirujano=turno.get("cirujano_nombre") or "Médico Cirujano",
            enlace_firma=enlace_firma
        )
        
        # Enviar mensaje vía WhatsApp (función síncrona)
        jid = telefono if "@" in telefono else f"{telefono}@s.whatsapp.net"
        res_wa = whatsapp_manager.enviar_mensaje(jid, mensaje_final)
        
        # Actualizar estado de envío en BD
        actualizar_turno_quirofano(turno_id, {
            "consentimiento_estado": "enviado_whatsapp",
            "consentimiento_enviado_at": "now()"
        })
        
        return {
            "success": True, 
            "mensaje": "Consentimiento enviado por WhatsApp exitosamente.",
            "enlace_firma": enlace_firma,
            "resultado_wa": res_wa
        }
    except Exception as e:
        logger.error(f"Error al enviar consentimiento por WhatsApp: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# Endpoint Público para la pantalla de firma del paciente
@app.get("/api/consentimiento-publico/{token}")
def obtener_datos_consentimiento_publico(token: str):
    try:
        turno = get_consentimiento_by_token(token)
        if not turno:
            raise HTTPException(status_code=404, detail="El enlace de consentimiento no es válido o ha caducado.")
            
        paciente = turno.get("pacientes") or {}
        # 1. Intentar resolver el texto legal desde el Nomenclador Multidimensional
        practica_cod = turno.get("practica_codigo") or ""
        practica_id = turno.get("practica_id") or ""
        practica_nombre = turno.get("practica_nombre") or ""
        
        # Si el turno no tiene práctica definida o es genérica, buscar en la asesoría quirúrgica vinculada
        if (not practica_cod or not practica_nombre or practica_nombre == "Nueva Cirugía / Procedimiento") and turno.get("asesoria_id"):
            try:
                res_as = supabase.table("asesorias_quirurgicas").select("practica_codigo, practica_nombre, presupuesto_id").eq("id", turno["asesoria_id"]).limit(1).execute()
                if res_as.data:
                    as_row = res_as.data[0]
                    if as_row.get("practica_codigo"):
                        practica_cod = as_row["practica_codigo"]
                    if as_row.get("practica_nombre") and as_row["practica_nombre"] != "Nueva Cirugía / Procedimiento":
                        practica_nombre = as_row["practica_nombre"]
                    elif as_row.get("presupuesto_id"):
                        res_p = supabase.table("presupuestos").select("items_presupuesto(servicios_precios(codigo, nombre_prestacion))").eq("id", as_row["presupuesto_id"]).limit(1).execute()
                        if res_p.data and res_p.data[0].get("items_presupuesto"):
                            item_sp = res_p.data[0]["items_presupuesto"][0].get("servicios_precios") or {}
                            if item_sp.get("codigo"):
                                practica_cod = item_sp["codigo"]
                            if item_sp.get("nombre_prestacion"):
                                practica_nombre = item_sp["nombre_prestacion"]
                    
                    # Actualizar turno con los datos reales encontrados
                    if practica_nombre and practica_nombre != "Nueva Cirugía / Procedimiento":
                        supabase.table("turnos_quirofano").update({
                            "practica_nombre": practica_nombre,
                            "practica_codigo": practica_cod or None
                        }).eq("id", turno["id"]).execute()
                        turno["practica_nombre"] = practica_nombre
                        turno["practica_codigo"] = practica_cod
            except Exception as e_as:
                logger.warning(f"Aviso al resolver práctica desde asesoría vinculada: {e_as}")

        cuerpo_template = None
        titulo_consentimiento = "Consentimiento Informado Quirúrgico"
        
        resumen_practica = get_practica_resumen_operativo(practica_id or practica_cod or practica_nombre)
        if resumen_practica and resumen_practica.get("habilitar_consentimiento") and resumen_practica.get("texto_consentimiento"):
            cuerpo_template = resumen_practica["texto_consentimiento"]
            titulo_consentimiento = resumen_practica.get("titulo_consentimiento") or "Consentimiento Informado Quirúrgico"
            
        # 2. Fallback a configuración de quirófano general si la práctica no tiene plantilla asignada
        if not cuerpo_template:
            config = get_configuracion_quirofano()
            plantillas = config.get("plantillas_consentimiento") or []
            plantilla_sel = None
            p_nom_low = (practica_nombre or "").lower()
            for pl in plantillas:
                pl_id = str(pl.get("id") or "").lower()
                pl_tipo = str(pl.get("tipo") or "").lower()
                if (pl_id and pl_id in p_nom_low) or (pl_tipo and pl_tipo in p_nom_low):
                    plantilla_sel = pl
                    break
            if not plantilla_sel and len(plantillas) > 0:
                plantilla_sel = plantillas[0]
                
            if plantilla_sel:
                cuerpo_template = plantilla_sel.get("cuerpo")
                titulo_consentimiento = plantilla_sel.get("titulo") or titulo_consentimiento
            else:
                cuerpo_template = (
                    "Por medio de la presente, yo {paciente}, DNI {dni}, declaro que he sido debidamente informado "
                    "por el Dr. {cirujano} acerca de la intervención de {cirugia} en mi {ojo_intervenido} en {quirofano}. "
                    "Comprendo y acepto los riesgos, beneficios y cuidados médicos indicados."
                )
            
        ojo = turno.get("ojo") or "OD"
        ojo_desc = "OJO DERECHO (OD)" if ojo == "OD" else "OJO IZQUIERDO (OI)" if ojo == "OI" else "AMBOS OJOS (AO)"
        
        cuerpo_renderizado = render_consent_template(
            cuerpo_template,
            {
                "paciente": paciente.get("nombre") or "Paciente",
                "dni": paciente.get("dni") or "-",
                "cirujano": turno.get("cirujano_nombre") or "Médico Cirujano",
                "medico": turno.get("cirujano_nombre") or "Médico Cirujano",
                "practica": turno.get("practica_nombre") or "Cirugía Oftalmológica",
                "cirugia": turno.get("practica_nombre") or "Cirugía Oftalmológica",
                "ojo_intervenido": ojo_desc,
                "ojo": ojo_desc,
                "quirofano": (turno.get("quirofanos") or {}).get("nombre") or "Quirófano Central",
                "fecha": str(turno.get("fecha_cirugia") or ""),
                "fecha_cirugia": str(turno.get("fecha_cirugia") or ""),
                "hora_cirugia": str(turno.get("hora_inicio") or "")[:5],
                "hora_inicio": str(turno.get("hora_inicio") or "")[:5]
            }
        )
        
        return {
            "success": True,
            "turno": {
                "id": turno.get("id"),
                "fecha_cirugia": turno.get("fecha_cirugia"),
                "hora_inicio": turno.get("hora_inicio"),
                "practica_nombre": turno.get("practica_nombre"),
                "ojo": turno.get("ojo"),
                "ojo_desc": ojo_desc,
                "cirujano_nombre": turno.get("cirujano_nombre"),
                "quirofano_nombre": (turno.get("quirofanos") or {}).get("nombre") or "Quirófano",
                "tipo_anestesia": turno.get("tipo_anestesia"),
                "consentimiento_estado": turno.get("consentimiento_estado"),
                "consentimiento_pdf_url": turno.get("consentimiento_pdf_url")
            },
            "paciente": {
                "nombre": paciente.get("nombre"),
                "dni": paciente.get("dni"),
                "nro_hc": paciente.get("nro_hc"),
                "obra_social": paciente.get("obra_social")
            },
            "consentimiento": {
                "titulo": titulo_consentimiento,
                "cuerpo": cuerpo_renderizado
            }
        }
    except Exception as e:
        logger.error(f"Error al obtener consentimiento público: {e}")
        raise HTTPException(status_code=500, detail=str(e))

class FirmaPayload(BaseModel):
    firma_base64: str
    ip_origen: Optional[str] = "127.0.0.1"
    user_agent: Optional[str] = "Mobile-Web"

@app.post("/api/consentimiento-publico/{token}/firmar")
async def firmar_consentimiento_publico(token: str, payload: FirmaPayload):
    try:
        res = registrar_firma_consentimiento(
            token=token,
            firma_base64=payload.firma_base64,
            ip_origen=payload.ip_origen or "Web-Client",
            user_agent=payload.user_agent or "Browser"
        )
        if not res.get("success"):
            raise HTTPException(status_code=400, detail=res.get("error") or "Error al registrar la firma")
            
        # Opcional: Enviar confirmación por WhatsApp
        turno = get_consentimiento_by_token(token)
        if turno:
            paciente = turno.get("pacientes") or {}
            tel = paciente.get("telefono")
            if tel:
                config = get_configuracion_quirofano()
                conf_msg = config.get("whatsapp_mensaje_confirmacion") or (
                    "¡Muchas gracias {paciente}! Hemos registrado tu consentimiento firmado digitalmente para tu cirugía del {fecha_cirugia}. "
                    "Recordá concurrir con 8 horas de ayuno total."
                )
                msg_ok = conf_msg.format(
                    paciente=paciente.get("nombre") or "Paciente",
                    fecha_cirugia=str(turno.get("fecha_cirugia") or "")
                )
                jid = tel if "@" in tel else f"{tel}@s.whatsapp.net"
                try:
                    whatsapp_manager.enviar_mensaje(jid, msg_ok)
                except Exception as err_w:
                    logger.warning(f"No se pudo enviar confirmación por WhatsApp tras firma: {err_w}")
                    
        return res
    except Exception as e:
        logger.error(f"Error al firmar consentimiento: {e}")
        raise HTTPException(status_code=500, detail=str(e))




# ====================================================================
# ENDPOINTS: PRESTADORES (INSTRUMENTADORES, ANESTESISTAS)
# ====================================================================

@app.get("/api/prestadores")
def listar_prestadores_endpoint(rol: Optional[str] = None, solo_activos: bool = False):
    try:
        items = get_prestadores(rol=rol, solo_activos=solo_activos)
        return {"success": True, "prestadores": items}
    except Exception as e:
        logger.error(f"Error listando prestadores: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/prestadores")
def crear_prestador_endpoint(payload: Dict[str, Any] = Body(...)):
    try:
        nuevo = crear_prestador(payload)
        return {"success": True, "prestador": nuevo}
    except Exception as e:
        logger.error(f"Error creando prestador: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.put("/api/prestadores/{prestador_id}")
def actualizar_prestador_endpoint(prestador_id: str, payload: Dict[str, Any] = Body(...)):
    try:
        act = actualizar_prestador(prestador_id, payload)
        return {"success": True, "prestador": act}
    except Exception as e:
        logger.error(f"Error actualizando prestador {prestador_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/api/prestadores/{prestador_id}")
def eliminar_prestador_endpoint(prestador_id: str):
    try:
        ok = eliminar_prestador(prestador_id)
        return {"success": ok}
    except Exception as e:
        logger.error(f"Error eliminando prestador {prestador_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))



@app.get("/api/asesorias-quirurgicas/{asesoria_id}/consentimiento")
def obtener_consentimiento_asesoria(asesoria_id: str):
    if not supabase:
        return {"success": False, "consentimiento": None}
    try:
        resp = supabase.table("turnos_quirofano").select("*, pacientes(*)").eq("asesoria_id", asesoria_id).order("created_at", desc=True).limit(1).execute()
        if not resp.data:
            return {"success": False, "consentimiento": None}
        t = resp.data[0]
        return {
            "success": True,
            "consentimiento": {
                "turno_id": t.get("id"),
                "estado": t.get("consentimiento_estado"),
                "token": t.get("consentimiento_token"),
                "pdf_url": t.get("consentimiento_pdf_url"),
                "firmado_at": t.get("consentimiento_firmado_at"),
                "firma_ip": t.get("consentimiento_firma_ip"),
                "fecha_cirugia": t.get("fecha_cirugia"),
                "hora_inicio": t.get("hora_inicio"),
                "practica_nombre": t.get("practica_nombre"),
                "cirujano_nombre": t.get("cirujano_nombre")
            }
        }
    except Exception as e:
        logger.error(f"Error al obtener consentimiento de asesoría {asesoria_id}: {e}")
        return {"success": False, "error": str(e)}

@app.get("/api/turnos-quirofano/{turno_id}/parte-quirurgico")
def obtener_o_generar_parte_quirurgico(turno_id: str):
    """
    Retorna o genera el Protocolo / Parte Quirúrgico Oficial en PDF para el turno indicado.
    """
    try:
        from app.db import supabase
        from app.services.pdf_service import generar_pdf_parte_quirurgico
        if not supabase:
            raise HTTPException(status_code=500, detail="Sin conexión a BD")
            
        t_resp = supabase.table("turnos_quirofano").select("*, pacientes(*), quirofanos(nombre, codigo)").eq("id", turno_id).limit(1).execute()
        if not t_resp.data:
            raise HTTPException(status_code=404, detail="Turno no encontrado")
            
        turno_item = t_resp.data[0]
        paciente_data = turno_item.get("pacientes") or {}
        
        pdf_filename = generar_pdf_parte_quirurgico(turno_item, paciente_data)
        pdf_rel_url = f"/static/{pdf_filename}"
        
        supabase.table("turnos_quirofano").update({"parte_quirurgico_pdf_url": pdf_rel_url}).eq("id", turno_id).execute()
        if turno_item.get("asesoria_id"):
            supabase.table("asesorias_quirurgicas").update({"parte_quirurgico_pdf_url": pdf_rel_url}).eq("id", turno_item["asesoria_id"]).execute()
            
        return {
            "success": True,
            "pdf_url": pdf_rel_url,
            "filename": pdf_filename,
            "turno_id": turno_id
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error generando Parte Quirúrgico para turno {turno_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/turnos-quirofano/{turno_id}/checklist-seguridad")
def registrar_checklist_seguridad(turno_id: str, payload: Dict[str, Any] = Body(...)):
    """
    Registra las verificaciones de la Pausa Quirúrgica OMS (Sign-In / Time-Out / Sign-Out).
    """
    try:
        from app.db import guardar_checklist_seguridad_turno
        res = guardar_checklist_seguridad_turno(turno_id, payload)
        if not res.get("success"):
            raise HTTPException(status_code=400, detail=res.get("error") or "Error guardando checklist")
        return res
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error guardando checklist para turno {turno_id}: {e}")
@app.post("/api/turnos-quirofano/{turno_id}/subir-consentimiento-geclisa")
def subir_consentimiento_geclisa_endpoint(turno_id: str, payload: Optional[Dict[str, Any]] = None):
    """
    Sube el Consentimiento Informado firmado digitalmente a la Historia Clínica de Geclisa.
    """
    try:
        usuario_crm = (payload or {}).get("usuario_crm") if isinstance(payload, dict) else None
        res = subir_consentimiento_turno_a_geclisa(turno_id, usuario_crm=usuario_crm)
        if not res.get("success"):
            raise HTTPException(status_code=400, detail=res.get("error") or "Error al subir consentimiento a Geclisa.")
        return res
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error subiendo consentimiento a Geclisa para turno {turno_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/turnos-quirofano/{turno_id}/subir-parte-quirurgico-geclisa")
def subir_parte_quirurgico_geclisa_endpoint(turno_id: str, payload: Optional[Dict[str, Any]] = None):
    """
    Sube el Protocolo / Parte Quirúrgico Oficial a la Historia Clínica de Geclisa.
    """
    try:
        usuario_crm = (payload or {}).get("usuario_crm") if isinstance(payload, dict) else None
        res = subir_parte_quirurgico_turno_a_geclisa(turno_id, usuario_crm=usuario_crm)
        if not res.get("success"):
            raise HTTPException(status_code=400, detail=res.get("error") or "Error al subir protocolo quirúrgico a Geclisa.")
        return res
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error subiendo protocolo quirúrgico a Geclisa para turno {turno_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/api/turnos-quirofano/{turno_id}/desvincular-documento-geclisa/{tipo_doc}")
def desvincular_documento_geclisa_endpoint(turno_id: str, tipo_doc: str):
    """
    Elimina el documento en Geclisa y limpia el estado de sincronización.
    """
    if tipo_doc not in ["consentimiento", "parte_quirurgico"]:
        raise HTTPException(status_code=400, detail="tipo_doc debe ser 'consentimiento' o 'parte_quirurgico'.")
    try:
        res = desvincular_documento_geclisa_turno(turno_id, tipo_doc)
        if not res.get("success"):
            raise HTTPException(status_code=400, detail=res.get("error") or "Error al desvincular documento de Geclisa.")
        return res
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error desvinculando documento {tipo_doc} de turno {turno_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/geclisa/pacientes/{paciente_id}/archivos")
def listar_archivos_paciente_geclisa(paciente_id: str):
    """
    Lista todos los documentos y archivos adjuntos en la Historia Clínica de Geclisa para el paciente.
    Soporta paciente_id como UUID del CRM, DNI o Ficha ID directamente.
    """
    try:
        from app.services.geclisa_client import geclisa_client
        ficha_id = None
        paciente_nombre = ""
        paciente_crm_id = None
        dni = None

        # 1. Intentar buscar en Supabase (por UUID, DNI o ficha_id)
        if supabase:
            try:
                res_paciente = None
                # Búsqueda por UUID solo si tiene formato UUID válido
                if len(str(paciente_id)) == 36 and '-' in str(paciente_id):
                    res_paciente = supabase.table("pacientes").select("id, nombre, dni, geclisa_ficha_id").eq("id", paciente_id).execute()

                if not res_paciente or not res_paciente.data:
                    # Búsqueda por geclisa_ficha_id si es numérico
                    if str(paciente_id).isdigit():
                        res_paciente = supabase.table("pacientes").select("id, nombre, dni, geclisa_ficha_id").eq("geclisa_ficha_id", int(paciente_id)).execute()

                if not res_paciente or not res_paciente.data:
                    # Búsqueda por DNI
                    res_paciente = supabase.table("pacientes").select("id, nombre, dni, geclisa_ficha_id").ilike("dni", f"%{paciente_id}%").execute()

                if res_paciente and res_paciente.data and len(res_paciente.data) > 0:
                    pac = res_paciente.data[0]
                    paciente_crm_id = pac.get("id")
                    ficha_id = pac.get("geclisa_ficha_id")
                    dni = pac.get("dni")
                    paciente_nombre = pac.get("nombre") or ""
            except Exception as db_err:
                logger.warning(f"Aviso al consultar paciente en Supabase para archivos: {db_err}")

        # 2. Si no se resolvió ficha_id pero tenemos DNI o entrada numérica
        if not ficha_id:
            if str(paciente_id).isdigit():
                if len(str(paciente_id)) < 7:
                    ficha_id = int(paciente_id)
                else:
                    dni = str(paciente_id)

        if not ficha_id and dni:
            res_d = geclisa_client.buscar_paciente_por_dni(str(dni))
            if res_d.get("encontrado") and res_d.get("ficha_id"):
                ficha_id = int(res_d["ficha_id"])
                if paciente_crm_id and supabase:
                    try:
                        supabase.table("pacientes").update({"geclisa_ficha_id": ficha_id}).eq("id", paciente_crm_id).execute()
                    except Exception as upd_err:
                        logger.warning(f"No se pudo guardar geclisa_ficha_id: {upd_err}")

        if not ficha_id:
            return {"success": False, "encontrado": False, "mensaje": "El paciente no tiene Ficha registrada en Geclisa.", "archivos": []}
            
        archivos_raw = geclisa_client.listar_archivos_historia_clinica(ficha_id)
        archivos_norm = []
        for arc in archivos_raw:
            as_id = arc.get("asId") or arc.get("id")
            if not as_id:
                continue
            ext = (arc.get("extension") or "pdf").lower().replace(".", "")
            titulo = arc.get("titulo") or f"Documento #{as_id}"
            archivos_norm.append({
                "id": arc.get("id") or as_id,
                "as_id": as_id,
                "asId": as_id,
                "titulo": titulo,
                "fecha": arc.get("fecha") or "",
                "hora": arc.get("hora") or "",
                "prestador": arc.get("preNombre") or arc.get("prestador") or "",
                "preNombre": arc.get("preNombre") or "",
                "clase": arc.get("acNombre") or arc.get("clase") or "Clínicos",
                "acNombre": arc.get("acNombre") or "Clínicos",
                "observaciones": arc.get("observaciones") or "",
                "formato": ext,
                "extension": ext,
                "url": f"/api/geclisa/archivos/{as_id}/ver",
                "download_url": f"/api/geclisa/archivos/{as_id}/descargar"
            })
            
        return {
            "success": True,
            "encontrado": True,
            "ficha_id": ficha_id,
            "paciente_nombre": paciente_nombre,
            "archivos": archivos_norm,
            "total_archivos": len(archivos_norm)
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error al listar archivos de Geclisa para paciente {paciente_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/geclisa/archivos/{as_id}/ver")
def ver_archivo_geclisa_endpoint(as_id: int):
    """
    Transmite el contenido binario del archivo de Geclisa para visualización directa en navegadores / iframe.
    """
    try:
        from app.services.geclisa_client import geclisa_client
        file_bytes, content_type, filename = geclisa_client.descargar_archivo_historia_clinica(as_id)
        if not file_bytes:
            raise HTTPException(status_code=404, detail=f"No se pudo obtener el archivo #{as_id} desde Geclisa.")
        return Response(
            content=file_bytes,
            media_type=content_type,
            headers={
                "Content-Disposition": f"inline; filename=\"{filename}\"",
                "Cache-Control": "public, max-age=3600"
            }
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error al visualizar archivo #{as_id} de Geclisa: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/geclisa/archivos/{as_id}/descargar")
def descargar_archivo_geclisa_endpoint(as_id: int, nombre: Optional[str] = None):
    """
    Transmite el archivo de Geclisa forzando la descarga con un nombre de archivo legible.
    """
    try:
        from app.services.geclisa_client import geclisa_client
        import re
        file_bytes, content_type, filename = geclisa_client.descargar_archivo_historia_clinica(as_id)
        if not file_bytes:
            raise HTTPException(status_code=404, detail=f"No se pudo descargar el archivo #{as_id} desde Geclisa.")
            
        download_name = filename
        if nombre:
            ext = filename.split(".")[-1] if "." in filename else "pdf"
            clean_name = re.sub(r'[^a-zA-Z0-9_\-\.]', '_', nombre).strip('_')
            if clean_name:
                if not clean_name.lower().endswith(f".{ext.lower()}"):
                    clean_name += f".{ext}"
                download_name = clean_name
                
        return Response(
            content=file_bytes,
            media_type=content_type,
            headers={
                "Content-Disposition": f"attachment; filename=\"{download_name}\"",
                "Cache-Control": "public, max-age=3600"
            }
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error al descargar archivo #{as_id} de Geclisa: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/api/geclisa/archivos/{as_id}")
def eliminar_archivo_geclisa_endpoint(as_id: int):
    """
    Elimina un archivo adjunto del repositorio de Geclisa por su ID de archivo (asId / hcaId).
    """
    try:
        from app.services.geclisa_client import geclisa_client
        res = geclisa_client.eliminar_archivo_historia_clinica(as_id, as_id)
        if not res.get("success"):
            raise HTTPException(status_code=400, detail=res.get("mensaje") or "Error al eliminar archivo de Geclisa.")
        return res
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error eliminando archivo #{as_id} de Geclisa: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# ====================================================================
# ENDPOINTS: INTEGRACIÓN GECLISA ELEMENTOS & LENTES INTRAOCULARES
# ====================================================================

@app.get("/api/geclisa/elementos/buscar")
def buscar_elementos_geclisa_endpoint(q: str = Query(..., min_length=1, description="Término de búsqueda o código GTIN")):
    """
    Busca elementos en Geclisa por nombre, código comercial o GTIN vía /api/Elementos/autocomplete.
    """
    try:
        from app.services.geclisa_client import geclisa_client
        elementos = geclisa_client.buscar_elementos(q)
        return {"success": True, "elementos": elementos, "total": len(elementos)}
    except Exception as e:
        logger.error(f"Error al buscar elementos en Geclisa con query '{q}': {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/geclisa/elementos/{ele_id}/stock-lotes")
def obtener_stock_lotes_geclisa_endpoint(ele_id: int):
    """
    Obtiene el stock consolidado (Quirófano y Consignación) y lotes activos para un eleId de Geclisa.
    """
    try:
        from app.services.geclisa_client import geclisa_client
        resumen = geclisa_client.obtener_resumen_stock_lotes(ele_id)
        return {"success": True, "resumen": resumen}
    except Exception as e:
        logger.error(f"Error al obtener stock y lotes para eleId {ele_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/modelos-lio")
def listar_modelos_lio_endpoint(solo_activos: bool = False):
    try:
        items = get_modelos_lio(solo_activos=solo_activos)
        return {"success": True, "modelos": items}
    except Exception as e:
        logger.error(f"Error listando modelos de LIO: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/modelos-lio/{modelo_id}")
def obtener_modelo_lio_endpoint(modelo_id: str):
    try:
        modelo = get_modelo_lio_por_id(modelo_id)
        if not modelo:
            raise HTTPException(status_code=404, detail="Modelo de LIO no encontrado.")
        items = get_modelos_lio_items(modelo_id)
        modelo["items"] = items
        return {"success": True, "modelo": modelo}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error obteniendo modelo de LIO {modelo_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/modelos-lio")
def crear_modelo_lio_endpoint(payload: Dict[str, Any] = Body(...)):
    try:
        nuevo = crear_modelo_lio(payload)
        return {"success": True, "modelo": nuevo}
    except Exception as e:
        logger.error(f"Error creando modelo de LIO: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.put("/api/modelos-lio/{modelo_id}")
def actualizar_modelo_lio_endpoint(modelo_id: str, payload: Dict[str, Any] = Body(...)):
    try:
        act = actualizar_modelo_lio(modelo_id, payload)
        return {"success": True, "modelo": act}
    except Exception as e:
        logger.error(f"Error actualizando modelo de LIO {modelo_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/api/modelos-lio/{modelo_id}")
def eliminar_modelo_lio_endpoint(modelo_id: str):
    try:
        ok = eliminar_modelo_lio(modelo_id)
        return {"success": ok}
    except Exception as e:
        logger.error(f"Error eliminando modelo de LIO {modelo_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# --- ÍTEMS / SKUS (MAPEADOS POR GTIN / DIOPTRÍA) ---

@app.get("/api/modelos-lio/{modelo_id}/items")
def listar_items_modelo_lio_endpoint(modelo_id: str):
    try:
        items = get_modelos_lio_items(modelo_id)
        return {"success": True, "items": items, "total": len(items)}
    except Exception as e:
        logger.error(f"Error listando items para modelo LIO {modelo_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/modelos-lio/{modelo_id}/items")
def crear_item_modelo_lio_endpoint(modelo_id: str, payload: Dict[str, Any] = Body(...)):
    try:
        payload["modelo_lio_id"] = modelo_id
        nuevo_item = crear_modelo_lio_item(payload)
        return {"success": True, "item": nuevo_item}
    except Exception as e:
        logger.error(f"Error creando item para modelo LIO {modelo_id}: {e}")
        raise HTTPException(status_code=400, detail=f"No se pudo guardar la graduación (verifique que no esté duplicada): {str(e)}")

@app.delete("/api/modelos-lio/items/{item_id}")
def eliminar_item_modelo_lio_endpoint(item_id: str):
    try:
        ok = eliminar_modelo_lio_item(item_id)
        return {"success": ok}
    except Exception as e:
        logger.error(f"Error eliminando item LIO {item_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/modelos-lio-items")
def listar_todos_items_lio_endpoint(modelo_lio_id: Optional[str] = Query(None), q: Optional[str] = Query(None)):
    try:
        items = get_all_modelos_lio_items(modelo_lio_id, q)
        return {"success": True, "items": items, "total": len(items)}
    except Exception as e:
        logger.error(f"Error listando todos los items LIO: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/modelos-lio/validar-gtin")
def validar_gtin_endpoint(gtin: str = Query(...), exclude_id: Optional[str] = Query(None)):
    try:
        res = validar_gtin_unico(gtin, exclude_id)
        return {"success": True, **res}
    except Exception as e:
        logger.error(f"Error validando GTIN {gtin}: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/modelos-lio/sincronizar-masivo-alcon")
def sincronizar_masivo_alcon_endpoint():
    """
    Escanea los elementos de Geclisa para la marca Alcon, los cruza contra el catálogo maestro
    de 3.895 GTINs, y crea/actualiza automáticamente las familias y graduaciones en el CRM.
    """
    try:
        # 1. Búsqueda exhaustiva en Geclisa para encontrar todos los elementos Alcon
        terminos_busqueda = [
            "CLAREON", "PANOPTIX", "VIVITY", "SN60WF", "ACRYSOF", "AU00T0",
            "TFNT", "CNA0", "SY60WF", "MA60AC", "MN60", "SND1", "SV25", "DFT",
            "0038065", "38065", "CNW", "CNA", "ALCON", "ACRY"
        ]
        elementos_dict = {}
        for term in terminos_busqueda:
            try:
                res = geclisa_client.buscar_elementos(term, limite=100)
                for el in res:
                    if el.get("eleId"):
                        elementos_dict[el["eleId"]] = el
            except Exception as ex_t:
                logger.warning(f"Falla al buscar término '{term}' en Geclisa: {ex_t}")

        elementos_geclisa = list(elementos_dict.values())
        logger.info(f"[SINCRONIZACION_ALCON] {len(elementos_geclisa)} elementos únicos obtenidos de Geclisa.")

        # 2. Cruzar contra catálogo Alcon GTIN
        coincidencias = alcon_catalog_service.cruzar_elementos_geclisa(elementos_geclisa)
        logger.info(f"[SINCRONIZACION_ALCON] {len(coincidencias)} coincidencias exactas con el catálogo de 3.895 GTINs.")

        # 3. Sincronizar en DB
        resultado = sincronizar_lentes_masivos(coincidencias)

        return {
            "success": True,
            "total_geclisa_analizados": len(elementos_geclisa),
            "total_coincidencias_gtin": len(coincidencias),
            "resultado": resultado
        }
    except Exception as e:
        logger.error(f"Error en sincronización masiva Alcon: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/alcon/catalogo-completo")
def obtener_catalogo_alcon_completo():
    """
    Retorna la lista completa de los 3.895 SKUs de Alcon con GTIN, Nombres y Dioptrías.
    """
    try:
        items = alcon_catalog_service.get_catalogo_completo()
        return {
            "success": True,
            "total": len(items),
            "items": items
        }
    except Exception as e:
        logger.error(f"Error al obtener catálogo Alcon: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# ====================================================================
# ENDPOINTS: CATÁLOGO MAESTRO UNIFICADO DE GTINs, LENTES E INSUMOS (ABM)
# ====================================================================

class ItemCatalogoMaestroPayload(BaseModel):
    gtin_14: str
    gtin_12: Optional[str] = None
    marca: str = "Alcon"
    nombre_producto: str
    internacional: Optional[str] = None
    categoria: str = "LIO"
    familia_nombre: Optional[str] = None
    modelo_lio_id: Optional[str] = None
    tipo_optica: Optional[str] = "Monofocal Asférico"
    dioptria: Optional[float] = None
    es_torico: bool = False
    torico_valor: Optional[str] = None
    constante_a: Optional[float] = 118.9
    acd_estimado: Optional[float] = 5.0
    geclisa_ele_id: Optional[int] = None
    geclisa_ele_cod: Optional[str] = None
    activo: bool = True
    origen: str = "MANUAL"
    observaciones: Optional[str] = None

class ActualizarItemCatalogoMaestroPayload(BaseModel):
    gtin_14: Optional[str] = None
    gtin_12: Optional[str] = None
    marca: Optional[str] = None
    nombre_producto: Optional[str] = None
    internacional: Optional[str] = None
    categoria: Optional[str] = None
    familia_nombre: Optional[str] = None
    modelo_lio_id: Optional[str] = None
    tipo_optica: Optional[str] = None
    dioptria: Optional[float] = None
    es_torico: Optional[bool] = None
    torico_valor: Optional[str] = None
    constante_a: Optional[float] = None
    acd_estimado: Optional[float] = None
    geclisa_ele_id: Optional[int] = None
    geclisa_ele_cod: Optional[str] = None
    activo: Optional[bool] = None
    observaciones: Optional[str] = None

@app.get("/api/catalogo-maestro")
def listar_catalogo_maestro_endpoint(
    busqueda: Optional[str] = None,
    marca: Optional[str] = "ALL",
    categoria: Optional[str] = "ALL",
    solo_activos: bool = True,
    limit: int = 50,
    offset: int = 0
):
    try:
        data = get_catalogo_maestro(
            busqueda=busqueda,
            marca=marca,
            categoria=categoria,
            solo_activos=solo_activos,
            limit=limit,
            offset=offset
        )
        return {"success": True, "total": data["total"], "items": data["items"]}
    except Exception as e:
        logger.error(f"Error al listar catalogo maestro: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/catalogo-maestro")
def crear_item_catalogo_maestro_endpoint(payload: ItemCatalogoMaestroPayload):
    try:
        nuevo = crear_item_catalogo_maestro(payload.dict())
        return {"success": True, "item": nuevo}
    except Exception as e:
        logger.error(f"Error al crear item en catalogo maestro: {e}")
        raise HTTPException(status_code=400, detail=str(e))

@app.put("/api/catalogo-maestro/{item_id}")
def actualizar_item_catalogo_maestro_endpoint(item_id: str, payload: ActualizarItemCatalogoMaestroPayload):
    try:
        datos = {k: v for k, v in payload.dict().items() if v is not None}
        act = actualizar_item_catalogo_maestro(item_id, datos)
        return {"success": True, "item": act}
    except Exception as e:
        logger.error(f"Error al actualizar item #{item_id} en catalogo maestro: {e}")
        raise HTTPException(status_code=400, detail=str(e))

@app.delete("/api/catalogo-maestro/{item_id}")
def eliminar_item_catalogo_maestro_endpoint(item_id: str, fisico: bool = False):
    try:
        ok = eliminar_item_catalogo_maestro(item_id, fisico=fisico)
        return {"success": ok}
    except Exception as e:
        logger.error(f"Error al eliminar item #{item_id} de catalogo maestro: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/catalogo-maestro/sincronizar-geclisa")
def sincronizar_catalogo_maestro_con_geclisa_endpoint():
    """
    Cruza los registros del catálogo maestro contra Geclisa por GTIN o código comercial y asocia los eleId (Modo Estricto B).
    """
    try:
        from app.services.geclisa_client import geclisa_client
        
        # 1. Barrido exhaustivo multi-prefijo en Geclisa
        terminos_busqueda = [
            "0038065", "38065", "CLAREON", "PANOPTIX", "VIVITY", "SN60WF", "ACRYSOF", "AU00T0",
            "TFNT", "CNA0", "SY60WF", "MA60AC", "MN60", "SND1", "SV25", "DFT", "PXY", "CNW", "CNA",
            "ALCON", "TECNIS", "ICB", "ZCB", "DCB", "ZEISS", "LISA", "RAYNER", "EMV", "600U",
            "HOYA", "BAUSCH", "ANILLO", "VISCO", "PROVISC", "DUOVISC", "HEALON", "DISCOVISC", "GAS"
        ]

        elementos_dict = {}
        for term in terminos_busqueda:
            try:
                res = geclisa_client.buscar_elementos(term, limite=200)
                for el in res:
                    if el.get("eleId"):
                        elementos_dict[el["eleId"]] = el
            except Exception as e_t:
                logger.warning(f"Error buscando término '{term}' en Geclisa: {e_t}")

        elementos_geclisa = list(elementos_dict.values())
        if not elementos_geclisa:
            return {"success": True, "total_sincronizados": 0, "mensaje": "No se encontraron elementos en Geclisa."}

        # 2. Indexar elementos de Geclisa por GTIN a 14 y 12 dígitos
        ele_map_gtin = {}
        for el in elementos_geclisa:
            g = str(el.get("eleCod") or "").strip()
            if g:
                ele_map_gtin[g.zfill(14)] = el
                ele_map_gtin[g.lstrip("0")] = el

        # 3. Traer todos los ítems de catalogo_maestro_gtin (con paginación interna)
        todos_items = []
        offset = 0
        while True:
            cat_res = supabase.table("catalogo_maestro_gtin").select("id, gtin_14, gtin_12, modelo_lio_id, familia_nombre, dioptria, es_torico, torico_valor, nombre_producto").range(offset, offset + 999).execute()
            rows = cat_res.data or []
            todos_items.extend(rows)
            if len(rows) < 1000:
                break
            offset += 1000

        sincronizados = 0

        # Traer familias de modelos_lio para asignación automática si no está vinculado
        fam_res = supabase.table("modelos_lio").select("id, modelo, marca").execute()
        familias_crm = fam_res.data or []

        for it in todos_items:
            g14 = str(it.get("gtin_14") or "").strip()
            g12 = str(it.get("gtin_12") or "").strip()
            match = ele_map_gtin.get(g14) or (ele_map_gtin.get(g12) if g12 else None)

            if match:
                ele_id = match.get("eleId")
                ele_cod = match.get("eleCod")
                ele_nom = match.get("eleNombre")

                # Actualizar registro en catalogo_maestro_gtin
                supabase.table("catalogo_maestro_gtin").update({
                    "geclisa_ele_id": ele_id,
                    "geclisa_ele_cod": ele_cod,
                    "updated_at": "now()"
                }).eq("id", it["id"]).execute()

                # Vincular en modelos_lio_items si corresponde
                mod_id = it.get("modelo_lio_id")
                if not mod_id and it.get("familia_nombre"):
                    # Intentar buscar familia por nombre aproximado
                    fn = str(it["familia_nombre"]).lower()
                    for f in familias_crm:
                        if f["modelo"].lower() in fn or fn in f["modelo"].lower() or ("vivity" in fn and "vivity" in f["modelo"].lower()) or ("panoptix" in fn and "panoptix" in f["modelo"].lower()):
                            mod_id = f["id"]
                            break

                if mod_id and it.get("dioptria") is not None:
                    try:
                        crear_modelo_lio_item({
                            "modelo_lio_id": mod_id,
                            "geclisa_ele_id": ele_id,
                            "geclisa_ele_cod": ele_cod,
                            "geclisa_nombre": ele_nom,
                            "dioptria": it["dioptria"],
                            "es_torico": it.get("es_torico", False),
                            "torico_valor": it.get("torico_valor")
                        })
                    except Exception:
                        pass

                sincronizados += 1

        return {
            "success": True,
            "total_geclisa": len(elementos_geclisa),
            "total_catalogo": len(todos_items),
            "total_sincronizados": sincronizados
        }
    except Exception as e:
        logger.error(f"Error en sincronizacion maestro Geclisa: {e}")
        raise HTTPException(status_code=500, detail=str(e))
        logger.error(f"Error en sincronizacion maestro Geclisa: {e}")
        raise HTTPException(status_code=500, detail=str(e))

class ResolverSkuPayload(BaseModel):
    modelo_lio_id: Optional[str] = None
    modelo_nombre: Optional[str] = None
    dioptria: float
    torico_valor: Optional[str] = None
    es_torico: Optional[bool] = False

@app.post("/api/modelos-lio/resolver-sku")
def resolver_sku_lio_endpoint(payload: ResolverSkuPayload):
    """
    Resuelve el SKU/GTIN de Geclisa para una combinación de Modelo, Dioptría y Toricidad,
    y consulta en tiempo real el stock en Quirófano (Dep 1) y Consignación (Dep 3) con sus lotes.
    """
    try:
        item = resolver_sku_lio(
            modelo_lio_id=payload.modelo_lio_id,
            modelo_nombre=payload.modelo_nombre,
            dioptria=payload.dioptria,
            torico_valor=payload.torico_valor,
            es_torico=bool(payload.es_torico)
        )
        if not item:
            return {
                "success": True,
                "mapeado": False,
                "mensaje": "Graduación no mapeada a código GTIN Geclisa.",
                "item": None,
                "stock": None,
                "resumen": None
            }

        # Consultar stock y lotes en vivo en Geclisa para este eleId
        from app.services.geclisa_client import geclisa_client
        resumen_stock = geclisa_client.obtener_resumen_stock_lotes(item["geclisa_ele_id"])

        return {
            "success": True,
            "mapeado": True,
            "item": item,
            "stock": resumen_stock,
            "resumen": resumen_stock
        }
    except Exception as e:
        logger.error(f"Error resolviendo SKU LIO: {e}")
        raise HTTPException(status_code=500, detail=str(e))




# ====================================================================
# ENDPOINTS: CONTROL DE QUIRÓFANO EN VIVO & RECEPCIÓN DEL DÍA
# ====================================================================

@app.get("/api/turnos-quirofano-dia")
def listar_turnos_dia_endpoint(
    fecha: str = Query(..., description="Fecha en formato YYYY-MM-DD"),
    quirofano_id: Optional[str] = Query(None)
):
    try:
        items = get_turnos_dia_ejecucion(fecha=fecha, quirofano_id=quirofano_id)
        return {"success": True, "turnos": items, "fecha": fecha}
    except Exception as e:
        logger.error(f"Error listando turnos del día ({fecha}): {e}")
        raise HTTPException(status_code=500, detail=str(e))

class CambiarEstadoPayload(BaseModel):
    estado: str

@app.put("/api/turnos-quirofano/{turno_id}/estado")
@app.put("/api/turnos-quirofano/{turno_id}/cambiar-estado")
def cambiar_estado_turno_endpoint(turno_id: str, payload: CambiarEstadoPayload):
    try:
        res = cambiar_estado_turno_quirofano(turno_id=turno_id, nuevo_estado=payload.estado)
        if not res.get("success"):
            raise HTTPException(status_code=400, detail=res.get("error") or "Error al actualizar estado")
        return res
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error en endpoint cambiar estado {turno_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# ====================================================================
# ENDPOINTS: AGENDA HOSPITALARIA GECLISA & GESTIÓN DE TURNOS
# ====================================================================

@app.get("/api/geclisa/prestadores")
def listar_prestadores_geclisa_endpoint(query: str = Query("", description="Término de búsqueda por nombre o matrícula")):
    """
    Lista el catálogo de prestadores activos de Geclisa para selección en usuarios y agenda.
    """
    try:
        from app.services.geclisa_client import geclisa_client
        prestadores = geclisa_client.buscar_prestadores(query)
        return {"success": True, "prestadores": prestadores, "total": len(prestadores)}
    except Exception as e:
        logger.error(f"Error al listar prestadores de Geclisa: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/geclisa/agenda")
def obtener_agenda_geclisa_endpoint(
    pre_id: Optional[int] = Query(None, description="ID del prestador en Geclisa"),
    fecha: Optional[str] = Query(None, description="Fecha en formato YYYY-MM-DD"),
    usuario_crm_id: Optional[str] = Query(None, description="ID del usuario en el CRM para resolver prestador por defecto")
):
    """
    Obtiene la agenda de turnos en vivo de Geclisa para un prestador o global de toda la clínica.
    Si pre_id no se especifica o es 0/None, obtiene la agenda global consolidada del día.
    """
    try:
        from app.services.geclisa_client import geclisa_client
        from datetime import datetime
        
        fecha_iso = fecha or datetime.now().strftime("%Y-%m-%d")
        prestador_info = None
        
        target_pre_id = pre_id
        prestador_info = None

        # 1. Si no viene pre_id, resolver desde el perfil del usuario del CRM o fallback 969
        if not target_pre_id and usuario_crm_id:
            try:
                p_resp = supabase.table("usuarios_perfil").select("geclisa_pre_id, geclisa_matricula, geclisa_prestador_nombre").eq("id", usuario_crm_id).limit(1).execute()
                if p_resp.data and p_resp.data[0].get("geclisa_pre_id"):
                    target_pre_id = int(p_resp.data[0]["geclisa_pre_id"])
                    prestador_info = {
                        "pre_id": target_pre_id,
                        "matricula": p_resp.data[0].get("geclisa_matricula") or "",
                        "nombre": p_resp.data[0].get("geclisa_prestador_nombre") or f"Prestador #{target_pre_id}"
                    }
            except Exception as e:
                logger.warning(f"Error resolviendo prestador de usuario: {e}")

        if not target_pre_id:
            target_pre_id = 969

        # 2. Consultar turnos del prestador
        turnos = geclisa_client.obtener_agenda_prestador(pre_id=target_pre_id, fecha_iso=fecha_iso)
        
        if not prestador_info:
            p_data = geclisa_client.obtener_prestador_por_id(target_pre_id)
            if p_data.get("encontrado"):
                prestador_info = {
                    "pre_id": target_pre_id,
                    "nombre": p_data.get("nombre"),
                    "matricula": p_data.get("matricula")
                }
            elif turnos:
                prestador_info = {
                    "pre_id": target_pre_id,
                    "nombre": turnos[0].get("prestador_nombre") or f"Prestador #{target_pre_id}",
                    "matricula": ""
                }
            else:
                prestador_info = {"pre_id": target_pre_id, "nombre": f"Prestador #{target_pre_id}", "matricula": ""}
        
        # 3. Calcular métricas agregadas de estados
        metricas = {
            "total": len(turnos),
            "reservado": sum(1 for t in turnos if t.get("estado_key") == "reservado"),
            "confirmado": sum(1 for t in turnos if t.get("estado_key") == "confirmado"),
            "ingresado": sum(1 for t in turnos if t.get("estado_key") == "ingresado"),
            "atendido": sum(1 for t in turnos if t.get("estado_key") == "atendido"),
            "cancelado": sum(1 for t in turnos if t.get("estado_key") == "cancelado")
        }
        
        # 4. Extraer catálogos dinámicos presentes en los turnos
        servicios_set = set()
        ubicaciones_set = set()
        consultorios_set = set()
        prestadores_dict = {}

        for t in turnos:
            if t.get("servicio"):
                servicios_set.add(t["servicio"])
            if t.get("ubicacion"):
                ubicaciones_set.add(t["ubicacion"])
            if t.get("consultorio"):
                consultorios_set.add(t["consultorio"])
            if t.get("prestador_id"):
                prestadores_dict[t["prestador_id"]] = t.get("prestador_nombre") or f"Prestador #{t['prestador_id']}"

        return {
            "success": True,
            "pre_id": target_pre_id,
            "prestador": prestador_info,
            "fecha": fecha_iso,
            "turnos": turnos,
            "metricas": metricas,
            "catalogos": {
                "servicios": sorted(list(servicios_set)),
                "ubicaciones": sorted(list(ubicaciones_set)),
                "consultorios": sorted(list(consultorios_set)),
                "prestadores": [{"pre_id": k, "nombre": v} for k, v in sorted(prestadores_dict.items(), key=lambda x: x[1])]
            }
        }
    except Exception as e:
        logger.error(f"Error al obtener agenda de Geclisa: {e}")
        raise HTTPException(status_code=500, detail=str(e))

class CambiarEstadoTurnoGeclisaPayload(BaseModel):
    nuevo_estado: str # 'reservado' | 'confirmado' | 'ingresado' | 'atendido' | 'cancelado'
    canal: Optional[int] = 7 # 7: WhatsApp, 1: Teléfono, 4: Personal
    motivo_id: Optional[int] = 1
    usuario_crm: Optional[str] = None

@app.put("/api/geclisa/agenda/turnos/{turno_id}/estado")
def cambiar_estado_turno_geclisa_endpoint(turno_id: int, payload: CambiarEstadoTurnoGeclisaPayload):
    """
    Cambia el estado de un turno en Geclisa (Reservado, Confirmado, Ingresado, Atendido, Cancelado).
    """
    try:
        from app.services.geclisa_client import geclisa_client
        res = geclisa_client.cambiar_estado_turno(
            turno_id=turno_id,
            nuevo_estado=payload.nuevo_estado,
            canal=payload.canal or 7,
            motivo_id=payload.motivo_id or 1
        )
        if not res.get("success"):
            raise HTTPException(status_code=400, detail=res.get("error") or "Error al actualizar estado en Geclisa.")
            
        log_event(
            nivel="INFO",
            modulo="GECLISA",
            accion="CAMBIO_ESTADO_TURNO",
            mensaje=f"Turno #{turno_id} cambiado a estado '{payload.nuevo_estado}' por {payload.usuario_crm or 'Usuario'}",
            detalles={"turno_id": turno_id, "nuevo_estado": payload.nuevo_estado, "usuario": payload.usuario_crm}
        )
        return res
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error al cambiar estado de turno #{turno_id} en Geclisa: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ====================================================================
# ENDPOINTS: CÁLCULO DE LENTE INTRAOCULAR (LIO) MULTILENTE Y STOCK
# ====================================================================

class GuardarCalculoLioPayload(BaseModel):
    turno_id: Optional[str] = None
    asesoria_id: Optional[str] = None
    paciente_id: Optional[str] = None
    lio_calculado_por: str
    opciones: List[Dict[str, Any]]
    confirmar: bool = True
    formula: Optional[str] = None
    target_refractivo: Optional[str] = None
    observaciones: Optional[str] = None
    ojo: Optional[str] = None

@app.get("/api/calculo-lio/pacientes")
def listar_pacientes_calculo_lio(
    cirujano_nombre: Optional[str] = None,
    estado_calculo: Optional[str] = "todos", # 'todos' | 'pendientes' | 'calculados'
    busqueda: Optional[str] = None
):
    """
    Lista todos los pacientes asignados a un cirujano (tanto turnos agendados como asesorías activas)
    para la gestión y cálculo de lentes intraoculares.
    """
    try:
        from datetime import datetime, timezone
        items = []
        turnos_vistos = set()
        asesorias_vistas = set()

        # 1. Consultar turnos de quirófano programados / activos
        q_turnos = supabase.table("turnos_quirofano").select(
            "*, pacientes(*), quirofanos(nombre, codigo, color), asesorias_quirurgicas(*)"
        ).neq("estado", "cancelado").order("fecha_cirugia", desc=False)

        res_turnos = q_turnos.execute()
        turnos_data = res_turnos.data or []

        for t in turnos_data:
            pac = t.get("pacientes") or {}
            as_id = t.get("asesoria_id")
            if as_id:
                asesorias_vistas.add(as_id)
            turnos_vistos.add(t["id"])

            cirujano = t.get("cirujano_nombre") or (t.get("asesorias_quirurgicas") or {}).get("medico_derivador_nombre") or "Sin Asignar"
            
            # Filtro por cirujano
            if cirujano_nombre and cirujano_nombre.lower() != "todos":
                if cirujano_nombre.lower() not in cirujano.lower():
                    continue

            es_calculado = bool(t.get("lio_calculado"))
            opciones = t.get("lio_calculo_opciones") or []

            # Si no tiene opciones pero tiene lente_tipo y lente_dioptria, armar opción 1 por compatibilidad
            if not opciones and t.get("lente_tipo"):
                opciones = [{
                    "id": "opt-base",
                    "tipo_opcion": "principal",
                    "etiqueta": "Plan A (Principal)",
                    "modelo": t.get("lente_tipo"),
                    "dioptria": t.get("lente_dioptria") or "+21.50",
                    "es_torico": bool(t.get("es_torico")),
                    "torico_valor": t.get("lente_torico_valor"),
                    "torico_eje": t.get("lente_torico_eje"),
                    "es_implantado": True
                }]

            item = {
                "tipo_registro": "turno",
                "turno_id": t["id"],
                "asesoria_id": as_id,
                "paciente_id": t.get("paciente_id"),
                "paciente": pac,
                "paciente_nombre": pac.get("nombre") or "Paciente",
                "paciente_dni": pac.get("dni") or "S/D",
                "paciente_telefono": pac.get("telefono") or "S/D",
                "paciente_obra_social": t.get("obra_social") or pac.get("obra_social") or "Particular",
                "geclisa_ficha_id": pac.get("geclisa_ficha_id"),
                "fecha_cirugia": t.get("fecha_cirugia"),
                "hora_inicio": t.get("hora_inicio"),
                "ojo": t.get("ojo") or "OD",
                "practica_nombre": t.get("practica_nombre") or "Cirugía Oftalmológica",
                "cirujano_nombre": cirujano,
                "quirofano_nombre": (t.get("quirofanos") or {}).get("nombre") or "Quirófano",
                "estado_turno": t.get("estado"),
                "lio_calculado": es_calculado,
                "lio_calculado_at": t.get("lio_calculado_at"),
                "lio_calculado_por": t.get("lio_calculado_por") or (cirujano if es_calculado else None),
                "lio_calculo_opciones": opciones,
                "lio_stock_reservado": bool(t.get("lio_stock_reservado")),
                "lio_stock_reservado_at": t.get("lio_stock_reservado_at"),
                "lio_stock_observaciones": t.get("lio_stock_observaciones") or "",
                "lente_tipo": t.get("lente_tipo"),
                "lente_dioptria": t.get("lente_dioptria"),
                "es_torico": bool(t.get("es_torico")),
                "lente_torico_valor": t.get("lente_torico_valor"),
                "lente_torico_eje": t.get("lente_torico_eje")
            }
            items.append(item)

        # 2. Consultar asesorías confirmadas o en proceso que aún no tengan turno agendado
        q_asesorias = supabase.table("asesorias_quirurgicas").select(
            "*, pacientes(*)"
        ).in_("estado", ["confirmado", "en_asesoramiento", "presupuesto_enviado", "en_analisis"]).order("created_at", desc=True)

        res_as = q_asesorias.execute()
        for a in (res_as.data or []):
            if a["id"] in asesorias_vistas:
                continue
            
            pac = a.get("pacientes") or {}
            cirujano = a.get("medico_derivador_nombre") or "Sin Asignar"
            
            if cirujano_nombre and cirujano_nombre.lower() != "todos":
                if cirujano_nombre.lower() not in cirujano.lower():
                    continue

            es_calculado = bool(a.get("lio_calculado"))
            opciones = a.get("lio_calculo_opciones") or []

            item = {
                "tipo_registro": "asesoria",
                "turno_id": None,
                "asesoria_id": a["id"],
                "paciente_id": a.get("paciente_id"),
                "paciente": pac,
                "paciente_nombre": pac.get("nombre") or "Paciente",
                "paciente_dni": pac.get("dni") or "S/D",
                "paciente_telefono": pac.get("telefono") or "S/D",
                "paciente_obra_social": a.get("cobertura_obra_social") or pac.get("obra_social") or "Particular",
                "geclisa_ficha_id": pac.get("geclisa_ficha_id"),
                "fecha_cirugia": None,
                "hora_inicio": None,
                "ojo": a.get("ojo") or "OD",
                "practica_nombre": a.get("practica_nombre") or "Cirugía Oftalmológica",
                "cirujano_nombre": cirujano,
                "quirofano_nombre": "Pendiente de Agendamiento",
                "estado_turno": "en_asesoramiento",
                "lio_calculado": es_calculado,
                "lio_calculado_at": a.get("lio_calculado_at"),
                "lio_calculado_por": a.get("lio_calculado_por") or (cirujano if es_calculado else None),
                "lio_calculo_opciones": opciones,
                "lio_stock_reservado": False,
                "lio_stock_reservado_at": None,
                "lio_stock_observaciones": "",
                "lente_tipo": None,
                "lente_dioptria": None,
                "es_torico": False,
                "lente_torico_valor": None,
                "lente_torico_eje": None
            }
            items.append(item)

        # Filtro de búsqueda por texto inicial
        if busqueda and busqueda.strip():
            b = busqueda.lower().strip()
            items = [
                it for it in items
                if b in it["paciente_nombre"].lower()
                or b in it["paciente_dni"].lower()
                or b in it["cirujano_nombre"].lower()
                or b in it["practica_nombre"].lower()
            ]

        # Extraer lista de cirujanos disponibles
        cirujanos_set = set()
        for it in items:
            if it["cirujano_nombre"] and it["cirujano_nombre"] != "Sin Asignar":
                cirujanos_set.add(it["cirujano_nombre"])

        # Métricas Globales (Calculadas sobre el conjunto completo de este cirujano/búsqueda)
        total_global = len(items)
        pendientes_global = len([it for it in items if not it["lio_calculado"]])
        calculados_global = len([it for it in items if it["lio_calculado"]])
        stock_pendiente_global = len([it for it in items if it["lio_calculado"] and not it["lio_stock_reservado"]])

        # Filtrado opcional por estado de cálculo
        items_filtrados = items
        if estado_calculo == "pendientes":
            items_filtrados = [it for it in items if not it["lio_calculado"]]
        elif estado_calculo == "calculados":
            items_filtrados = [it for it in items if it["lio_calculado"]]
        elif estado_calculo == "stock_pendiente":
            items_filtrados = [it for it in items if it["lio_calculado"] and not it["lio_stock_reservado"]]

        return {
            "success": True,
            "total": total_global,
            "pendientes_count": pendientes_global,
            "calculados_count": calculados_global,
            "stock_pendiente_count": stock_pendiente_global,
            "metricas": {
                "total": total_global,
                "pendientes": pendientes_global,
                "calculados": calculados_global,
                "stock_pendiente": stock_pendiente_global
            },
            "cirujanos": sorted(list(cirujanos_set)),
            "pacientes": items_filtrados
        }
    except Exception as e:
        logger.error(f"Error al listar pacientes para cálculo de LIO: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/calculo-lio/guardar")
def guardar_calculo_lio_endpoint(payload: GuardarCalculoLioPayload):
    """
    Guarda las opciones de LIO calculadas por el médico cirujano y sella el estado lio_calculado.
    No modifica cotizaciones ni precios del asesoramiento.
    """
    try:
        from datetime import datetime, timezone
        ahora_iso = datetime.now(timezone.utc).isoformat()
        opciones = payload.opciones or []

        # Determinar valores principales del primer lente / plan A
        modelo_ppal = None
        dioptria_ppal = None
        es_torico_ppal = False
        torico_valor_ppal = None
        torico_eje_ppal = None

        if opciones:
            ppal = next((op for op in opciones if op.get("tipo_opcion") == "principal"), opciones[0])
            modelo_ppal = ppal.get("modelo")
            dioptria_ppal = ppal.get("dioptria")
            es_torico_ppal = bool(ppal.get("es_torico"))
            torico_valor_ppal = ppal.get("torico_valor")
            torico_eje_ppal = ppal.get("torico_eje")

        es_confirmado = bool(payload.confirmar)

        upd_data = {
            "lio_calculado": es_confirmado,
            "lio_calculado_at": ahora_iso if es_confirmado else None,
            "lio_calculado_por": payload.lio_calculado_por if es_confirmado else None,
            "lio_calculo_opciones": opciones
        }

        # 1. Si hay turno_id, actualizar turnos_quirofano
        if payload.turno_id:
            turno_upd = {
                **upd_data,
                "lleva_lente": True,
                "updated_at": ahora_iso
            }
            if modelo_ppal:
                turno_upd["lente_tipo"] = modelo_ppal
            if dioptria_ppal:
                turno_upd["lente_dioptria"] = dioptria_ppal
            turno_upd["es_torico"] = es_torico_ppal
            if torico_valor_ppal is not None:
                turno_upd["lente_torico_valor"] = int(torico_valor_ppal) if str(torico_valor_ppal).isdigit() else 0
            if torico_eje_ppal is not None:
                turno_upd["lente_torico_eje"] = int(torico_eje_ppal) if str(torico_eje_ppal).isdigit() else 90

            supabase.table("turnos_quirofano").update(turno_upd).eq("id", payload.turno_id).execute()

        # 2. Si hay asesoria_id, sincronizar referencia técnica en asesorías (sin tocar presupuesto)
        if payload.asesoria_id:
            as_upd = {
                **upd_data,
                "updated_at": ahora_iso
            }
            if payload.ojo:
                as_upd["ojo"] = payload.ojo
            supabase.table("asesorias_quirurgicas").update(as_upd).eq("id", payload.asesoria_id).execute()

        accion_log = "CALCULO_LIO_CONFIRMADO" if es_confirmado else "CALCULO_LIO_BORRADOR"
        mensaje_log = f"Cálculo de LIO {'confirmado' if es_confirmado else 'guardado como borrador'} ({len(opciones)} opciones) por {payload.lio_calculado_por}"

        log_event(
            nivel="INFO",
            modulo="QUIROFANO",
            accion=accion_log,
            mensaje=mensaje_log,
            detalles={"turno_id": payload.turno_id, "asesoria_id": payload.asesoria_id, "opciones_count": len(opciones), "confirmado": es_confirmado}
        )

        return {
            "success": True,
            "mensaje": "Cálculo de LIO confirmado y sellado exitosamente." if es_confirmado else "Borrador de cálculo de LIO guardado.",
            "lio_calculado": es_confirmado,
            "lio_calculado_at": ahora_iso if es_confirmado else None,
            "opciones": opciones
        }
    except Exception as e:
        logger.error(f"Error al guardar cálculo de LIO: {e}")
        raise HTTPException(status_code=500, detail=str(e))

class ReabrirCalculoPayload(BaseModel):
    turno_id: Optional[str] = None
    asesoria_id: Optional[str] = None
    usuario: Optional[str] = None

@app.post("/api/calculo-lio/reabrir")
def reabrir_calculo_lio_endpoint(payload: ReabrirCalculoPayload):
    """
    Reabre un cálculo de LIO confirmado para permitir su rectificación o ajuste médico.
    """
    try:
        from datetime import datetime, timezone
        ahora_iso = datetime.now(timezone.utc).isoformat()
        
        upd = {
            "lio_calculado": False,
            "updated_at": ahora_iso
        }
        
        if payload.turno_id:
            supabase.table("turnos_quirofano").update(upd).eq("id", payload.turno_id).execute()
        if payload.asesoria_id:
            supabase.table("asesorias_quirurgicas").update(upd).eq("id", payload.asesoria_id).execute()
            
        log_event(
            nivel="INFO",
            modulo="QUIROFANO",
            accion="CALCULO_LIO_REABIERTO",
            mensaje=f"Cálculo de LIO reabierto para edición por {payload.usuario or 'Cirujano'}",
            detalles={"turno_id": payload.turno_id, "asesoria_id": payload.asesoria_id}
        )
        return {"success": True, "mensaje": "Cálculo reabierto para edición.", "lio_calculado": False}
    except Exception as e:
        logger.error(f"Error al reabrir cálculo de LIO: {e}")
        raise HTTPException(status_code=500, detail=str(e))

class ReservarStockPayload(BaseModel):
    reservado: bool = True
    observaciones: Optional[str] = None
    usuario: Optional[str] = None

@app.put("/api/turnos-quirofano/{turno_id}/reservar-stock")
def reservar_stock_lio_endpoint(turno_id: str, payload: ReservarStockPayload):
    """
    Marca o desmarca la reserva física / consignación del lente intraocular en Quirófano.
    """
    try:
        from datetime import datetime, timezone
        ahora_iso = datetime.now(timezone.utc).isoformat() if payload.reservado else None

        upd = {
            "lio_stock_reservado": payload.reservado,
            "lio_stock_reservado_at": ahora_iso,
            "lio_stock_observaciones": payload.observaciones or "",
            "updated_at": "now()"
        }

        res = supabase.table("turnos_quirofano").update(upd).eq("id", turno_id).execute()
        if not res.data:
            raise HTTPException(status_code=404, detail="Turno no encontrado")

        log_event(
            nivel="INFO",
            modulo="QUIROFANO",
            accion="RESERVA_STOCK_LIO",
            mensaje=f"Stock de LIO {'reservado' if payload.reservado else 'desmarcado'} para turno #{turno_id}",
            detalles={"turno_id": turno_id, "reservado": payload.reservado, "usuario": payload.usuario}
        )

        return {"success": True, "turno": res.data[0]}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error reservando stock de LIO para turno {turno_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))
